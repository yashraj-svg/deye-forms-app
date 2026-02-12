#!/usr/bin/env python
"""Check Excel invoice rates vs our calculator."""
import openpyxl
from forms.calculator.data_loader import load_pincode_master
from forms.calculator.freight_calculator import GlobalCourierCargo, QuoteInput, ShipmentItem
from forms.calculator.config import Settings

# Load Excel with calculated values
wb = openpyxl.load_workbook('ML model for global courier.xlsx', data_only=True)
ws = wb.active

# Print headers
print("Excel Headers:")
for i in range(1, 20):
    val = ws.cell(1, i).value
    if val:
        print(f"  Col {chr(64+i)}: {val}")

print("\n" + "="*80)
print("Comparing first 5 shipments:")
print("="*80)

# Load calculator
pins = load_pincode_master('.')
settings = Settings()
calc = GlobalCourierCargo(settings)

# Check first 10 data rows
for row_num in range(2, 12):
    awb = ws.cell(row_num, 3).value  # Col C: AWB NO.
    from_pin = str(ws.cell(row_num, 5).value or '').strip()  # Col E: From Pincode
    to_pin = str(ws.cell(row_num, 6).value or '').strip()  # Col F: To Pincode
    weight = ws.cell(row_num, 10).value  # Col J: WEIGHT
    
    # Excel pricing columns
    excel_base = ws.cell(row_num, 12).value  # Col L: AMOUNT (base freight)
    excel_fuel = ws.cell(row_num, 13).value  # Col M: 10% FSC
    excel_without_gst = ws.cell(row_num, 14).value  # Col N: Total Without GST
    excel_gst = ws.cell(row_num, 15).value  # Col O: GST
    excel_total = ws.cell(row_num, 16).value  # Col P: Subtotal (final)
    
    if not from_pin or not to_pin or not weight:
        continue
    
    # Calculate using our logic
    items = [ShipmentItem(weight_kg=float(weight), length_cm=0, breadth_cm=0, height_cm=0)]
    result = calc.calculate_quote(QuoteInput(from_pincode=from_pin, to_pincode=to_pin, items=items), pins)
    
    # Extract surcharges
    oda = result.surcharges.get('oda', 0)
    fuel = result.surcharges.get('fuel_surcharge', 0)
    
    print(f"\nRow {row_num}: AWB={awb} {from_pin}→{to_pin}, {weight}kg")
    print(f"  Excel:  Base={excel_base} Fuel={excel_fuel} GST={excel_gst} Total={excel_total}")
    print(f"  Calc:   Base={result.base_freight:.2f} ODA={oda:.2f} Fuel={fuel:.2f} GST={result.gst_amount:.2f} Total={result.total_after_gst:.2f}")
    
    # Calculate what Excel's base should be (base+ODA if our calc has ODA)
    excel_base_with_oda = result.base_freight + oda
    
    if excel_total:
        diff = result.total_after_gst - float(excel_total)
        match = "✓ OK" if abs(diff) < 1 else "✗ MISMATCH"
        print(f"  Diff:   {diff:+.2f} {match}")
        
        # Check if Excel base includes ODA
        if oda > 0:
            if excel_base and abs(float(excel_base) - excel_base_with_oda) < 1:
                print(f"  Note:   Excel base includes ODA (₹{excel_base} ≈ {excel_base_with_oda:.2f})")
