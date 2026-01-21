import csv
import os
from django.core.management.base import BaseCommand
from forms.models import PincodeData


class Command(BaseCommand):
    help = 'Import partner pincode data from CSV/Excel files into database'

    def handle(self, *args, **options):
        self.stdout.write('Starting partner data import...')
        
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
        
        # Import Rahul Delhivery CSV
        self.import_rahul_csv(base_dir)
        
        # Import Safexpress whitelist
        self.import_safexpress_whitelist(base_dir)
        
        # Assign partner regions
        self.assign_regions()
        
        count = PincodeData.objects.count()
        self.stdout.write(self.style.SUCCESS(f'✓ Import complete! Total pincodes: {count}'))

    def import_rahul_csv(self, base_dir):
        csv_path = os.path.join(
            base_dir,
            'Pickup & Drop Partner Charges',
            'Rahul Delhivery',
            'B2B_Pincode_List_globalcouriercargodc b2br_2025-12-08.csv'
        )
        
        if not os.path.exists(csv_path):
            self.stdout.write(self.style.WARNING(f'Rahul CSV not found: {csv_path}'))
            return
        
        self.stdout.write('Importing Rahul Delhivery data...')
        imported = 0
        
        with open(csv_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            batch = []
            
            for row in reader:
                pin = row.get('Pin', '').strip()
                if not pin or len(pin) != 6 or not pin.isdigit():
                    continue
                
                oda_raw = row.get('ODA', '').strip()
                is_oda = None
                deliverable = bool(oda_raw)  # Deliverable if ODA field has any value
                
                if oda_raw.upper() == 'TRUE':
                    is_oda = True
                elif oda_raw.upper() == 'FALSE':
                    is_oda = False
                
                batch.append(PincodeData(
                    pincode=pin,
                    city=row.get('Facility City', '').strip() or None,
                    state=row.get('Facility State', '').strip() or None,
                    is_oda=is_oda,
                    deliverable=deliverable,
                ))
                
                if len(batch) >= 1000:
                    PincodeData.objects.bulk_create(batch, ignore_conflicts=True)
                    imported += len(batch)
                    batch = []
                    self.stdout.write(f'  Imported {imported} pincodes...')
            
            if batch:
                PincodeData.objects.bulk_create(batch, ignore_conflicts=True)
                imported += len(batch)
        
        self.stdout.write(self.style.SUCCESS(f'✓ Rahul data: {imported} pincodes'))

    def import_safexpress_whitelist(self, base_dir):
        xlsx_path = os.path.join(
            base_dir,
            'Pickup & Drop Partner Charges',
            'Safexpress',
            '10421 SCHEDULE PINCODE AS ON 1 OCTOBER 2025 (1).xlsx'
        )
        
        if not os.path.exists(xlsx_path):
            self.stdout.write(self.style.WARNING(f'Safexpress Excel not found: {xlsx_path}'))
            return
        
        self.stdout.write('Importing Safexpress whitelist...')
        
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.stdout.write(self.style.WARNING('openpyxl not installed, skipping Safexpress import'))
            return
        
        wb = load_workbook(filename=xlsx_path, read_only=True)
        ws = wb.active
        
        pins = set()
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                pin = str(cell).strip()
                if len(pin) == 6 and pin.isdigit():
                    pins.add(pin)
        
        # Update existing records
        updated = PincodeData.objects.filter(pincode__in=pins).update(safexpress_is_oda=False)
        
        self.stdout.write(self.style.SUCCESS(f'✓ Safexpress whitelist: {updated} pincodes marked as non-ODA'))

    def assign_regions(self):
        self.stdout.write('Assigning partner regions...')
        
        # Import region mappings
        from forms.calculator.data_loader import _RAHUL_REGION_BY_CITY, _RAHUL_REGION_BY_STATE, _BLUEDART_REGION_BY_STATE
        
        updated = 0
        batch_updates = []
        
        for pincode_data in PincodeData.objects.all().iterator(chunk_size=1000):
            city_key = (pincode_data.city or '').strip().lower()
            state_key = (pincode_data.state or '').strip().lower()
            changed = False
            
            # Assign Rahul region (city priority, then state)
            if not pincode_data.global_cargo_region:
                pincode_data.global_cargo_region = _RAHUL_REGION_BY_CITY.get(city_key) or _RAHUL_REGION_BY_STATE.get(state_key)
                changed = True
            
            # Assign Bluedart region
            if not pincode_data.bluedart_region and state_key:
                pincode_data.bluedart_region = _BLUEDART_REGION_BY_STATE.get(state_key)
                changed = True
            
            if changed:
                batch_updates.append(pincode_data)
                updated += 1
            
            if len(batch_updates) >= 1000:
                PincodeData.objects.bulk_update(batch_updates, ['global_cargo_region', 'bluedart_region'])
                batch_updates = []
                self.stdout.write(f'  Assigned regions for {updated} pincodes...')
        
        if batch_updates:
            PincodeData.objects.bulk_update(batch_updates, ['global_cargo_region', 'bluedart_region'])
        
        self.stdout.write(self.style.SUCCESS(f'✓ Regions assigned: {updated} pincodes'))
