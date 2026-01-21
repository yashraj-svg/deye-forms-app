"""
Management command to import stock data from Excel file with year-wise sheets
Usage: python manage.py import_stock_data
"""
from django.core.management.base import BaseCommand
from forms.models import StockItem
import openpyxl
import os
from decimal import Decimal, InvalidOperation
from collections import defaultdict


class Command(BaseCommand):
    help = 'Import stock inventory data from Untitled spreadsheet (1).xlsx with year-wise sheets'

    def handle(self, *args, **options):
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
        excel_path = os.path.join(base_dir, 'Untitled spreadsheet (1).xlsx')
        
        if not os.path.exists(excel_path):
            self.stdout.write(self.style.ERROR(f'Excel file not found: {excel_path}'))
            return
        
        self.stdout.write(self.style.SUCCESS('Starting stock data import from Excel...'))
        
        # Clear existing data
        deleted_count = StockItem.objects.all().count()
        StockItem.objects.all().delete()
        self.stdout.write(f'Deleted {deleted_count} existing stock items')
        
        # Dictionary to aggregate quantities by (pcba_sn_new, year)
        aggregated_data = defaultdict(lambda: {
            'pcba_sn_old': '',
            'component_type': '',
            'specification': '',
            'quantity': Decimal('0'),
            'remark': '',
        })
        
        # Load workbook
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        total_rows_processed = 0
        
        # Process each sheet (year)
        for sheet_name in wb.sheetnames:
            try:
                year = int(sheet_name)
            except ValueError:
                self.stdout.write(f'Skipping non-numeric sheet: {sheet_name}')
                continue
            
            ws = wb[sheet_name]
            self.stdout.write(f'\nProcessing year {year}...')
            
            # Read header (first row)
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            header = [str(h).strip().replace('\xa0', ' ') if h else '' for h in header_row]
            
            rows_in_sheet = 0
            
            # Process data rows
            for row_values in ws.iter_rows(min_row=2, values_only=True):
                # Skip empty rows
                if not row_values or all(v is None or str(v).strip() == '' for v in row_values):
                    continue
                
                # Create row dict
                row = {}
                for i, header_name in enumerate(header):
                    if i < len(row_values):
                        row[header_name] = row_values[i]
                
                # Get PCBA SN (new) - fallback to old SN if new is missing
                pcba_new = row.get('PCBA SN* (new)', None)
                pcba_old = row.get('PCBA SN* (old)', None)
                
                if pcba_new:
                    pcba_new = str(pcba_new).strip()
                if pcba_old:
                    pcba_old = str(pcba_old).strip()
                
                # Use new SN if available, otherwise use old SN
                if not pcba_new and not pcba_old:
                    continue
                
                if not pcba_new:
                    pcba_new = pcba_old  # Fallback to old SN for legacy data
                
                # Parse quantity
                try:
                    qty_value = row.get('Quantity*(PCS)', 0)
                    if qty_value is None:
                        qty_value = 0
                    quantity = Decimal(str(qty_value))
                except (InvalidOperation, ValueError, TypeError):
                    quantity = Decimal('0')
                
                # Create key for aggregation
                key = (pcba_new, year)
                
                # Aggregate data - store old SN if not already set
                if not aggregated_data[key]['pcba_sn_old']:
                    aggregated_data[key]['pcba_sn_old'] = pcba_old or ''
                aggregated_data[key]['component_type'] = str(row.get('Component type', '') or '')
                aggregated_data[key]['specification'] = str(row.get('Specification', '') or '')
                aggregated_data[key]['quantity'] += quantity
                
                # Concatenate remarks if multiple entries
                remark = str(row.get('Remark', '') or '')
                if remark and remark not in aggregated_data[key]['remark']:
                    if aggregated_data[key]['remark']:
                        aggregated_data[key]['remark'] += '; ' + remark
                    else:
                        aggregated_data[key]['remark'] = remark
                
                rows_in_sheet += 1
                total_rows_processed += 1
            
            self.stdout.write(f'  Processed {rows_in_sheet} rows from {year}')
        
        # Create stock items from aggregated data
        self.stdout.write('\nCreating stock items...')
        imported = 0
        batch = []
        batch_size = 500
        
        for (pcba_new, year), data in aggregated_data.items():
            stock_item = StockItem(
                pcba_sn_old=data['pcba_sn_old'],
                pcba_sn_new=pcba_new,
                component_type=data['component_type'],
                specification=data['specification'],
                quantity=data['quantity'],
                remark=data['remark'],
                year=year
            )
            
            batch.append(stock_item)
            
            # Bulk create in batches
            if len(batch) >= batch_size:
                StockItem.objects.bulk_create(batch)
                imported += len(batch)
                self.stdout.write(f'  Imported {imported} items...')
                batch = []
        
        # Import remaining items
        if batch:
            StockItem.objects.bulk_create(batch)
            imported += len(batch)
        
        self.stdout.write(
            self.style.SUCCESS(
                f'\n✓ Import complete!\n'
                f'  Total rows processed: {total_rows_processed}\n'
                f'  Unique items imported: {imported} (after merging duplicates)'
            )
        )
        
        # Show year breakdown
        from django.db.models import Sum, Count
        years = StockItem.objects.values('year').annotate(
            count=Count('id'),
            total_qty=Sum('quantity')
        ).order_by('year')
        
        self.stdout.write('\nYear breakdown:')
        for y in years:
            self.stdout.write(f"  {y['year']}: {y['count']} items, {y['total_qty']} total quantity")
