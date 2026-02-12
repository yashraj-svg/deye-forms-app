#!/usr/bin/env python
"""
Management command to load Bigship ODA pincodes from Excel file into database
Usage: python manage.py load_bigship_pincodes
"""

from django.core.management.base import BaseCommand
from pathlib import Path
import openpyxl
from forms.models import PincodeData


class Command(BaseCommand):
    help = 'Load Bigship ODA pincodes from Bigship Serviceable Pincode.xlsx into database'

    def handle(self, *args, **options):
        # Try multiple possible paths for the Excel file
        possible_paths = [
            Path(__file__).resolve().parents[4] / "Bigship Serviceable Pincode.xlsx",
            Path.cwd() / "Bigship Serviceable Pincode.xlsx",
            Path(__file__).resolve().parents[2] / "Bigship Serviceable Pincode.xlsx",
        ]
        
        excel_file = None
        for path in possible_paths:
            if path.exists():
                excel_file = path
                break
        
        if not excel_file:
            self.stdout.write(self.style.ERROR(f"❌ Excel file not found in any of the expected locations:"))
            for path in possible_paths:
                self.stdout.write(f"  - {path}")
            return
        
        self.stdout.write(self.style.SUCCESS(f"📂 Loading from: {excel_file}"))
        
        try:
            wb = openpyxl.load_workbook(str(excel_file))
            ws = wb.active
            
            loaded_count = 0
            oda_count = 0
            updated_count = 0
            
            for row_idx in range(2, ws.max_row + 1):
                pincode_val = str(ws.cell(row_idx, 1).value or "").strip()
                city = str(ws.cell(row_idx, 2).value or "").strip()
                state = str(ws.cell(row_idx, 3).value or "").strip()
                is_oda_val = ws.cell(row_idx, 4).value
                
                if not pincode_val:
                    continue
                
                # Determine if ODA
                is_oda = bool(is_oda_val) if is_oda_val else None
                if isinstance(is_oda_val, str) and is_oda_val.lower() in ['true', 'yes', 'y', '1']:
                    is_oda = True
                elif isinstance(is_oda_val, str) and is_oda_val.lower() in ['false', 'no', 'n', '0']:
                    is_oda = False
                
                # Update or create pincode in database
                pincode_rec, created = PincodeData.objects.get_or_create(
                    pincode=pincode_val,
                    defaults={
                        'city': city,
                        'state': state,
                        'bigship_is_oda': is_oda,
                        'deliverable': True,
                    }
                )
                
                # Update if existing
                if not created:
                    old_oda = pincode_rec.bigship_is_oda
                    if old_oda != is_oda:
                        pincode_rec.bigship_is_oda = is_oda
                        pincode_rec.save()
                        updated_count += 1
                
                loaded_count += 1
                if is_oda:
                    oda_count += 1
            
            self.stdout.write(self.style.SUCCESS(f"✅ Successfully loaded {loaded_count} Bigship serviceable pincodes"))
            self.stdout.write(self.style.SUCCESS(f"🚨 ODA pincodes: {oda_count}"))
            self.stdout.write(self.style.SUCCESS(f"🔄 Updated existing records: {updated_count}"))
            
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"❌ Error loading pincodes: {e}"))
            import traceback
            traceback.print_exc()
