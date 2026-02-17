"""
Management command to update stock items from 'Current stock updated.xlsx'
This replaces all existing stock data with data from the Excel file.
"""

from django.core.management.base import BaseCommand
from django.db import transaction
from forms.models import StockItem
import openpyxl
from pathlib import Path
from datetime import date


class Command(BaseCommand):
    help = 'Update stock items from Current stock updated.xlsx file'

    def add_arguments(self, parser):
        parser.add_argument(
            '--confirm',
            action='store_true',
            help='Confirm that you want to replace all existing stock data',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Preview the changes without actually updating the database',
        )

    def handle(self, *args, **options):
        confirm = options.get('confirm')
        dry_run = options.get('dry_run')

        # Find the Excel file
        base_dir = Path(__file__).resolve().parent.parent.parent.parent
        excel_file = base_dir / 'Current stock updated.xlsx'

        if not excel_file.exists():
            self.stdout.write(self.style.ERROR(f'❌ Excel file not found: {excel_file}'))
            return

        self.stdout.write(self.style.SUCCESS(f'📁 Found Excel file: {excel_file}'))

        # Load the Excel file
        try:
            wb = openpyxl.load_workbook(str(excel_file), data_only=True)
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'❌ Error loading Excel file: {e}'))
            return

        # Check if "Current Stock" sheet exists
        if 'Current Stock' not in wb.sheetnames:
            self.stdout.write(self.style.ERROR(f'❌ Sheet "Current Stock" not found. Available sheets: {wb.sheetnames}'))
            wb.close()
            return

        ws = wb['Current Stock']
        self.stdout.write(self.style.SUCCESS(f'✅ Found "Current Stock" sheet with {ws.max_row} rows'))

        # Parse the data
        items_to_create = []
        skipped_count = 0
        total_balance = 0

        self.stdout.write('\n📊 Parsing Excel data...\n')

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if not row or all(cell is None for cell in row):
                continue

            # Extract data from columns
            sr_no = row[0] if len(row) > 0 else None
            serial_number = row[1] if len(row) > 1 else None
            component_type = row[2] if len(row) > 2 else None
            description = row[3] if len(row) > 3 else None
            balance = row[4] if len(row) > 4 else 0

            # Convert serial number to string and clean it
            if serial_number is not None:
                serial_number = str(serial_number).strip()
            else:
                serial_number = f"UNKNOWN_{row_num}"

            # Clean component type and description
            if component_type:
                component_type = str(component_type).strip()
            else:
                component_type = "Unknown"

            if description:
                description = str(description).strip()
            else:
                description = ""

            # Convert balance to number
            try:
                balance = int(balance) if balance else 0
            except (ValueError, TypeError):
                balance = 0

            # Skip items with 0 or negative balance if needed (optional)
            # if balance <= 0:
            #     skipped_count += 1
            #     continue

            total_balance += balance

            # Create StockItem object
            stock_item = StockItem(
                pcba_sn_new=serial_number,
                component_type=component_type,
                specification=description,
                quantity=balance,
                year=2026,  # Current year
                shipment_date=date.today(),  # Today's date
                remark=f"Imported from Current stock updated.xlsx on {date.today()}"
            )

            items_to_create.append(stock_item)

        # Also import from "Serial Number " sheet (items without serial numbers)
        if 'Serial Number ' in wb.sheetnames:
            self.stdout.write(self.style.SUCCESS(f'\n✅ Found "Serial Number " sheet, importing items without serial numbers...'))
            ws_serial = wb['Serial Number ']
            
            for row_num, row in enumerate(ws_serial.iter_rows(min_row=2, values_only=True), start=2):
                # Skip empty rows
                if not row or all(cell is None for cell in row):
                    continue

                # Extract data from columns (Serial Number, Component's, Description, BALANCE)
                serial_number = row[0] if len(row) > 0 else None
                component_type = row[1] if len(row) > 1 else None
                description = row[2] if len(row) > 2 else None
                balance = row[3] if len(row) > 3 else 0

                # Use the serial number from the sheet (Unknown 1, Unknown 2, etc.)
                if serial_number:
                    serial_number = str(serial_number).strip()
                else:
                    serial_number = f"UNKNOWN_{row_num}"

                # Clean component type and description
                if component_type:
                    component_type = str(component_type).strip()
                else:
                    component_type = "Unknown"

                if description:
                    description = str(description).strip()
                else:
                    description = ""

                # Convert balance to number
                try:
                    balance = int(balance) if balance else 0
                except (ValueError, TypeError):
                    balance = 0

                total_balance += balance

                # Create StockItem object
                stock_item = StockItem(
                    pcba_sn_new=serial_number,
                    component_type=component_type,
                    specification=description,
                    quantity=balance,
                    year=2026,
                    shipment_date=date.today(),
                    remark=f"Imported from Current stock updated.xlsx (Serial Number sheet) on {date.today()}"
                )

                items_to_create.append(stock_item)

        wb.close()

        # Display summary
        self.stdout.write(f'\n{"="*80}')
        self.stdout.write(self.style.WARNING('SUMMARY'))
        self.stdout.write(f'{"="*80}')
        self.stdout.write(f'📦 Total items to import: {len(items_to_create)}')
        self.stdout.write(f'⏭️  Skipped items: {skipped_count}')
        self.stdout.write(f'📊 Total quantity: {total_balance:,} units')
        self.stdout.write(f'{"="*80}\n')

        # Show first 5 items as preview
        self.stdout.write(self.style.WARNING('Preview (first 5 items):'))
        for i, item in enumerate(items_to_create[:5], 1):
            self.stdout.write(f'{i}. {item.pcba_sn_new} | {item.component_type} | {item.specification[:50]} | Qty: {item.quantity}')

        self.stdout.write(f'\n...(and {len(items_to_create) - 5} more items)\n')

        # Check if dry run
        if dry_run:
            self.stdout.write(self.style.WARNING('🔍 DRY RUN - No changes made to database'))
            return

        # Check confirmation
        if not confirm:
            self.stdout.write(self.style.ERROR('\n⚠️  This will REPLACE all existing stock data!'))
            self.stdout.write(self.style.WARNING(f'Current stock items in database: {StockItem.objects.count()}'))
            self.stdout.write(self.style.WARNING('\nTo proceed, run:'))
            self.stdout.write(self.style.SUCCESS('python manage.py update_current_stock --confirm'))
            return

        # Perform the update
        self.stdout.write(self.style.WARNING('\n⚠️  Starting database update...'))

        try:
            with transaction.atomic():
                # Delete all existing stock items
                old_count = StockItem.objects.count()
                self.stdout.write(f'🗑️  Deleting {old_count} existing stock items...')
                StockItem.objects.all().delete()

                # Create new stock items
                self.stdout.write(f'➕ Creating {len(items_to_create)} new stock items...')
                StockItem.objects.bulk_create(items_to_create, batch_size=100)

                new_count = StockItem.objects.count()

                self.stdout.write(f'\n{"="*80}')
                self.stdout.write(self.style.SUCCESS('✅ DATABASE UPDATE COMPLETE!'))
                self.stdout.write(f'{"="*80}')
                self.stdout.write(f'🗑️  Deleted: {old_count} items')
                self.stdout.write(f'➕ Created: {new_count} items')
                self.stdout.write(f'📊 Total quantity: {total_balance:,} units')
                self.stdout.write(f'{"="*80}\n')

        except Exception as e:
            self.stdout.write(self.style.ERROR(f'\n❌ Error updating database: {e}'))
            self.stdout.write(self.style.ERROR('Transaction rolled back - no changes made'))
            return
