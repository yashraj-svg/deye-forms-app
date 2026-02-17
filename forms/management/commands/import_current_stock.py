"""
Management command to import stock data from Current stock updated.xlsx
Replaces all existing stock data with data from "Current Stock" sheet only
"""

from django.core.management.base import BaseCommand
from django.db import transaction
from forms.models import StockItem
import openpyxl
from pathlib import Path
from decimal import Decimal


class Command(BaseCommand):
    help = 'Import current stock data from Excel file (Current Stock sheet only)'

    def add_arguments(self, parser):
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Preview import without making changes',
        )

    def handle(self, *args, **options):
        dry_run = options['dry_run']
        
        # Locate the Excel file
        excel_file = Path(__file__).resolve().parent.parent.parent.parent / 'Current stock updated.xlsx'
        
        if not excel_file.exists():
            self.stdout.write(self.style.ERROR(f'❌ Excel file not found: {excel_file}'))
            return
        
        self.stdout.write(self.style.SUCCESS(f'📂 Found Excel file: {excel_file}'))
        
        # Load workbook
        wb = openpyxl.load_workbook(str(excel_file), data_only=True)
        
        if dry_run:
            self.stdout.write(self.style.WARNING('🔍 DRY RUN MODE - No changes will be made'))
        
        def import_from_sheet(workbook, sheet_name, col_serial, col_component, col_desc, col_balance):
            """Import stock items from a specific sheet"""
            if sheet_name not in workbook.sheetnames:
                self.stdout.write(self.style.ERROR(f'❌ Sheet "{sheet_name}" not found'))
                return 0, 0
            
            ws = workbook[sheet_name]
            items_to_create = []
            total_qty = 0
            skipped = 0
            
            self.stdout.write(f'\n📊 Processing sheet: {sheet_name}')
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or len(row) < 2:
                    continue
                
                # Get values based on column mapping
                serial_number = row[ord(col_serial) - ord('A')] if col_serial else None
                component_type = row[ord(col_component) - ord('A')] if col_component else None
                description = row[ord(col_desc) - ord('A')] if col_desc else None
                balance = row[ord(col_balance) - ord('A')] if col_balance else None
                
                # Use component type as serial number if serial is empty
                if not serial_number or str(serial_number).strip() == '':
                    if component_type:
                        serial_number = f"NO_SN_{component_type}_{row_idx}"
                    else:
                        serial_number = f"NO_SN_{row_idx}"
                
                # Skip only if balance is None (not zero - we want to track zero stock items)
                if balance is None:
                    skipped += 1
                    continue
                
                # Handle component_type being None
                if component_type is None:
                    component_type = ""
                
                # Handle description being None
                if description is None:
                    description = ""
                
                try:
                    quantity = Decimal(str(balance))
                    total_qty += float(quantity)
                    
                    items_to_create.append(StockItem(
                        pcba_sn_new=str(serial_number).strip(),
                        component_type=str(component_type).strip(),
                        specification=str(description).strip(),
                        quantity=quantity,
                        year=2025,
                        shipment_date=None,
                    ))
                    
                except (ValueError, TypeError) as e:
                    self.stdout.write(self.style.WARNING(f'⚠️  Row {row_idx}: Invalid balance "{balance}" - {e}'))
                    skipped += 1
                    continue
            
            self.stdout.write(f'   ✓ Found {len(items_to_create)} valid items')
            self.stdout.write(f'   ✓ Total quantity: {total_qty:,.0f} units')
            if skipped > 0:
                self.stdout.write(f'   ⚠️  Skipped {skipped} rows')
            
            return items_to_create, total_qty
        
        # Import from "Current Stock" sheet only
        items_to_create, total_quantity = import_from_sheet(wb, "Current Stock", 'B', 'C', 'D', 'E')
        
        wb.close()
        
        if not items_to_create:
            self.stdout.write(self.style.ERROR('❌ No items to import'))
            return
        
        # Summary
        self.stdout.write(self.style.SUCCESS(f'\n📦 IMPORT SUMMARY:'))
        self.stdout.write(f'   Items to import: {len(items_to_create)}')
        self.stdout.write(f'   Total quantity: {total_quantity:,.0f} units')
        
        if dry_run:
            self.stdout.write(self.style.WARNING('\n✅ Dry run complete - no changes made'))
            return
        
        # Execute import in transaction
        with transaction.atomic():
            # Delete all existing stock items
            deleted_count = StockItem.objects.all().delete()[0]
            self.stdout.write(self.style.WARNING(f'\n🗑️  Deleted {deleted_count} old stock items'))
            
            # Bulk create new items
            StockItem.objects.bulk_create(items_to_create, batch_size=500)
            self.stdout.write(self.style.SUCCESS(f'✅ Created {len(items_to_create)} new stock items'))
        
        # Verify
        final_count = StockItem.objects.count()
        self.stdout.write(self.style.SUCCESS(f'\n✅ IMPORT COMPLETE!'))
        self.stdout.write(f'   Database now has: {final_count} items')
        self.stdout.write(f'   Total quantity: {total_quantity:,.0f} units')
