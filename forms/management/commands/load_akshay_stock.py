"""
Load stock items from Akshay India Shipping List Excel file
This command reads the Excel file and loads all data into StockItem model,
preserving exact serial numbers, component names, descriptions, and quantities.
"""

import openpyxl
from django.core.management.base import BaseCommand
from forms.models import StockItem
from django.db import transaction


class Command(BaseCommand):
    help = 'Load stock items from Akshay India Shipping List 2020 - 2025.xlsx'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default='Akshay India Shipping List 2020 - 2025.xlsx',
            help='Path to the Excel file'
        )
        parser.add_argument(
            '--clear',
            action='store_true',
            help='Clear existing stock items before loading'
        )

    def handle(self, *args, **options):
        file_path = options['file']
        clear_existing = options['clear']

        self.stdout.write(self.style.WARNING(f'\n📂 Loading stock data from: {file_path}'))

        try:
            # Load workbook
            wb = openpyxl.load_workbook(file_path)
            self.stdout.write(self.style.SUCCESS(f'✓ Opened workbook with {len(wb.sheetnames)} sheets'))
            self.stdout.write(f'   Sheets: {", ".join(wb.sheetnames)}')

            # Clear existing data if requested
            if clear_existing:
                count = StockItem.objects.count()
                StockItem.objects.all().delete()
                self.stdout.write(self.style.WARNING(f'\n🗑️  Deleted {count} existing stock items'))

            total_loaded = 0
            total_skipped = 0

            # Process each sheet (year)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Extract year from sheet name (e.g., "2025年" -> 2025)
                try:
                    year = int(''.join(filter(str.isdigit, sheet_name)))
                except ValueError:
                    year = 2025  # Default year

                self.stdout.write(f'\n📋 Processing sheet: {sheet_name} (Year: {year})')

                # Find the header row (contains "PCBA SN")
                header_row = None
                for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                    if row and any('PCBA SN' in str(cell) for cell in row if cell):
                        header_row = row_idx
                        break

                if not header_row:
                    self.stdout.write(self.style.WARNING(f'   ⚠️  Could not find header row, skipping sheet'))
                    continue

                self.stdout.write(f'   Header found at row {header_row}')

                # Get column indices
                headers = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
                col_indices = {}
                for idx, header in enumerate(headers):
                    if header:
                        header_str = str(header).lower().strip()
                        if 'pcba' in header_str and 'old' in header_str:
                            col_indices['pcba_sn_old'] = idx
                        elif 'pcba' in header_str and 'new' in header_str:
                            col_indices['pcba_sn_new'] = idx
                        elif 'component type' in header_str or 'component' in header_str:
                            col_indices['component_type'] = idx
                        elif 'specification' in header_str:
                            col_indices['specification'] = idx
                        elif 'quantity' in header_str or 'pcs' in header_str:
                            col_indices['quantity'] = idx
                        elif 'remark' in header_str:
                            col_indices['remark'] = idx

                if 'pcba_sn_new' not in col_indices:
                    self.stdout.write(self.style.ERROR(f'   ❌ Could not find PCBA SN (new) column'))
                    continue

                self.stdout.write(f'   Column mapping: {col_indices}')

                # Process data rows
                batch = []
                batch_size = 500
                row_count = 0
                skipped = 0

                for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                    # Get PCBA SN (new)
                    pcba_sn_new = row[col_indices['pcba_sn_new']] if 'pcba_sn_new' in col_indices else None
                    
                    # Convert to string and clean
                    if pcba_sn_new:
                        pcba_sn_new = str(pcba_sn_new).strip()
                        # Handle scientific notation for serial numbers
                        if '.' in pcba_sn_new and 'e' not in pcba_sn_new.lower():
                            pcba_sn_new = pcba_sn_new.split('.')[0]
                        # Skip if empty after cleaning
                        if pcba_sn_new == '' or pcba_sn_new.lower() == 'none':
                            pcba_sn_new = None
                    else:
                        pcba_sn_new = None

                    # Get other fields
                    pcba_sn_old = row[col_indices['pcba_sn_old']] if 'pcba_sn_old' in col_indices else None
                    if pcba_sn_old:
                        pcba_sn_old = str(pcba_sn_old).strip()
                        if '.' in pcba_sn_old and 'e' not in pcba_sn_old.lower():
                            pcba_sn_old = pcba_sn_old.split('.')[0]
                        if pcba_sn_old == '' or pcba_sn_old.lower() == 'none':
                            pcba_sn_old = None

                    component_type = row[col_indices['component_type']] if 'component_type' in col_indices else None
                    if component_type:
                        component_type = str(component_type).strip()
                        if component_type == '' or component_type == 'None':
                            component_type = None

                    specification = row[col_indices['specification']] if 'specification' in col_indices else None
                    if specification:
                        specification = str(specification).strip()
                        if specification == '' or specification == 'None':
                            specification = None

                    # Get quantity
                    quantity_raw = row[col_indices['quantity']] if 'quantity' in col_indices else None
                    try:
                        quantity = float(quantity_raw) if quantity_raw else 0
                    except (ValueError, TypeError):
                        quantity = 0

                    remark = row[col_indices['remark']] if 'remark' in col_indices else None
                    if remark:
                        remark = str(remark).strip()
                        if remark == '' or remark == 'None':
                            remark = None

                    # Skip if both serial number AND (component_type or specification) are missing
                    # Allow items with blank serial if they have component/spec info
                    if not pcba_sn_new and not component_type and not specification:
                        skipped += 1
                        continue
                    
                    # If no serial but has component/spec, use component+spec as identifier
                    if not pcba_sn_new and (component_type or specification):
                        # This is allowed - items like tools, fixtures, etc with no serial
                        pass
                    elif not pcba_sn_new:
                        # No serial and no useful info
                        skipped += 1
                        continue

                    # Create StockItem
                    stock_item = StockItem(
                        pcba_sn_old=pcba_sn_old or None,
                        pcba_sn_new=pcba_sn_new or None,  # Allow NULL for non-serialized items
                        component_type=component_type or None,
                        specification=specification or None,
                        quantity=quantity,
                        remark=remark or None,
                        year=year
                    )
                    batch.append(stock_item)
                    row_count += 1

                    # Batch insert
                    if len(batch) >= batch_size:
                        with transaction.atomic():
                            StockItem.objects.bulk_create(batch, ignore_conflicts=False)
                        self.stdout.write(self.style.SUCCESS(f'      ✓ Inserted {len(batch)} items'))
                        total_loaded += len(batch)
                        batch = []

                # Insert remaining items
                if batch:
                    with transaction.atomic():
                        StockItem.objects.bulk_create(batch, ignore_conflicts=False)
                    self.stdout.write(self.style.SUCCESS(f'      ✓ Inserted {len(batch)} items'))
                    total_loaded += len(batch)

                total_skipped += skipped
                self.stdout.write(self.style.SUCCESS(f'   ✅ Loaded {row_count} items from {sheet_name} (skipped {skipped} empty rows)'))

            self.stdout.write(self.style.SUCCESS(f'\n\n✅ COMPLETE: Loaded {total_loaded} stock items from {len(wb.sheetnames)} sheets'))
            if total_skipped > 0:
                self.stdout.write(self.style.WARNING(f'   ⚠️  Skipped {total_skipped} empty/invalid rows'))

            # Show summary
            self.stdout.write(f'\n📊 Database Summary:')
            self.stdout.write(f'   Total items: {StockItem.objects.count()}')
            self.stdout.write(f'   Years: {list(StockItem.objects.values_list("year", flat=True).distinct().order_by("year"))}')
            self.stdout.write(f'   Component types: {StockItem.objects.values("component_type").distinct().count()}')

        except FileNotFoundError:
            self.stdout.write(self.style.ERROR(f'❌ File not found: {file_path}'))
            self.stdout.write(f'   Make sure the file is in the project root directory')
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'❌ Error loading stock data: {str(e)}'))
            import traceback
            traceback.print_exc()
