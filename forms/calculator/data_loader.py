from __future__ import annotations

import csv
import glob
import os
import re
from dataclasses import dataclass
from functools import lru_cache
from typing import Dict, Optional, Tuple


@dataclass
class PincodeRecord:
    pincode: str
    state: Optional[str] = None
    city: Optional[str] = None
    is_metro: Optional[bool] = None
    is_oda: Optional[bool] = None
    safexpress_is_oda: Optional[bool] = None
    deliverable: Optional[bool] = True
    # Partner-specific region/zone codes
    safexpress_zone: Optional[str] = None
    bluedart_region: Optional[str] = None
    global_cargo_region: Optional[str] = None


class PincodeDB:
    def __init__(self) -> None:
        self._db: Dict[str, PincodeRecord] = {}

    def get(self, pincode: str) -> Optional[PincodeRecord]:
        return self._db.get(str(pincode))

    def __contains__(self, pincode: str) -> bool:
        return str(pincode) in self._db

    def add(self, rec: PincodeRecord) -> None:
        self._db[str(rec.pincode)] = rec

    def size(self) -> int:
        return len(self._db)

    def records(self):
        return self._db.values()


def _safe_bool(x) -> Optional[bool]:
    if isinstance(x, bool):
        return x
    if x is None:
        return None
    s = str(x).strip().lower()
    if s in {"y", "yes", "true", "1"}:
        return True
    if s in {"n", "no", "false", "0"}:
        return False
    return None


def discover_file(patterns: Tuple[str, ...]) -> Optional[str]:
    for pat in patterns:
        matches = glob.glob(pat)
        if matches:
            return matches[0]
    return None


# -----------------------------
# Partner region helpers
# -----------------------------

_RAHUL_REGION_BY_CITY = {
    # C1 - Cities
    "raipur": "C1",
    "nagpur": "C1",
    # E1 - Cities
    "kolkata": "E1",
    "bhubaneswar": "E1",
    "bhubaneshwar": "E1",
    "patna": "E1",
    # E2 - Cities
    "jamshedpur": "E2",
    # N1 - Cities
    "new delhi": "N1",
    # Zone-based mapping for Global Cargo (18 zones)
    # AMB - Ambala zone
    "chandigarh": "AMB",
    "ambala": "AMB",
    "shimla": "AMB",
    "jammu": "AMB",
    "amritsar": "AMB",
    "ludhiana": "AMB",
    "jalandhar": "AMB",
    "dehradun": "AMB",
    "haridwar": "AMB",
    # JAI - Jaipur
    "jaipur": "JAI",
    "jodhpur": "JAI",
    "kota": "JAI",
    "udaipur": "JAI",
    # DEL - Gurgaon/Delhi
    "delhi": "DEL",
    "gurgaon": "DEL",
    "gurugram": "DEL",
    "noida": "DEL",
    "faridabad": "DEL",
    "ghaziabad": "DEL",
    "sahibabad": "DEL",
    "agra": "DEL",
    "aligarh": "DEL",
    "moradabad": "DEL",
    # AMD - Ahmedabad
    "ahmedabad": "AMD",
    "surat": "AMD",
    "vadodara": "AMD",
    "rajkot": "AMD",
    "gandhinagar": "AMD",
    # PNQ - Pune
    "pune": "PNQ",
    "pimpri": "PNQ",
    # BOM - Mumbai
    "mumbai": "BOM",
    "thane": "BOM",
    "navi mumbai": "BOM",
    "nashik": "BOM",
    # NAG - Nagpur
    "nagpur": "NAG",
    "aurangabad": "NAG",
    # IDR - Indore
    "indore": "IDR",
    "bhopal": "IDR",
    "raipur": "IDR",
    # BLR - Bangalore
    "bangalore": "BLR",
    "bengaluru": "BLR",
    "mysore": "BLR",
    "mangalore": "BLR",
    # HYD - Hyderabad
    "hyderabad": "HYD",
    "vijayawada": "HYD",
    "visakhapatnam": "HYD",
    # MAA - Chennai
    "chennai": "MAA",
    "madurai": "MAA",
    "salem": "MAA",
    # CJB - Coimbatore
    "coimbatore": "CJB",
    "kochi": "CJB",
    "cochin": "CJB",
    "trivandrum": "CJB",
    "kozhikode": "CJB",
    "kollam": "CJB",
    # BBI - Bhubaneswar
    "bhubaneswar": "BBI",
    "cuttack": "BBI",
    # LOK - Lucknow
    "lucknow": "LOK",
    "kanpur": "LOK",
    "varanasi": "LOK",
    # PAT - Patna
    "patna": "PAT",
    # NJP - Siliguri
    "siliguri": "NJP",
    "gangtok": "NJP",
    # CCU - Kolkata
    "kolkata": "CCU",
    "howrah": "CCU",
    "jamshedpur": "CCU",
    "ranchi": "CCU",
    # GAU - Guwahati
    "guwahati": "GAU",
}

_RAHUL_REGION_BY_STATE = {
    # Zone-based mapping for Global Cargo
    "chandigarh": "AMB",
    "himachal pradesh": "AMB",
    "jammu and kashmir": "AMB",
    "punjab": "AMB",
    "uttarakhand": "AMB",
    "rajasthan": "JAI",
    "delhi": "DEL",
    "haryana": "DEL",
    "gujarat": "AMD",
    "dadra and nagar haveli": "AMD",
    "daman and diu": "AMD",
    "goa": "PNQ",
    "maharashtra": "BOM",
    "madhya pradesh": "IDR",
    "chhattisgarh": "IDR",
    "karnataka": "BLR",
    "andhra pradesh": "HYD",
    "telangana": "HYD",
    "tamil nadu": "MAA",
    "puducherry": "MAA",
    "pondicherry": "MAA",
    "kerala": "CJB",
    "odisha": "BBI",
    "orissa": "BBI",
    "uttar pradesh": "LOK",
    "bihar": "PAT",
    "sikkim": "NJP",
    "west bengal": "CCU",
    "jharkhand": "CCU",
    "assam": "GAU",
    "meghalaya": "GAU",
    "tripura": "GAU",
    "arunachal pradesh": "GAU",
    "mizoram": "GAU",
    "manipur": "GAU",
    "nagaland": "GAU",
}


_BLUEDART_REGION_BY_STATE = {
    "himachal pradesh": "NORTH",
    "punjab": "NORTH",
    "haryana": "NORTH",
    "uttarakhand": "NORTH",
    "uttar pradesh": "NORTH",
    "rajasthan": "NORTH",
    "delhi": "NORTH",
    "bihar": "EAST",
    "odisha": "EAST",
    "orissa": "EAST",
    "west bengal": "EAST",
    "jharkhand": "EAST",
    "maharashtra": "WEST",
    "madhya pradesh": "WEST",
    "gujarat": "WEST",
    "chhattisgarh": "WEST",
    "goa": "WEST",
    "daman and diu": "WEST",
    "dadra and nagar haveli": "WEST",
    "karnataka": "SOUTH",
    "tamil nadu": "SOUTH",
    "kerala": "SOUTH",
    "andhra pradesh": "SOUTH",
    "telangana": "SOUTH",
    "pondicherry": "SOUTH",
    "puducherry": "SOUTH",
    "nagaland": "NE",
    "mizoram": "NE",
    "manipur": "NE",
    "meghalaya": "NE",
    "tripura": "NE",
    "arunachal pradesh": "NE",
    # If needed, treat "andaman and nicobar" separately later
    "jammu and kashmir": "J&K",
    "ladakh": "J&K",
}


def _assign_partner_regions(rec: PincodeRecord) -> None:
    state_key = (rec.state or "").strip().lower()
    city_key = (rec.city or "").strip().lower()
    # Always reassign zones from latest mapping (city takes priority over state)
    rec.global_cargo_region = _RAHUL_REGION_BY_CITY.get(city_key) or _RAHUL_REGION_BY_STATE.get(state_key)
    if state_key:
        if rec.bluedart_region is None:
            rec.bluedart_region = _BLUEDART_REGION_BY_STATE.get(state_key)


def _load_safexpress_non_oda_from_txt(base_dir: str) -> set[str]:
    """Return set of pincodes listed as non-ODA in the bundled text file."""
    path = os.path.join(
        base_dir,
        "forms",
        "calculator",
        "data",
        "safexpress_non_oda_pincodes.txt",
    )
    if not os.path.exists(path):
        return set()

    pins: set[str] = set()
    try:
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                pin = line.strip()
                if re.fullmatch(r"\d{6}", pin):
                    pins.add(pin)
    except Exception:
        return set()
    return pins


def _load_safexpress_non_oda_from_excel(base_dir: str) -> set[str]:
    """Return set of pincodes listed as non-ODA in the Safexpress schedule Excel."""
    path = os.path.join(
        base_dir,
        "Pickup & Drop Partner Charges",
        "Safexpress",
        "10421 SCHEDULE PINCODE AS ON 1 OCTOBER 2025 (1).xlsx",
    )
    if not os.path.exists(path):
        return set()
    try:
        from openpyxl import load_workbook
    except Exception:
        return set()

    pins: set[str] = set()
    try:
        wb = load_workbook(filename=path, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                pin = str(cell).strip()
                if re.fullmatch(r"\d{6}", pin):
                    pins.add(pin)
    except Exception:
        return set()
    return pins


def _load_safexpress_non_oda(base_dir: str) -> set[str]:
    """Return set of pincodes listed as non-ODA for Safexpress (txt + excel)."""
    pins = set()
    pins.update(_load_safexpress_non_oda_from_txt(base_dir))
    pins.update(_load_safexpress_non_oda_from_excel(base_dir))
    return pins


@lru_cache(maxsize=1)
def _load_global_cargo_oda_list(base_dir: str) -> Dict[str, Tuple[bool, Optional[str], Optional[str]]]:
    """
    Load official Global Courier Cargo B2B pincode list to get accurate ODA status.
    Returns dict: {pincode: (is_oda, city, state)} where is_oda=True means ODA area.
    """
    oda_map: Dict[str, Tuple[bool, Optional[str], Optional[str]]] = {}
    
    # Look for the official B2B pincode file
    patterns = (
        os.path.join(base_dir, "B2B_Pincode_List_globalcouriercargodc b2br updated.csv"),
        os.path.join(base_dir, "B2B_Pincode_List_GlobalCourierCargoDC b2br updated.csv"),
        os.path.join(base_dir, "*B2B*globalcourier*.csv"),
    )
    
    b2b_file = discover_file(patterns)
    if not b2b_file:
        return oda_map
    
    try:
        with open(b2b_file, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                pin = str(row.get("Pin", "")).strip()
                if not pin:
                    continue
                oda_str = str(row.get("ODA", "")).strip().upper()
                city = str(row.get("Facility City", "")).strip() or None
                state = str(row.get("Facility State", "")).strip() or None
                # ODA=TRUE means it's an Out of Delivery Area, ODA=FALSE means regularly deliverable
                is_oda = oda_str == "TRUE"
                oda_map[pin] = (is_oda, city, state)
    except Exception:
        pass
    
    return oda_map


def load_pincode_master(base_dir: str) -> PincodeDB:
    """
    Load pincode data from Django database (cached).
    Falls back to CSV if database is empty.
    Applies official Global Courier Cargo B2B list for accurate ODA status.
    """
    db = PincodeDB()
    
    # Load official Global Courier Cargo ODA list early
    global_cargo_oda_map = _load_global_cargo_oda_list(base_dir)
    
    # Try loading from Django database first
    try:
        from forms.models import PincodeData
        count = PincodeData.objects.count()
        
        if count > 0:
            # Load from database
            for record in PincodeData.objects.all():
                rec = PincodeRecord(
                    pincode=record.pincode,
                    city=record.city,
                    state=record.state,
                    is_oda=record.is_oda,
                    deliverable=record.deliverable,
                    safexpress_is_oda=record.safexpress_is_oda,
                    global_cargo_region=record.global_cargo_region,
                    bluedart_region=record.bluedart_region,
                )
                # Override is_oda with official Global Courier Cargo list if available
                if str(rec.pincode) in global_cargo_oda_map:
                    is_oda, city, state = global_cargo_oda_map[str(rec.pincode)]
                    rec.is_oda = is_oda
                    if city:
                        rec.city = city
                    if state:
                        rec.state = state
                # Assign partner regions if not already set in database
                _assign_partner_regions(rec)
                db.add(rec)

            # Apply Safexpress non-ODA whitelist to DB-loaded records
            safexpress_non_oda = _load_safexpress_non_oda(base_dir)
            if safexpress_non_oda:
                for rec in db.records():
                    rec.safexpress_is_oda = False if rec.pincode in safexpress_non_oda else True
            return db
    except Exception:
        pass  # Fall back to file loading
    
    # Fallback: Load from CSV files (legacy)
    rahul_csv = discover_file(
        (
            os.path.join(
                base_dir,
                "Pickup & Drop Partner Charges",
                "Rahul Delhivery",
                "*.csv",
            ),
        )
    )

    if rahul_csv:
        try:
            with open(rahul_csv, newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    pin = str(row.get("Pin") or "").strip()
                    if not re.fullmatch(r"\d{6}", pin):
                        continue

                    oda_raw = row.get("ODA")
                    is_oda = _safe_bool(oda_raw)
                    deliverable = bool(oda_raw and str(oda_raw).strip())

                    rec = db.get(pin) or PincodeRecord(pincode=pin)
                    rec.city = row.get("Facility City") or rec.city
                    rec.state = row.get("Facility State") or rec.state
                    rec.is_oda = is_oda
                    rec.deliverable = deliverable
                    _assign_partner_regions(rec)
                    db.add(rec)
        except Exception:
            pass

    safexpress_non_oda = _load_safexpress_non_oda(base_dir)
    if safexpress_non_oda:
        for rec in db.records():
            rec.safexpress_is_oda = False if rec.pincode in safexpress_non_oda else True
        for pin in safexpress_non_oda:
            rec = db.get(pin) or PincodeRecord(pincode=pin)
            rec.safexpress_is_oda = False
            _assign_partner_regions(rec)
            db.add(rec)

    # Apply official Global Courier Cargo ODA list to all records (overrides other sources)
    if global_cargo_oda_map:
        for pin, (is_oda, city, state) in global_cargo_oda_map.items():
            rec = db.get(pin) or PincodeRecord(pincode=pin)
            rec.is_oda = is_oda
            if city:
                rec.city = city
            if state:
                rec.state = state
            _assign_partner_regions(rec)
            db.add(rec)

    return db


def lookup_pincode(db: PincodeDB, pincode: str) -> Optional[PincodeRecord]:
    return db.get(str(pincode))
