#!/usr/bin/env python3
"""
Bigship Calculator - Support for CFT, LTL, MPS service types
Franchise partner rates - effective 1st Jan 2026
"""

from pathlib import Path
from typing import Dict, Optional, List
from forms.calculator.data_loader import PincodeRecord, PincodeDB
from forms.calculator.freight_calculator import (
    BaseCarrier, QuoteInput, QuoteResult, ShipmentItem, Settings, _round_money
)
import openpyxl


class BigshipPincodeDB:
    """Load and manage Bigship serviceable pincodes"""
    
    def __init__(self, excel_path: str = "Bigship Serviceable Pincode.xlsx"):
        self._serviceable = set()
        self._oda_pincodes = set()
        self._pincode_details: Dict[str, dict] = {}
        self._load_from_excel(excel_path)
    
    def _load_from_excel(self, excel_path):
        """Load serviceable pincodes from Excel file"""
        try:
            # Convert to Path object and resolve
            if isinstance(excel_path, str):
                excel_path = Path(excel_path)
            excel_path = excel_path.resolve()
            
            print(f"[Bigship] Loading pincodes from: {excel_path}")
            print(f"[Bigship] File exists: {excel_path.exists()}")
            print(f"[Bigship] Current working directory: {Path.cwd()}")
            
            # List files in project root for debugging
            project_root = excel_path.parent
            print(f"[Bigship] Project root: {project_root}")
            if project_root.exists():
                excel_files = list(project_root.glob('*.xlsx'))
                print(f"[Bigship] Excel files in project root: {[f.name for f in excel_files]}")
                all_files = list(project_root.glob('*'))
                print(f"[Bigship] Total files in project root: {len(all_files)}")
            
            if not excel_path.exists():
                print(f"[Bigship] ERROR: File not found at {excel_path}")
                print(f"[Bigship] This will result in ALL pincodes being marked as non-serviceable")
                return
            
            wb = openpyxl.load_workbook(str(excel_path))
            ws = wb.active
            
            loaded_count = 0
            for row_idx in range(2, ws.max_row + 1):
                pincode = str(ws.cell(row_idx, 1).value or "").strip()
                city = str(ws.cell(row_idx, 2).value or "").strip()
                state = str(ws.cell(row_idx, 3).value or "").strip()
                is_oda = ws.cell(row_idx, 4).value
                
                if pincode:
                    self._serviceable.add(pincode)
                    self._pincode_details[pincode] = {
                        "city": city,
                        "state": state,
                        "is_oda": is_oda
                    }
                    
                    if is_oda or str(is_oda).lower() in ["true", "yes", "y", "1"]:
                        self._oda_pincodes.add(pincode)
                    loaded_count += 1
            
            print(f"[Bigship] Successfully loaded {loaded_count} serviceable pincodes")
            print(f"[Bigship] ODA pincodes: {len(self._oda_pincodes)}")
        except Exception as e:
            print(f"[Bigship] ERROR loading pincodes: {e}")
            import traceback
            traceback.print_exc()
    
    def is_serviceable(self, pincode: str) -> bool:
        """Check if pincode is serviceable by Bigship"""
        return str(pincode).strip() in self._serviceable
    
    def is_oda(self, pincode: str) -> bool:
        """Check if pincode is ODA (Out of Delivery Area)"""
        return str(pincode).strip() in self._oda_pincodes
    
    def get_details(self, pincode: str) -> Optional[dict]:
        """Get pincode details (city, state, etc.)"""
        return self._pincode_details.get(str(pincode).strip())


class Bigship(BaseCarrier):
    """Bigship courier calculator - CFT, LTL, MPS service types)"""
    
    name = "Bigship"
    
    # Service types
    SERVICE_TYPES = {
        "CFT": "CFT - Courier Freight Transport (Light Weight)",
        "LTL": "LTL - Less Than Truckload (Heavy Shipments)",
        "MPS": "MPS - Metro Parcel Service (Non-ODA Only)"
    }
    
    # Franchise Partner Rates - CFT (Cool Food Transport)
    # Weight slabs in kg, Rate per kg
    CFT_RATES = {
        # Weight range: (min, max): rate_per_kg
        "slab_1": {"range": (0, 5), "rate": 80.0, "min_charge": 400},
        "slab_2": {"range": (5, 10), "rate": 70.0, "min_charge": 500},
        "slab_3": {"range": (10, 25), "rate": 60.0, "min_charge": 800},
        "slab_4": {"range": (25, 50), "rate": 50.0, "min_charge": 1500},
        "slab_5": {"range": (50, 100), "rate": 40.0, "min_charge": 2500},
        "slab_6": {"range": (100, 250), "rate": 35.0, "min_charge": 4000},
    }
    
    # Franchise Partner Rates - LTL (Less Than Truckload)
    LTL_RATES = {
        "slab_1": {"range": (0, 5), "rate": 50.0, "min_charge": 250},
        "slab_2": {"range": (5, 10), "rate": 45.0, "min_charge": 350},
        "slab_3": {"range": (10, 25), "rate": 40.0, "min_charge": 600},
        "slab_4": {"range": (25, 50), "rate": 35.0, "min_charge": 1000},
        "slab_5": {"range": (50, 100), "rate": 30.0, "min_charge": 2000},
        "slab_6": {"range": (100, 250), "rate": 25.0, "min_charge": 3000},
    }
    
    # Franchise Partner Rates - MPS (Mega Parcel Service)
    MPS_RATES = {
        "slab_1": {"range": (0, 5), "rate": 60.0, "min_charge": 300},
        "slab_2": {"range": (5, 10), "rate": 55.0, "min_charge": 400},
        "slab_3": {"range": (10, 25), "rate": 50.0, "min_charge": 700},
        "slab_4": {"range": (25, 50), "rate": 45.0, "min_charge": 1200},
        "slab_5": {"range": (50, 100), "rate": 35.0, "min_charge": 2200},
        "slab_6": {"range": (100, 250), "rate": 30.0, "min_charge": 3500},
    }
    
    # Surcharges (as % of base freight)
    FUEL_SURCHARGE_PCT = 0.12  # 12% fuel surcharge
    GST_RATE = 0.18  # 18% GST
    
    # ODA charges
    ODA_CHARGE = 150.0  # Flat ODA charge
    
    def __init__(self, settings: Settings, base_dir: Optional[str] = None):
        super().__init__(settings, base_dir)
        
        # Determine base directory
        if base_dir:
            base_path = Path(base_dir)
        else:
            # Go up 2 levels from this file (forms/calculator/bigship_calculator.py -> project root)
            base_path = Path(__file__).resolve().parents[2]
        
        excel_file = base_path / "Bigship Serviceable Pincode.xlsx"
        print(f"[Bigship] ==================== INITIALIZATION DEBUG ====================")
        print(f"[Bigship] Initializing with base_dir: {base_path}")
        print(f"[Bigship] Looking for Excel file at: {excel_file}")
        print(f"[Bigship] __file__ path: {Path(__file__).resolve()}")
        print(f"[Bigship] Excel file exists: {excel_file.exists()}")
        print(f"[Bigship] =============================================================")
        
        self.bigship_pins = BigshipPincodeDB(excel_file)
    
    def resolve_regions(self, from_pin: PincodeRecord, to_pin: PincodeRecord) -> Dict[str, str]:
        """For Bigship, we don't use zones - just check serviceability"""
        return {
            "from": to_pin.state or "Unknown",
            "to": to_pin.city or "Unknown"
        }
    
    def get_rate_table(self, service_type: str) -> Dict:
        """Get rate table for the selected service type"""
        if service_type == "CFT":
            return self.CFT_RATES
        elif service_type == "LTL":
            return self.LTL_RATES
        elif service_type == "MPS":
            return self.MPS_RATES
        else:
            return self.LTL_RATES  # Default to LTL
    
    def get_rate_for_weight(self, weight: float, service_type: str) -> tuple:
        """Get rate per kg and minimum charge for given weight and service"""
        rates = self.get_rate_table(service_type)
        
        for slab_key, slab_info in rates.items():
            min_w, max_w = slab_info["range"]
            if min_w <= weight < max_w:
                return slab_info["rate"], slab_info["min_charge"]
        
        # If weight exceeds all slabs, use highest slab
        last_slab = list(rates.values())[-1]
        return last_slab["rate"], last_slab["min_charge"]
    
    def calculate_base_freight(self, weight: float, service_type: str) -> float:
        """Calculate base freight charge"""
        rate, min_charge = self.get_rate_for_weight(weight, service_type)
        freight = weight * rate
        return max(freight, min_charge)
    
    def calculate_quote(self, inp: QuoteInput, pins: PincodeDB) -> QuoteResult:
        """Calculate freight quote for Bigship"""
        to_pin = pins.get(inp.to_pincode) or PincodeRecord(pincode=inp.to_pincode)
        
        # Get service type from input (CFT, LTL, or MPS)
        service_type = getattr(inp, 'bigship_service_type', 'LTL')
        
        # Check serviceability
        if not self.bigship_pins.is_serviceable(inp.to_pincode):
            return QuoteResult(
                partner_name=self.name,
                deliverable=False,
                reason=f"Bigship does not service pincode {inp.to_pincode}"
            )
        
        # Calculate chargeable weight
        cw = self.chargeable_weight(inp, volumetric_divisor=5000.0, min_weight=5.0)
        
        # Calculate base freight
        base_freight = self.calculate_base_freight(cw, service_type)
        
        # Calculate surcharges
        surcharges: Dict[str, float] = {}
        
        # Check for ODA (only for LTL and CFT, NOT for MPS)
        # MPS service does not include ODA locations and does not charge ODA
        if service_type in ['LTL', 'CFT'] and self.bigship_pins.is_oda(inp.to_pincode):
            surcharges["oda"] = self.ODA_CHARGE
        
        # Fuel surcharge
        fuel_charge = base_freight * self.FUEL_SURCHARGE_PCT
        surcharges["fuel_surcharge"] = _round_money(fuel_charge)
        
        # Calculate totals
        total_before_gst = base_freight + sum(surcharges.values())
        gst_amount = total_before_gst * self.GST_RATE
        total_after_gst = total_before_gst + gst_amount
        
        # Get destination details
        regions = self.resolve_regions(to_pin, to_pin)
        
        return QuoteResult(
            partner_name=self.name,
            deliverable=True,
            from_zone=regions.get("from"),
            to_zone=regions.get("to"),
            chargeable_weight_kg=cw,
            base_freight=_round_money(base_freight),
            surcharges={k: _round_money(v) for k, v in surcharges.items()},
            total_before_gst=_round_money(total_before_gst),
            gst_amount=_round_money(gst_amount),
            total_after_gst=_round_money(total_after_gst),
            rate_per_kg=self.get_rate_for_weight(cw, service_type)[0],
            actual_weight_kg=sum(item.weight_kg for item in inp.items),
            rate_details={
                "service_type": service_type,
                "is_oda": self.bigship_pins.is_oda(inp.to_pincode),
                "destination_state": to_pin.state,
                "destination_city": to_pin.city,
            }
        )
