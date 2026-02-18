import os
import uuid
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden, JsonResponse
from functools import wraps
import openpyxl
from pathlib import Path
from decimal import Decimal
from datetime import datetime
from django.utils import timezone

# Permission decorators
def stock_manager_required(view_func):
    """Allow only superusers, SnehalShinde, or NileshBagad"""
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        allowed_users = {"SnehalShinde", "NileshBagad"}
        if not (request.user.is_superuser or request.user.username in allowed_users):
            return HttpResponseForbidden("⛔ Access Denied: Only stock managers (Snehal/Nilesh) or superusers can access this page.")
        return view_func(request, *args, **kwargs)
    return _wrapped_view


def logistic_access_required(view_func):
    """Allow only superusers or SakshiChorghe"""
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        allowed_users = {"SakshiChorghe"}
        if not (request.user.is_superuser or request.user.username in allowed_users):
            return HttpResponseForbidden("⛔ Access Denied: Only SakshiChorghe or superusers can access this page.")
        return view_func(request, *args, **kwargs)
    return _wrapped_view

# Helper function to read current stock from database
def get_current_stock_data():
    """Read current stock data from StockItem database"""
    try:
        from .models import StockItem
        
        # Get all stock items ordered by component type and serial number
        stock_items_qs = StockItem.objects.all().order_by('component_type', 'pcba_sn_new')
        
        stock_items = []
        total_balance = 0
        
        for item in stock_items_qs:
            stock_data = {
                'serial_number': item.pcba_sn_new or 'N/A',
                'component': item.component_type or 'Unknown',
                'description': item.specification or '',
                'balance': int(item.quantity) if item.quantity else 0
            }
            stock_items.append(stock_data)
            total_balance += stock_data['balance']
        
        return stock_items, len(stock_items), total_balance
    except Exception as e:
        print(f"Error reading stock data from database: {e}")
        return None, None, None

# Update New Shipments view (same interface as send_stock)
@login_required
@stock_manager_required
def update_new_shipments(request):
    from django.forms import formset_factory
    from .forms import StockRequisitionItemForm
    ItemFormSet = formset_factory(StockRequisitionItemForm, extra=0, can_delete=True)

    if request.method == 'POST':
        # Step 1: Data entry POST, show confirmation page
        if 'confirm_update' not in request.POST:
            item_formset = ItemFormSet(request.POST, prefix='form')
            if item_formset.is_valid():
                # Prepare summary data for confirmation, but pass the formset to the template
                total_qty = 0
                for form in item_formset:
                    if form.cleaned_data.get('serial_number') and not form.cleaned_data.get('DELETE'):
                        qty = form.cleaned_data.get('quantity_received') or 0
                        total_qty += float(qty)
                return render(request, 'forms/update_new_shipments_confirm.html', {
                    'item_formset': item_formset,
                    'total_qty': total_qty,
                })
            else:
                return render(request, 'forms/update_new_shipments.html', {
                    'item_formset': item_formset,
                    'empty_form': item_formset.empty_form,
                })
        # Step 2: Confirm Update POST, save to DB and send email
        else:
            # Save to DB using the formset so all edits and unique items are saved
            from .models import StockItem
            from django.core.mail import EmailMultiAlternatives
            from django.utils.html import escape
            item_formset = ItemFormSet(request.POST, prefix='form')
            items = []
            total_qty = 0
            shipment_date = request.POST.get('shipment_date')
            if item_formset.is_valid():
                for form in item_formset:
                    if form.cleaned_data.get('DELETE'):
                        continue
                    sn = form.cleaned_data.get('serial_number')
                    ct = form.cleaned_data.get('component_type')
                    desc = form.cleaned_data.get('description')
                    qty = form.cleaned_data.get('quantity_received') or 0
                    total_qty += float(qty)
                    items.append({'serial_number': sn, 'component_type': ct, 'description': desc, 'quantity_received': qty})
                    StockItem.objects.create(
                        pcba_sn_new=sn,
                        component_type=ct,
                        specification=desc,
                        quantity=qty,
                        year=(int(shipment_date[:4]) if shipment_date else 2026),
                        shipment_date=shipment_date
                    )
                
                # Send email notification
                try:
                    # Build email content
                    items_html = ""
                    for item in items:
                        items_html += f"""
                        <tr style="border-bottom: 1px solid #e2e8f0;">
                            <td style="padding: 10px; text-align: left;">{escape(item['serial_number'])}</td>
                            <td style="padding: 10px; text-align: left;">{escape(item['component_type'])}</td>
                            <td style="padding: 10px; text-align: left;">{escape(item['description'])}</td>
                            <td style="padding: 10px; text-align: right;">{item['quantity_received']}</td>
                        </tr>
                        """
                    
                    email_html = f"""
                    <html style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f8fafc;">
                    <body style="margin: 0; padding: 20px; background: #f8fafc;">
                        <div style="max-width: 700px; margin: 0 auto; background: white; border-radius: 12px; box-shadow: 0 1px 3px rgba(15, 23, 42, 0.08); overflow: hidden;">
                            <!-- Header -->
                            <div style="background: linear-gradient(135deg, #059669, #10b981); padding: 30px 20px; text-align: center; color: white;">
                                <h1 style="margin: 0; font-size: 28px; font-weight: 700;">📦 New Shipment Received</h1>
                                <p style="margin: 10px 0 0 0; font-size: 14px; opacity: 0.9;">Stock Update Notification</p>
                            </div>
                            
                            <!-- Content -->
                            <div style="padding: 30px;">
                                <p style="margin: 0 0 20px 0; font-size: 15px; color: #475569;">
                                    Hello,<br><br>
                                    A new shipment has been successfully added to the inventory system. Details are provided below:
                                </p>
                                
                                <!-- Shipment Details -->
                                <div style="background: #f1f5f9; border-left: 4px solid #059669; padding: 15px; border-radius: 6px; margin-bottom: 20px;">
                                    <p style="margin: 0; font-size: 14px; font-weight: 600; color: #0f172a;">Shipment Date:</p>
                                    <p style="margin: 5px 0 0 0; font-size: 14px; color: #475569;">{escape(shipment_date or 'Not specified')}</p>
                                </div>
                                
                                <!-- Items Table -->
                                <p style="margin: 0 0 12px 0; font-size: 13px; font-weight: 700; color: #475569; text-transform: uppercase;">📋 Items Added ({len(items)})</p>
                                <table style="width: 100%; border-collapse: collapse; margin-bottom: 20px; font-size: 13px;">
                                    <thead>
                                        <tr style="background: #f1f5f9; border-bottom: 2px solid #cbd5e1;">
                                            <th style="padding: 10px; text-align: left; font-weight: 600; color: #475569;">Serial Number</th>
                                            <th style="padding: 10px; text-align: left; font-weight: 600; color: #475569;">Component Type</th>
                                            <th style="padding: 10px; text-align: left; font-weight: 600; color: #475569;">Description</th>
                                            <th style="padding: 10px; text-align: right; font-weight: 600; color: #475569;">Quantity</th>
                                        </tr>
                                    </thead>
                                    <tbody>
                                        {items_html}
                                    </tbody>
                                </table>
                                
                                <!-- Total -->
                                <div style="background: #ecfdf5; border-left: 4px solid #16a34a; padding: 15px; border-radius: 6px; margin-bottom: 20px;">
                                    <p style="margin: 0; font-size: 14px; font-weight: 600; color: #15803d;">
                                        💰 Total Quantity Received: <span style="font-size: 20px;">{int(total_qty)}</span> units
                                    </p>
                                </div>
                                
                                <p style="margin: 20px 0 0 0; font-size: 13px; color: #64748b;">
                                    You can view and manage this shipment in the inventory management system.
                                </p>
                            </div>
                            
                            <!-- Footer -->
                            <div style="background: #f1f5f9; padding: 20px; text-align: center; border-top: 1px solid #e2e8f0;">
                                <p style="margin: 0; font-size: 12px; color: #64748b;">
                                    This is an automated notification from Deye Forms App.<br>
                                    <a href="https://www.deyeindia.in" style="color: #059669; text-decoration: none;">www.deyeindia.in</a>
                                </p>
                            </div>
                        </div>
                    </body>
                    </html>
                    """
                    
                    # Prepare email
                    subject = f"📦 New Shipment Received - {len(items)} Items ({int(total_qty)} units)"
                    from_email = os.environ.get('DEFAULT_FROM_EMAIL', 'yashraj@deyeindia.com')
                    recipient_list = ['snehal@deyeindia.com', 'nilesh@deyeindia.com']
                    
                    # Send email asynchronously in background thread using SendGrid API
                    import threading
                    import time
                    def _send_shipment_email():
                        try:
                            time.sleep(0.1)  # Give thread a moment to start
                            print(f"[EMAIL] 📦 Starting to send shipment email to {recipient_list}")
                            from forms.emails import send_sendgrid_email
                            success = send_sendgrid_email(recipient_list, subject, email_html, 'New shipment received')
                            if success:
                                print(f"[EMAIL] ✅ Shipment email sent for {len(items)} items via SendGrid API")
                            else:
                                print(f"[EMAIL] ⚠️ Shipment email sending had issues")
                        except Exception as email_err:
                            print(f"[EMAIL] ❌ Shipment email error: {str(email_err)}")
                            import traceback
                            traceback.print_exc()
                    
                    email_thread = threading.Thread(target=_send_shipment_email, daemon=False)
                    email_thread.start()
                    print(f"[EMAIL] 🧵 Background thread started for shipment with {len(items)} items")
                    
                except Exception as e:
                    # Log error but don't crash the app
                    print(f"Email preparation error: {str(e)}")
                
                return render(request, 'forms/update_new_shipments_confirm.html', {
                    'items': items,
                    'total_qty': total_qty,
                    'shipment_date': shipment_date,
                    'success_message': 'Shipments confirmed and saved!'
                })
            else:
                # If invalid, return to edit page with errors
                return render(request, 'forms/update_new_shipments.html', {
                    'item_formset': item_formset,
                    'empty_form': item_formset.empty_form,
                })
    else:
        item_formset = ItemFormSet(prefix='form')
    return render(request, 'forms/update_new_shipments.html', {
        'item_formset': item_formset,
        'empty_form': item_formset.empty_form,
    })
from django.views.decorators.http import require_GET
from django.contrib.auth.decorators import login_required
@login_required
@require_GET
def available_battery_serials(request):
    from django.db.models import Count
    inward_counts = InwardForm.objects.values('battery_id').annotate(count_inward=Count('id'))
    outward_counts = OutwardForm.objects.values('battery_id').annotate(count_outward=Count('id'))
    outward_map = {row['battery_id']: row['count_outward'] for row in outward_counts}
    available_serials = []
    for row in inward_counts:
        serial = row['battery_id']
        if not serial:
            continue
        inward = row['count_inward']
        outward = outward_map.get(serial, 0)
        if inward > outward:
            available_serials.append(serial)
    return JsonResponse({'serials': available_serials})
from django.views.decorators.http import require_GET
import json
from django.contrib.auth.decorators import login_required
# API endpoint for available inverter serials (inwarded but not outwarded)
@login_required
@require_GET
def available_inverter_serials(request):
    # For each inverter_id, count inward and outward occurrences
    from django.db.models import Count
    inward_counts = InwardForm.objects.values('inverter_id').annotate(count_inward=Count('id'))
    outward_counts = OutwardForm.objects.values('inverter_id_outward').annotate(count_outward=Count('id'))
    outward_map = {row['inverter_id_outward']: row['count_outward'] for row in outward_counts}
    available_serials = []
    for row in inward_counts:
        serial = row['inverter_id']
        inward = row['count_inward']
        outward = outward_map.get(serial, 0)
        if inward > outward:
            available_serials.append(serial)
    return JsonResponse({'serials': available_serials})
from django.db.models import OuterRef, Subquery, Exists
# Stock Inverters View
from .models import InwardForm, OutwardForm

from django.contrib.auth.decorators import login_required

@login_required
def stock_inverters_view(request):
    # Get all inverter_ids from OutwardForm
    outward_ids = set(OutwardForm.objects.values_list('inverter_id_outward', flat=True))
    # Get all InwardForm entries whose inverter_id is NOT in OutwardForm
    stock_inverters = InwardForm.objects.exclude(inverter_id__in=outward_ids).order_by('-created_at')
    return render(request, 'forms/stock_inverters.html', {'stock_inverters': stock_inverters})
# Simple static view for hierarchy page
from django.contrib.auth.decorators import login_required

@login_required
def hierarchy_static_view(request):
    return render(request, 'Include/hierarchy.html')
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from collections import defaultdict
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.contrib.auth import login
from django.http import HttpResponse, JsonResponse, HttpResponseForbidden
from django.core.paginator import Paginator
from django.core.mail import send_mail, EmailMultiAlternatives
from django.conf import settings
from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.urls import reverse
from datetime import date, timedelta
import csv

from .models import (
    RepairingForm,
    InwardForm,
    OutwardForm,
    ServiceReportForm,
    LeaveRequest,
    UpcomingEvent,
    UserProfile,
    StockRequisition,
    DispatchedStock,
    StockItem,
    LogisticBooking
)

from .forms import (
    RepairingFormForm,
    InwardFormForm,
    OutwardFormForm,
    ServiceReportFormForm,
    UserRegistrationForm,
    FreightCalculatorForm,
)

from .calculator import get_all_partner_quotes, QuoteInput
from .calculator.freight_calculator import ShipmentItem
from .calculator.data_loader import load_pincode_master, lookup_pincode


# ─────────────────────────────
# HOME
# ─────────────────────────────

@login_required
def simple_home(request):
    today = date.today()
    month_start = date(today.year, today.month, 1)
    if today.month == 12:
        month_end = date(today.year + 1, 1, 1) - timedelta(days=1)
    else:
        month_end = date(today.year, today.month + 1, 1) - timedelta(days=1)

    from .models import Holiday

    upcoming_items = []

    festival_holidays = Holiday.objects.filter(
        category='Festival',
        date__gte=month_start,
        date__lte=month_end,
    ).order_by('date')
    for holiday in festival_holidays:
        upcoming_items.append({
            'event_date': holiday.date,
            'title': holiday.name,
            'location': '',
            'description': 'Festival Holiday',
            'event_kind': 'Festival',
        })

    birthday_profiles = UserProfile.objects.select_related('user').filter(
        user__is_active=True,
        date_of_birth__isnull=False,
    )
    for profile in birthday_profiles:
        dob = profile.date_of_birth
        if not dob:
            continue
        try:
            next_birthday = date(today.year, dob.month, dob.day)
        except ValueError:
            continue

        if next_birthday < month_start:
            try:
                next_birthday = date(today.year + 1, dob.month, dob.day)
            except ValueError:
                continue

        if next_birthday < month_start or next_birthday > month_end:
            continue

        employee_name = profile.user.get_full_name() or profile.user.username
        upcoming_items.append({
            'event_date': next_birthday,
            'title': f"{employee_name}'s Birthday",
            'location': '',
            'description': 'Team member birthday',
            'event_kind': 'Birthday',
        })

    upcoming_items = sorted(upcoming_items, key=lambda item: item['event_date'])

    team_members = []
    if hasattr(request.user, 'profile'):
        team_members = request.user.profile.get_team_members()
    
    # Check if user can view team attendance (manager or superuser)
    from django.contrib.auth.models import User
    is_manager = User.objects.filter(profile__manager=request.user).exists()
    show_team_attendance = is_manager or request.user.is_superuser
    
    return render(request, 'forms/simple_home_modern.html', {
        'events': upcoming_items,
        'user': request.user, 
        'team_members': team_members,
        'show_team_attendance': show_team_attendance
    })


@login_required
def change_password(request):
    """Allow users to change their password with current password confirmation"""
    from django import forms
    from django.contrib.auth import update_session_auth_hash
    
    class PasswordChangeForm(forms.Form):
        current_password = forms.CharField(
            widget=forms.PasswordInput(attrs={'class': 'form-input', 'placeholder': 'Current Password'}),
            label='Current Password'
        )
        new_password = forms.CharField(
            widget=forms.PasswordInput(attrs={'class': 'form-input', 'placeholder': 'New Password'}),
            label='New Password',
            min_length=8
        )
        confirm_password = forms.CharField(
            widget=forms.PasswordInput(attrs={'class': 'form-input', 'placeholder': 'Confirm New Password'}),
            label='Confirm New Password',
            min_length=8
        )
        
        def clean(self):
            cleaned_data = super().clean()
            current_password = cleaned_data.get('current_password')
            new_password = cleaned_data.get('new_password')
            confirm_password = cleaned_data.get('confirm_password')
            
            if new_password != confirm_password:
                raise forms.ValidationError("New passwords do not match.")
            
            if new_password == current_password:
                raise forms.ValidationError("New password must be different from current password.")
            
            return cleaned_data
    
    if request.method == 'POST':
        form = PasswordChangeForm(request.POST)
        if form.is_valid():
            current_password = form.cleaned_data.get('current_password')
            new_password = form.cleaned_data.get('new_password')
            
            # Verify current password
            if not request.user.check_password(current_password):
                form.add_error('current_password', 'Current password is incorrect.')
                return render(request, 'forms/change_password.html', {'form': form, 'user': request.user})
            
            # Update password
            request.user.set_password(new_password)
            request.user.save()
            
            # Keep user logged in after password change
            update_session_auth_hash(request, request.user)
            
            from django.contrib import messages
            messages.success(request, '✅ Password changed successfully!')
            return redirect('forms:simple_home')
    else:
        form = PasswordChangeForm()
    
    return render(request, 'forms/change_password.html', {'form': form, 'user': request.user})


@login_required
def user_profile(request):
    """View and edit user profile"""
    from django import forms
    from .models import UserProfile
    
    # Create or get user profile
    profile, created = UserProfile.objects.get_or_create(user=request.user)
    
    class ProfileEditForm(forms.ModelForm):
        first_name = forms.CharField(
            max_length=30,
            required=False,
            widget=forms.TextInput(attrs={'class': 'form-input', 'placeholder': 'First Name'})
        )
        last_name = forms.CharField(
            max_length=30,
            required=False,
            widget=forms.TextInput(attrs={'class': 'form-input', 'placeholder': 'Last Name'})
        )
        email = forms.EmailField(
            required=False,
            widget=forms.EmailInput(attrs={'class': 'form-input', 'placeholder': 'Email Address'})
        )
        
        class Meta:
            model = UserProfile
            fields = ['profile_photo', 'phone']
            widgets = {
                'profile_photo': forms.FileInput(attrs={'class': 'form-input', 'accept': 'image/*'}),
                'phone': forms.TextInput(attrs={'class': 'form-input', 'placeholder': 'Phone Number'}),
            }
    
    if request.method == 'POST':
        form = ProfileEditForm(request.POST, request.FILES, instance=profile)
        if form.is_valid():
            # Update user fields
            request.user.first_name = form.cleaned_data.get('first_name', '')
            request.user.last_name = form.cleaned_data.get('last_name', '')
            request.user.email = form.cleaned_data.get('email', '')
            request.user.save()
            
            # Save profile (including photo)
            form.save()
            
            from django.contrib import messages
            messages.success(request, '✅ Profile updated successfully!')
            return redirect('forms:user_profile')
    else:
        form = ProfileEditForm(instance=profile, initial={
            'first_name': request.user.first_name,
            'last_name': request.user.last_name,
            'email': request.user.email,
        })
    
    return render(request, 'forms/user_profile.html', {
        'form': form,
        'user': request.user,
        'profile': profile
    })


@login_required
def user_id_card(request):
    """Display user ID card"""
    profile, created = UserProfile.objects.get_or_create(user=request.user)
    
    return render(request, 'forms/user_id_card.html', {
        'user': request.user,
        'profile': profile
    })


@login_required
def download_id_card(request):
    """Download ID card as image"""
    from django.http import HttpResponse
    from io import BytesIO
    try:
        from PIL import Image, ImageDraw, ImageFont
    except ImportError:
        from django.contrib import messages
        messages.error(request, 'ID card download feature requires Pillow library.')
        return redirect('forms:user_id_card')
    
    profile, created = UserProfile.objects.get_or_create(user=request.user)
    
    # Create ID card image (850x540 pixels)
    width, height = 850, 540
    img = Image.new('RGB', (width, height), color='white')
    draw = ImageDraw.Draw(img)
    
    # Background gradient effect (simple two-tone)
    for y in range(height):
        shade = int(255 - (y / height) * 50)
        draw.rectangle([(0, y), (width, y+1)], fill=(shade, shade, 255))
    
    # Company header
    draw.rectangle([(0, 0), (width, 100)], fill=(102, 126, 234))
    
    try:
        # Try to load fonts
        title_font = ImageFont.truetype("arial.ttf", 40)
        heading_font = ImageFont.truetype("arial.ttf", 24)
        text_font = ImageFont.truetype("arial.ttf", 20)
        small_font = ImageFont.truetype("arial.ttf", 16)
    except:
        # Fallback to default font
        title_font = ImageFont.load_default()
        heading_font = ImageFont.load_default()
        text_font = ImageFont.load_default()
        small_font = ImageFont.load_default()
    
    # Draw company name
    draw.text((width//2, 50), "DEYE INDIA", fill='white', font=title_font, anchor='mm')
    
    # Draw "EMPLOYEE ID CARD" text
    draw.text((width//2, 140), "EMPLOYEE ID CARD", fill=(102, 126, 234), font=heading_font, anchor='mm')
    
    # Profile photo placeholder/actual photo
    photo_x, photo_y = 80, 200
    photo_size = 150
    if profile.profile_photo:
        try:
            from django.conf import settings
            import os
            photo_path = os.path.join(settings.MEDIA_ROOT, profile.profile_photo.name)
            if os.path.exists(photo_path):
                photo = Image.open(photo_path)
                photo = photo.resize((photo_size, photo_size))
                img.paste(photo, (photo_x, photo_y))
        except:
            # Draw placeholder if photo fails to load
            draw.rectangle([(photo_x, photo_y), (photo_x + photo_size, photo_y + photo_size)], fill=(200, 200, 200), outline=(102, 126, 234), width=3)
            draw.text((photo_x + photo_size//2, photo_y + photo_size//2), "PHOTO", fill=(102, 126, 234), font=text_font, anchor='mm')
    else:
        # Draw placeholder
        draw.rectangle([(photo_x, photo_y), (photo_x + photo_size, photo_y + photo_size)], fill=(220, 220, 220), outline=(102, 126, 234), width=3)
        draw.text((photo_x + photo_size//2, photo_y + photo_size//2), "NO PHOTO", fill=(102, 126, 234), font=small_font, anchor='mm')
    
    # Employee details
    details_x = 280
    y_pos = 210
    line_height = 40
    
    # Name
    name = request.user.get_full_name() or request.user.username
    draw.text((details_x, y_pos), f"Name:", fill=(60, 60, 60), font=text_font)
    draw.text((details_x + 150, y_pos), name.upper(), fill=(20, 20, 20), font=text_font)
    
    # Employee ID
    y_pos += line_height
    emp_id = profile.employee_id or "N/A"
    draw.text((details_x, y_pos), f"Employee ID:", fill=(60, 60, 60), font=text_font)
    draw.text((details_x + 150, y_pos), emp_id, fill=(20, 20, 20), font=text_font)
    
    # Department
    y_pos += line_height
    department = profile.department or "N/A"
    draw.text((details_x, y_pos), f"Department:", fill=(60, 60, 60), font=text_font)
    draw.text((details_x + 150, y_pos), department, fill=(20, 20, 20), font=text_font)
    
    # Designation
    y_pos += line_height
    designation = profile.designation or "N/A"
    draw.text((details_x, y_pos), f"Designation:", fill=(60, 60, 60), font=text_font)
    draw.text((details_x + 150, y_pos), designation, fill=(20, 20, 20), font=text_font)
    
    # Joining Date
    y_pos += line_height
    joining = profile.joining_date.strftime("%d %b %Y") if profile.joining_date else "N/A"
    draw.text((details_x, y_pos), f"Joined:", fill=(60, 60, 60), font=text_font)
    draw.text((details_x + 150, y_pos), joining, fill=(20, 20, 20), font=text_font)
    
    # Footer
    draw.rectangle([(0, height-60), (width, height)], fill=(102, 126, 234))
    draw.text((width//2, height-30), "Deye Energy Solutions | www.deyeindia.com", fill='white', font=small_font, anchor='mm')
    
    # Save to bytes
    buffer = BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='image/png')
    response['Content-Disposition'] = f'attachment; filename="employee_id_{request.user.username}.png"'
    return response


@login_required
def stock_home(request):
    """Stock management home with 3 options"""
    return render(request, 'forms/stock_home.html')


@login_required
def received_stock(request):
    if not request.user.is_superuser:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("You do not have permission to view received stock.")
    """Received stock dashboard with filtering"""
    from .models import StockItem
    from django.db.models import Q
    
    # Get filters from query params
    selected_year = request.GET.get('year', 'all')
    selected_component = request.GET.get('component', 'all')
    search_query = request.GET.get('search', '').strip()
    
    # Get all available years and component types
    years = StockItem.objects.values_list('year', flat=True).distinct().order_by('-year')
    all_component_types = StockItem.objects.values_list('component_type', flat=True).distinct().order_by('component_type')
    all_component_types = [ct for ct in all_component_types if ct]  # Remove empty values
    
    # Start with all items
    stock_items = StockItem.objects.all()
    
    # Apply year filter
    if selected_year != 'all':
        try:
            year_int = int(selected_year)
            stock_items = stock_items.filter(year=year_int)
        except ValueError:
            pass
    
    # Apply component type filter
    if selected_component != 'all':
        stock_items = stock_items.filter(component_type=selected_component)
    
    # Apply search filter
    if search_query:
        stock_items = stock_items.filter(
            Q(pcba_sn_new__icontains=search_query) |
            Q(pcba_sn_old__icontains=search_query) |
            Q(specification__icontains=search_query) |
            Q(remark__icontains=search_query)
        )
    
    # Get statistics
    total_items = stock_items.count()
    total_quantity = stock_items.aggregate(total=Sum('quantity'))['total'] or 0
    component_types = stock_items.values('component_type').annotate(
        count=Count('id'),
        total_qty=Sum('quantity')
    ).order_by('-count')[:10]

    # Merge duplicates by PCBA SN across all years in the filtered set
    # For items with serial numbers, merge across years
    # For items without serial numbers, keep them separate
    merged = {}
    for item in stock_items.order_by('pcba_sn_new'):
        # For non-serialized items (blank serial), use a unique key based on component+spec+year
        if not item.pcba_sn_new:
            # Create unique key for non-serialized items
            key = f"no_sn_{item.component_type or 'unknown'}_{item.specification or 'unknown'}_{item.year}"
        else:
            # For serialized items, merge across all years
            key = item.pcba_sn_new
        
        if key not in merged:
            merged[key] = {
                'pcba_sn_new': item.pcba_sn_new,
                'pcba_sn_old': item.pcba_sn_old,
                'component_type': item.component_type,
                'specification': item.specification,
                'quantity': item.quantity,
                'years': {item.year} if item.year else set(),
                'remark': item.remark or '',
            }
        else:
            merged[key]['quantity'] += item.quantity
            if item.pcba_sn_old and not merged[key]['pcba_sn_old']:
                merged[key]['pcba_sn_old'] = item.pcba_sn_old
            if item.component_type and not merged[key]['component_type']:
                merged[key]['component_type'] = item.component_type
            if item.specification and not merged[key]['specification']:
                merged[key]['specification'] = item.specification
            if item.remark:
                if merged[key]['remark'] and item.remark not in merged[key]['remark']:
                    merged[key]['remark'] = f"{merged[key]['remark']}; {item.remark}"
                elif not merged[key]['remark']:
                    merged[key]['remark'] = item.remark
            if item.year:
                merged[key]['years'].add(item.year)

    merged_list = list(merged.values())

    # Sort merged list by latest year desc, then component_type, then PCBA
    # Handle None values for serial numbers (blank serial items)
    merged_list.sort(key=lambda x: (
        -(max(x['years']) if x['years'] else 0),
        x['component_type'] or '',
        x['pcba_sn_new'] or ''  # Convert None to empty string for sorting
    ))

    # Override stats to reflect merged items
    total_items = len(merged_list)
    total_quantity = sum([m['quantity'] for m in merged_list])

    # Paginate merged results
    from django.core.paginator import Paginator
    paginator = Paginator(merged_list, 50)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        'years': years,
        'selected_year': selected_year,
        'all_component_types': all_component_types,
        'selected_component': selected_component,
        'search_query': search_query,
        'stock_items': page_obj,
        'total_items': total_items,
        'total_quantity': total_quantity,
        'component_types': component_types,
    }
    return render(request, 'forms/received_stock.html', context)


@login_required
@stock_manager_required
def remaining_stock(request):
    """Remaining stock view - Display current inventory from Excel file"""
    stock_items, total_rows, total_balance = get_current_stock_data()
    
    context = {
        'stock_items': stock_items,
        'total_rows': total_rows,
        'total_balance': total_balance,
    }
    return render(request, 'forms/remaining_stock.html', context)


@login_required
def send_stock(request):
    """Stock requisition form (supports multiple items in one submission)."""
    try:
        from django.forms import formset_factory
        from .forms import StockRequisitionHeaderForm, StockRequisitionItemForm
        from .models import StockRequisition
        from .emails import send_bulk_requisition_email

        ItemFormSet = formset_factory(StockRequisitionItemForm, extra=1, can_delete=True)

        if request.method == 'POST':
            header_form = StockRequisitionHeaderForm(request.POST)
            item_formset = ItemFormSet(request.POST, prefix='form')

            if header_form.is_valid() and item_formset.is_valid():
                header_cd = header_form.cleaned_data
                created_requisitions = []
                
                for form in item_formset:
                    # Skip the form if it has DELETE checked
                    if form.cleaned_data.get('DELETE'):
                        continue
                    
                    # Skip the form if it's completely empty (no serial number)
                    if not form.cleaned_data.get('serial_number'):
                        continue
                    
                    # Create requisition only for forms with valid serial number
                    cd = form.cleaned_data
                    requisition = StockRequisition.objects.create(
                        serial_number=cd.get('serial_number'),
                        component_type=cd.get('component_type', ''),
                        description=cd.get('description', ''),
                        manager_name=header_cd['manager_name'],
                        quantity_required=cd.get('quantity_required') or 0,
                        required_to=header_cd['required_to'],
                        status='pending',
                    )
                    created_requisitions.append(requisition)

                created_count = len(created_requisitions)
                
                # Send consolidated email with all requisitions
                if created_requisitions:
                    try:
                        send_bulk_requisition_email(
                            created_requisitions,
                            header_cd['manager_name'],
                            header_cd['required_to']
                        )
                        print(f"✅ Email sent successfully for {created_count} requisitions")
                    except Exception as e:
                        print(f"❌ Error sending email: {str(e)}")
                        import traceback
                        traceback.print_exc()

                success_message = f'Stock requisition submitted successfully for {created_count} item(s)! Approvers have been notified.'
                # Fresh forms after successful submit
                header_form = StockRequisitionHeaderForm()
                item_formset = ItemFormSet(prefix='form')
                return render(request, 'forms/send_stock.html', {
                    'header_form': header_form,
                    'item_formset': item_formset,
                    'empty_form': item_formset.empty_form,
                    'success_message': success_message,
                })
            else:
                # Form validation failed - return with errors
                print(f"Form validation failed:")
                print(f"Header form errors: {header_form.errors}")
                print(f"Formset errors: {item_formset.errors}")
                return render(request, 'forms/send_stock.html', {
                    'header_form': header_form,
                    'item_formset': item_formset,
                    'empty_form': item_formset.empty_form,
                })
        else:
            header_form = StockRequisitionHeaderForm()
            item_formset = ItemFormSet(prefix='form')

        return render(request, 'forms/send_stock.html', {
            'header_form': header_form,
            'item_formset': item_formset,
            'empty_form': item_formset.empty_form,
        })
    except Exception as e:
        print(f"❌ CRITICAL ERROR in send_stock view: {str(e)}")
        import traceback
        traceback.print_exc()
        from django.http import HttpResponse
        return HttpResponse(f"Error: {str(e)}<br><br>Please contact support.", status=500)


@login_required
@stock_manager_required
def dispatched_stock(request):
    """Dispatch approved stock requisitions."""
    if request.method == 'POST':
        try:
            requisition_id = request.POST.get('requisition_id')
            if not requisition_id:
                return JsonResponse({'success': False, 'error': 'No requisition selected'})
            
            requisition = StockRequisition.objects.get(id=requisition_id)
            
            # Validate requisition is approved
            if requisition.status != 'approved':
                return JsonResponse({'success': False, 'error': 'Requisition not approved'})
            
            # Get form data
            dispatch_date = request.POST.get('dispatch_date')
            courier_name = request.POST.get('courier_name')
            tracking_number = request.POST.get('tracking_number', '')
            quantity_dispatched = int(request.POST.get('quantity_dispatched', 0))
            dispatch_remarks = request.POST.get('dispatch_remarks', '')
            
            # Validate quantity
            approved_qty = requisition.approved_quantity or requisition.quantity_required
            if quantity_dispatched > approved_qty:
                return JsonResponse({
                    'success': False, 
                    'error': f'Cannot dispatch {quantity_dispatched}. Approved quantity is {approved_qty}'
                })
            
            # Create dispatch record
            dispatch = DispatchedStock.objects.create(
                requisition=requisition,
                serial_number=requisition.serial_number,
                component_type=requisition.component_type,
                quantity_dispatched=quantity_dispatched,
                engineer_name=requisition.manager_name,
                dispatch_location=requisition.required_to,
                dispatch_date=dispatch_date,
                courier_name=courier_name,
                tracking_number=tracking_number,
                dispatch_remarks=dispatch_remarks,
                dispatched_by=request.user
            )
            
            # Update stock item quantity
            try:
                stock_item = StockItem.objects.get(pcba_sn_new=requisition.serial_number)
                stock_item.current_quantity = (stock_item.current_quantity or 0) - quantity_dispatched
                stock_item.save()
            except StockItem.DoesNotExist:
                pass
            
            # Update requisition status to dispatched if fully dispatched
            if quantity_dispatched >= approved_qty:
                requisition.status = 'dispatched'
                requisition.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Stock dispatched successfully! Tracking: {tracking_number}',
                'dispatch_id': dispatch.id
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    # GET request - show form
    approved_requisitions = StockRequisition.objects.filter(
        status='approved'
    ).order_by('-approval_date')
    
    return render(request, 'forms/dispatched_form.html', {
        'approved_requisitions': approved_requisitions
    })


@login_required
def stock_serial_search(request):
    """Return up to 10 matching serial suggestions from received stock."""
    from .models import StockItem
    q = (request.GET.get('q') or '').strip()
    results = []
    if q:
        qs = StockItem.objects.filter(pcba_sn_new__icontains=q).values(
            'pcba_sn_new', 'component_type', 'specification', 'year'
        ).order_by('-year')[:10]
        for row in qs:
            results.append({
                'pcba_sn_new': row['pcba_sn_new'],
                'component_type': row['component_type'] or '',
                'specification': row['specification'] or '',
                'year': row['year'] or ''
            })
    return JsonResponse({'results': results})


@login_required
def stock_serial_details(request):
    """Return details for an exact serial: latest year item, component type and specification."""
    from .models import StockItem
    sn = (request.GET.get('sn') or '').strip()
    if not sn:
        return JsonResponse({'found': False})
    item = StockItem.objects.filter(pcba_sn_new=sn).order_by('-year').first()
    if not item:
        return JsonResponse({'found': False})
    return JsonResponse({
        'found': True,
        'pcba_sn_new': item.pcba_sn_new,
        'component_type': item.component_type or '',
        'specification': item.specification or '',
        'year': item.year,
        'quantity': float(item.quantity) if item.quantity is not None else 0
    })


@login_required
def stock_info_api(request):
    """Return stock info for a requisition ID."""
    req_id = request.GET.get('requisition_id')
    if not req_id:
        return JsonResponse({'success': False, 'error': 'No requisition ID provided'})
    
    try:
        requisition = StockRequisition.objects.get(id=req_id)
        
        # Get current stock
        current_stock = 0
        try:
            stock_item = StockItem.objects.get(pcba_sn_new=requisition.serial_number)
            current_stock = stock_item.current_quantity or 0
        except StockItem.DoesNotExist:
            pass
        
        # Calculate quantities
        approved_qty = requisition.approved_quantity or requisition.quantity_required
        remaining = current_stock - approved_qty
        
        return JsonResponse({
            'success': True,
            'serial_number': requisition.serial_number,
            'component_type': requisition.component_type,
            'engineer_name': requisition.manager_name,
            'dispatch_location': requisition.required_to,
            'quantity_approved': approved_qty,
            'current_stock': current_stock,
            'remaining_stock': remaining,
            'approved_by': requisition.approved_by or '',
            'approval_date': requisition.approval_date.strftime('%Y-%m-%d') if requisition.approval_date else ''
        })
    except StockRequisition.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Requisition not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# ─────────────────────────────
# FORMS DASHBOARD
# ─────────────────────────────

@login_required
def select_form_page(request):
    return render(request, 'forms/select_form.html')


@login_required
@logistic_access_required
def logistic_booking_create(request):
    from django.contrib.auth.models import User

    engineers = User.objects.filter(is_active=True).order_by('first_name', 'last_name', 'username')

    courier_partner_base = [
        'Safexpress',
        'Bigship',
        'Bluedart',
        'Anjani',
        'Global Cargo',
    ]

    capacity_options = [
        '2.2 KW', '3 KW', '3.3 KW', '4 KW', '5 KW', '6 KW', '8 KW', '10 KW', '12 KW',
        '15 KW', '18 KW', '20 KW', '25 KW', '30 KW', '33 KW', '35 KW', '40 KW',
        '50 KW', '60 KW', '75 KW', '80 KW', '100 KW', '125 KW', '136 KW'
    ]

    inverter_type_options = [
        '1PH - Hybrid',
        '3PH - Hybrid',
        '1PH - Ongrid',
        '3PH - Ongrid',
        'All In One - Hybrid',
        'Other',
    ]

    battery_model_options = [
        'RW - L 2.5 Neutral',
        'RW - M 5.3 Neutral',
        'RW - M 5.3 Pro Neutral (M6)',
        'RW - M 6.1 Neutral',
        'RW - M 6.1 B Neutral No Remark',
        'SE - G 5.1 - Pro B Neutral No Remark',
        'AI - W 5.1 Neutral',
        'AI - W 5.1 - B Neutral No Remark',
        'BOS - GM 5.1 Neutral',
        'GB - LM 4.0 Neutral',
        'GB - LB Neutral',
        'SE - G 5.3',
        'SE - G 5.3 Pro',
        'SE - F 5',
        'Other',
    ]

    pickup_status_options = [
        'Pickup Done',
        'Pickup Not Done',
        'Order Cancelled',
        'Only Invoice',
    ]

    delivery_status_options = [
        'Delivered',
        'Undelivered',
        'OFD',
        'Order Cancelled',
        'In-Transit',
        'RTO Delivered',
        'Only Invoice',
    ]

    pickup_status_options = [
        'Pickup Done',
        'Pickup Not Done',
        'Order Cancelled',
        'Only Invoice',
    ]

    delivery_status_options = [
        'Delivered',
        'Undelivered',
        'OFD',
        'Order Cancelled',
        'In-Transit',
        'RTO Delivered',
        'Only Invoice',
    ]

    existing_partners = list(
        LogisticBooking.objects.values_list('courier_partner', flat=True).distinct()
    )
    extra_partners = [p for p in existing_partners if p and p not in courier_partner_base]
    courier_partner_choices = courier_partner_base + sorted(extra_partners)

    if request.method == 'POST':
        data = request.POST
        errors = []

        def _get_value(key, required=False):
            value = (data.get(key) or '').strip()
            if required and not value:
                errors.append(f"{key.replace('_', ' ').title()} is required")
            return value

        engineer_id = _get_value('engineer_id', required=True)
        engineer = None
        if engineer_id:
            engineer = User.objects.filter(id=engineer_id).first()
        if not engineer:
            errors.append('Engineer name is required')

        customer_name = _get_value('customer_name', required=True)
        contact_details = _get_value('contact_details', required=True)

        pickup_pincode = _get_value('pickup_pincode', required=True)
        pickup_state = _get_value('pickup_state')
        pickup_district = _get_value('pickup_district')
        pickup_city = _get_value('pickup_city')
        pickup_address = _get_value('pickup_address')

        delivery_name = _get_value('delivery_name', required=True)
        delivery_contact = _get_value('delivery_contact', required=True)
        delivery_pincode = _get_value('delivery_pincode', required=True)
        delivery_state = _get_value('delivery_state')
        delivery_district = _get_value('delivery_district')
        delivery_city = _get_value('delivery_city')
        delivery_address = _get_value('delivery_address')

        courier_partner = _get_value('courier_partner', required=True)
        courier_partner_other = _get_value('courier_partner_other')
        if courier_partner == '__other__':
            if not courier_partner_other:
                errors.append('Other courier partner is required')
            courier_partner = courier_partner_other

        shipment_weight_raw = _get_value('shipment_weight', required=True)
        invoice_number = _get_value('invoice_number', required=True)
        awb_number = _get_value('awb_number')

        pickup_status = _get_value('pickup_status')
        pickup_date_raw = _get_value('pickup_date')
        delivery_status = _get_value('delivery_status')
        delivery_date_raw = _get_value('delivery_date')
        remark = _get_value('remark')

        try:
            shipment_weight = Decimal(shipment_weight_raw) if shipment_weight_raw else None
        except Exception:
            shipment_weight = None
            errors.append('Shipment weight must be a valid number')

        pickup_date = None
        delivery_date = None
        if pickup_date_raw:
            try:
                pickup_date = datetime.strptime(pickup_date_raw, '%Y-%m-%d').date()
            except ValueError:
                errors.append('Pickup date must be valid')
        if delivery_date_raw:
            try:
                delivery_date = datetime.strptime(delivery_date_raw, '%Y-%m-%d').date()
            except ValueError:
                errors.append('Delivery date must be valid')

        object_types = data.getlist('object_type')
        object_capacities = data.getlist('object_capacity')
        object_variants = data.getlist('object_variant')
        object_serials = data.getlist('object_serial_number')
        object_qtys = data.getlist('object_quantity')

        items = []
        for idx, obj_type in enumerate(object_types):
            obj_type = (obj_type or '').strip()
            if not obj_type:
                continue
            capacity = (object_capacities[idx] if idx < len(object_capacities) else '').strip()
            variant = (object_variants[idx] if idx < len(object_variants) else '').strip()
            serial_number = (object_serials[idx] if idx < len(object_serials) else '').strip()
            qty_raw = (object_qtys[idx] if idx < len(object_qtys) else '').strip()
            try:
                quantity = int(qty_raw) if qty_raw else 1
            except Exception:
                quantity = 1
            if quantity < 1:
                quantity = 1
            items.append({
                'object_type': obj_type,
                'object_capacity': capacity,
                'object_variant': variant,
                'object_serial_number': serial_number,
                'object_quantity': quantity,
            })

        if not items:
            errors.append('At least one object row is required')

        if errors:
            return render(request, 'forms/logistic_form.html', {
                'errors': errors,
                'form_data': data,
                'items': items,
                'courier_partners': courier_partner_choices,
                'capacity_options': capacity_options,
                'inverter_type_options': inverter_type_options,
                'battery_model_options': battery_model_options,
                'pickup_status_options': pickup_status_options,
                'delivery_status_options': delivery_status_options,
                'engineers': engineers,
            })

        batch_id = str(uuid.uuid4())
        for item in items:
            LogisticBooking.objects.create(
                created_by=request.user,
                engineer=engineer,
                customer_name=customer_name,
                contact_details=contact_details,
                pickup_pincode=pickup_pincode,
                pickup_state=pickup_state,
                pickup_district=pickup_district,
                pickup_city=pickup_city,
                pickup_address=pickup_address,
                delivery_name=delivery_name,
                delivery_contact=delivery_contact,
                delivery_pincode=delivery_pincode,
                delivery_state=delivery_state,
                delivery_district=delivery_district,
                delivery_city=delivery_city,
                delivery_address=delivery_address,
                object_type=item['object_type'],
                object_capacity=item['object_capacity'],
                object_variant=item['object_variant'],
                object_serial_number=item['object_serial_number'],
                object_quantity=item['object_quantity'],
                courier_partner=courier_partner,
                shipment_weight=shipment_weight,
                awb_number=awb_number,
                invoice_number=invoice_number,
                pickup_status=pickup_status,
                pickup_date=pickup_date,
                delivery_status=delivery_status,
                delivery_date=delivery_date,
                remark=remark,
                batch_id=batch_id,
            )

        return render(request, 'forms/logistic_form.html', {
            'success_message': 'Logistic booking saved successfully.',
            'courier_partners': courier_partner_choices,
            'capacity_options': capacity_options,
            'inverter_type_options': inverter_type_options,
            'battery_model_options': battery_model_options,
            'pickup_status_options': pickup_status_options,
            'delivery_status_options': delivery_status_options,
            'engineers': engineers,
            'items': [{'object_type': 'Inverter', 'object_capacity': '', 'object_serial_number': '', 'object_quantity': 1}],
        })

    return render(request, 'forms/logistic_form.html', {
        'courier_partners': courier_partner_choices,
        'capacity_options': capacity_options,
        'inverter_type_options': inverter_type_options,
        'battery_model_options': battery_model_options,
        'pickup_status_options': pickup_status_options,
        'delivery_status_options': delivery_status_options,
        'engineers': engineers,
        'items': [{'object_type': 'Inverter', 'object_capacity': '', 'object_serial_number': '', 'object_quantity': 1}],
    })


@login_required
@logistic_access_required
def logistic_booking_list(request):
    from django.contrib.auth.models import User

    bookings = LogisticBooking.objects.all()

    from_date_str = (request.GET.get('from_date') or '').strip()
    to_date_str = (request.GET.get('to_date') or '').strip()
    engineer_id = (request.GET.get('engineer') or '').strip()
    filled_by_id = (request.GET.get('filled_by') or '').strip()
    courier = (request.GET.get('courier') or '').strip()
    object_type = (request.GET.get('object_type') or '').strip()
    search = (request.GET.get('search') or '').strip()

    if from_date_str:
        try:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
            bookings = bookings.filter(created_at__date__gte=from_date)
        except ValueError:
            from_date = None
    if to_date_str:
        try:
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
            bookings = bookings.filter(created_at__date__lte=to_date)
        except ValueError:
            to_date = None

    if engineer_id:
        bookings = bookings.filter(engineer_id=engineer_id)
    if filled_by_id:
        bookings = bookings.filter(created_by_id=filled_by_id)
    if courier:
        bookings = bookings.filter(courier_partner=courier)
    if object_type:
        bookings = bookings.filter(object_type=object_type)
    if search:
        bookings = bookings.filter(
            Q(customer_name__icontains=search)
            | Q(delivery_name__icontains=search)
            | Q(object_serial_number__icontains=search)
            | Q(awb_number__icontains=search)
            | Q(invoice_number__icontains=search)
        )

    bookings = bookings.order_by('-created_at')[:200]

    engineers = User.objects.filter(is_active=True).order_by('first_name', 'last_name', 'username')
    couriers = list(
        LogisticBooking.objects.values_list('courier_partner', flat=True).distinct()
    )
    couriers = sorted([c for c in couriers if c])
    return render(request, 'forms/logistic_list.html', {
        'bookings': bookings,
        'engineers': engineers,
        'couriers': couriers,
        'filters': {
            'from_date': from_date_str,
            'to_date': to_date_str,
            'engineer': engineer_id,
            'filled_by': filled_by_id,
            'courier': courier,
            'object_type': object_type,
            'search': search,
        },
    })


@login_required
@logistic_access_required
def logistic_booking_edit(request, booking_id):
    booking = get_object_or_404(LogisticBooking, id=booking_id)
    from django.contrib.auth.models import User

    courier_partner_base = [
        'Safexpress',
        'Bigship',
        'Bluedart',
        'Anjani',
        'Global Cargo',
    ]

    existing_partners = list(
        LogisticBooking.objects.values_list('courier_partner', flat=True).distinct()
    )
    extra_partners = [p for p in existing_partners if p and p not in courier_partner_base]
    courier_partner_choices = courier_partner_base + sorted(extra_partners)

    capacity_options = [
        '2.2 KW', '3 KW', '3.3 KW', '4 KW', '5 KW', '6 KW', '8 KW', '10 KW', '12 KW',
        '15 KW', '18 KW', '20 KW', '25 KW', '30 KW', '33 KW', '35 KW', '40 KW',
        '50 KW', '60 KW', '75 KW', '80 KW', '100 KW', '125 KW', '136 KW'
    ]

    inverter_type_options = [
        '1PH - Hybrid',
        '3PH - Hybrid',
        '1PH - Ongrid',
        '3PH - Ongrid',
        'All In One - Hybrid',
        'Other',
    ]

    battery_model_options = [
        'RW - L 2.5 Neutral',
        'RW - M 5.3 Neutral',
        'RW - M 5.3 Pro Neutral (M6)',
        'RW - M 6.1 Neutral',
        'RW - M 6.1 B Neutral No Remark',
        'SE - G 5.1 - Pro B Neutral No Remark',
        'AI - W 5.1 Neutral',
        'AI - W 5.1 - B Neutral No Remark',
        'BOS - GM 5.1 Neutral',
        'GB - LM 4.0 Neutral',
        'GB - LB Neutral',
        'SE - G 5.3',
        'SE - G 5.3 Pro',
        'SE - F 5',
        'Other',
    ]

    pickup_status_options = [
        'Pickup Pending',
        'Picked',
    ]

    delivery_status_options = [
        'In Transit',
        'Delivered',
    ]

    engineers = User.objects.filter(is_active=True).order_by('first_name', 'last_name', 'username')

    errors = []
    if request.method == 'POST':
        data = request.POST

        def _get_value(key, required=False):
            value = (data.get(key) or '').strip()
            if required and not value:
                errors.append(f"{key.replace('_', ' ').title()} is required")
            return value

        engineer_id = _get_value('engineer_id', required=True)
        engineer = User.objects.filter(id=engineer_id).first() if engineer_id else None
        if not engineer:
            errors.append('Engineer name is required')

        booking.customer_name = _get_value('customer_name', required=True)
        booking.contact_details = _get_value('contact_details', required=True)

        booking.pickup_pincode = _get_value('pickup_pincode', required=True)
        booking.pickup_state = _get_value('pickup_state')
        booking.pickup_district = _get_value('pickup_district')
        booking.pickup_city = _get_value('pickup_city')
        booking.pickup_address = _get_value('pickup_address')

        booking.delivery_name = _get_value('delivery_name', required=True)
        booking.delivery_contact = _get_value('delivery_contact', required=True)
        booking.delivery_pincode = _get_value('delivery_pincode', required=True)
        booking.delivery_state = _get_value('delivery_state')
        booking.delivery_district = _get_value('delivery_district')
        booking.delivery_city = _get_value('delivery_city')
        booking.delivery_address = _get_value('delivery_address')

        booking.object_type = _get_value('object_type', required=True)
        booking.object_capacity = _get_value('object_capacity')
        booking.object_variant = _get_value('object_variant')
        booking.object_serial_number = _get_value('object_serial_number')
        qty_raw = _get_value('object_quantity', required=True)
        try:
            booking.object_quantity = int(qty_raw)
        except Exception:
            errors.append('Object quantity must be a valid number')

        courier_partner = _get_value('courier_partner', required=True)
        courier_partner_other = _get_value('courier_partner_other')
        if courier_partner == '__other__':
            if not courier_partner_other:
                errors.append('Other courier partner is required')
            courier_partner = courier_partner_other
        booking.courier_partner = courier_partner

        weight_raw = _get_value('shipment_weight', required=True)
        try:
            booking.shipment_weight = Decimal(weight_raw)
        except Exception:
            errors.append('Shipment weight must be a valid number')

        booking.awb_number = _get_value('awb_number')
        booking.invoice_number = _get_value('invoice_number', required=True)

        booking.pickup_status = _get_value('pickup_status')
        pickup_date_raw = _get_value('pickup_date')
        booking.delivery_status = _get_value('delivery_status')
        delivery_date_raw = _get_value('delivery_date')
        booking.remark = _get_value('remark')

        if pickup_date_raw:
            try:
                booking.pickup_date = datetime.strptime(pickup_date_raw, '%Y-%m-%d').date()
            except ValueError:
                errors.append('Pickup date must be valid')
        else:
            booking.pickup_date = None

        if delivery_date_raw:
            try:
                booking.delivery_date = datetime.strptime(delivery_date_raw, '%Y-%m-%d').date()
            except ValueError:
                errors.append('Delivery date must be valid')
        else:
            booking.delivery_date = None

        booking.engineer = engineer

        if not errors:
            booking.save()
            return redirect('forms:logistic_list')

    return render(request, 'forms/logistic_edit.html', {
        'booking': booking,
        'errors': errors,
        'engineers': engineers,
        'courier_partners': courier_partner_choices,
        'capacity_options': capacity_options,
        'inverter_type_options': inverter_type_options,
        'battery_model_options': battery_model_options,
        'pickup_status_options': pickup_status_options,
        'delivery_status_options': delivery_status_options,
    })


@login_required
def forms_data_overview(request):
    context = {
        'inwards': InwardForm.objects.order_by('-created_at')[:50],
        'outwards': OutwardForm.objects.order_by('-created_at')[:50],
        'service_reports': ServiceReportForm.objects.order_by('-created_at')[:50],
        'repairs': RepairingForm.objects.order_by('-created_at')[:50],
    }
    return render(request, 'forms/data_overview.html', context)


@login_required
def repairing_form_page(request):
    if request.method == 'POST':
        print('\n[RepairingForm] ===== POST DATA RECEIVED =====')
        print(f'All POST keys: {list(request.POST.keys())}')
        print(f'Full POST data: {dict(request.POST.lists())}')
        
        # Template uses 'Tested_by' instead of 'tested_by'
        data = request.POST.copy()
        if 'Tested_by' in data and 'tested_by' not in data:
            data['tested_by'] = data.get('Tested_by')

        form = RepairingFormForm(data)
        if form.is_valid():
            print(f'[RepairingForm] Form valid. Cleaned data keys: {list(form.cleaned_data.keys())}')
            instance = form.save(commit=False)
            # Collect multiple checkbox values for fault_problems
            problems = request.POST.getlist('fault_problems')
            instance.fault_problems = problems if problems else []
            print(f'[RepairingForm] Fault problems captured: {instance.fault_problems}')
            instance.user = request.user
            instance.save()
            print(f'[RepairingForm] ✓ Saved successfully. ID: {instance.id}')
            return redirect(f"{request.path}?success=1")
        else:
            print('[RepairingForm] Validation errors:', form.errors.as_json())
            return render(request, 'forms/repairing_form.html', {
                'form': form,
                'success': False,
                'form_errors': form.errors,
            })
    else:
        form = RepairingFormForm()
        return render(request, 'forms/repairing_form.html', {'form': form})


@login_required
def inward_form_page(request):
    if request.method == 'POST':
        print('\n[InwardForm] ===== POST DATA RECEIVED =====')
        print(f'All POST keys: {list(request.POST.keys())}')
        print(f'Full POST data: {dict(request.POST.lists())}')
        
        form = InwardFormForm(request.POST)
        if form.is_valid():
            print(f'[InwardForm] Form valid. Cleaned data keys: {list(form.cleaned_data.keys())}')
            print(f'[InwardForm] Accessories: {form.cleaned_data.get("accessories")}')
            instance = form.save(commit=False)
            instance.user = request.user
            instance.save()
            print(f'[InwardForm] ✓ Saved successfully. ID: {instance.id}')
            return redirect(f"{request.path}?success=1")
        # Log errors to console for debugging
        print('[InwardForm] Validation errors:', form.errors.as_json())
        return render(request, 'forms/inward_form.html', {
            'form': form,
            'success': False,
            'form_errors': form.errors,
        })
    form = InwardFormForm()
    return render(request, 'forms/inward_form.html', {'form': form})


@login_required
def outward_form_page(request):
    # For each inverter_id, count inward and outward occurrences
    from django.db.models import Count
    inward_counts = InwardForm.objects.values('inverter_id').annotate(count_inward=Count('id'))
    outward_counts = OutwardForm.objects.values('inverter_id_outward').annotate(count_outward=Count('id'))
    outward_map = {row['inverter_id_outward']: row['count_outward'] for row in outward_counts}
    available_serials = []
    for row in inward_counts:
        serial = row['inverter_id']
        inward = row['count_inward']
        outward = outward_map.get(serial, 0)
        if inward > outward:
            available_serials.append(serial)
    if request.method == 'POST':
        print('\n[OutwardForm] ===== POST DATA RECEIVED =====')
        print(f'All POST keys: {list(request.POST.keys())}')
        print(f'Full POST data: {dict(request.POST.lists())}')
        
        form = OutwardFormForm(request.POST)
        if form.is_valid():
            print(f'[OutwardForm] Form valid. Cleaned data keys: {list(form.cleaned_data.keys())}')
            print(f'[OutwardForm] Accessories: {form.cleaned_data.get("accessories")}')
            print(f'[OutwardForm] Control card changed: {form.cleaned_data.get("control_card_changed")}')
            instance = form.save(commit=False)
            instance.user = request.user
            instance.save()
            print(f'[OutwardForm] ✓ Saved successfully. ID: {instance.id}')
            return redirect(f"{request.path}?success=1")
        print('[OutwardForm] Validation errors:', form.errors.as_json())
        return render(request, 'forms/outward_form.html', {
            'form': form,
            'success': False,
            'form_errors': form.errors,
            'available_serials': available_serials,
        })
    form = OutwardFormForm()
    return render(request, 'forms/outward_form.html', {'form': form, 'available_serials': available_serials})


@login_required
def service_form_page(request):
    if request.method == 'POST':
        print('\n[ServiceReportForm] ===== POST DATA RECEIVED =====')
        print(f'All POST keys: {list(request.POST.keys())}')
        print(f'Full POST data: {dict(request.POST.lists())}')
        
        data = request.POST.copy()
        # Map signature hidden inputs to model fields if present
        eng_sig = data.get('engData')
        cust_sig = data.get('custData')
        print(f'[ServiceReportForm] Engineer signature data present: {bool(eng_sig)}')
        print(f'[ServiceReportForm] Customer signature data present: {bool(cust_sig)}')
        
        form = ServiceReportFormForm(data)
        if form.is_valid():
            print(f'[ServiceReportForm] Form valid. Cleaned data keys: {list(form.cleaned_data.keys())}')
            instance = form.save(commit=False)
            instance.user = request.user
            if eng_sig:
                instance.engineer_signature_data = eng_sig
            if cust_sig:
                instance.customer_signature_data = cust_sig
            
            # Parse PV data from POST parameters (pv1_voltage, pv1_current, etc.)
            pv_data = []
            for i in range(1, 9):  # PV 1-8
                pv_entry = {
                    'voltage': data.get(f'pv{i}_voltage', ''),
                    'current': data.get(f'pv{i}_current', ''),
                    'earthing': data.get(f'pv{i}_earthing', ''),
                    'panel': data.get(f'pv{i}_panel', ''),
                    'observation': data.get(f'pv{i}_observation', ''),
                }
                pv_data.append(pv_entry)
            instance.pv_data = pv_data
            print(f'[ServiceReportForm] PV data captured: {instance.pv_data}')
            
            # Parse AC data from POST parameters (ac_rn_voltage, ac_yn_voltage, etc.)
            ac_data = {
                'R-N': {
                    'voltage': data.get('ac_rn_voltage', ''),
                    'current': data.get('ac_rn_current', ''),
                    'earthing': data.get('ac_rn_earthing', ''),
                },
                'Y-N': {
                    'voltage': data.get('ac_yn_voltage', ''),
                    'current': data.get('ac_yn_current', ''),
                    'earthing': data.get('ac_yn_earthing', ''),
                },
                'B-N': {
                    'voltage': data.get('ac_bn_voltage', ''),
                    'current': data.get('ac_bn_current', ''),
                    'earthing': data.get('ac_bn_earthing', ''),
                },
                'R-Y': {
                    'voltage': data.get('ac_ry_voltage', ''),
                    'current': data.get('ac_ry_current', ''),
                    'earthing': data.get('ac_ry_earthing', ''),
                },
                'Y-B': {
                    'voltage': data.get('ac_yb_voltage', ''),
                    'current': data.get('ac_yb_current', ''),
                    'earthing': data.get('ac_yb_earthing', ''),
                },
                'B-R': {
                    'voltage': data.get('ac_br_voltage', ''),
                    'current': data.get('ac_br_current', ''),
                    'earthing': data.get('ac_br_earthing', ''),
                },
                'N-PE': {
                    'voltage': data.get('ac_npe_voltage', ''),
                    'current': data.get('ac_npe_current', ''),
                    'earthing': data.get('ac_npe_earthing', ''),
                },
            }
            instance.ac_data = ac_data
            print(f'[ServiceReportForm] AC data captured: {instance.ac_data}')
            
            instance.save()
            print(f'[ServiceReportForm] ✓ Saved successfully. ID: {instance.id}')
            return redirect(f"{request.path}?success=1")
        print('[ServiceReportForm] Validation errors:', form.errors.as_json())
        return render(request, 'forms/service_form.html', {
            'form': form,
            'success': False,
            'form_errors': form.errors,
        })
    form = ServiceReportFormForm()
    return render(request, 'forms/service_form.html', {'form': form})


@login_required
def service_report_pdf(request, report_id):
    """Generate PDF view for a service report"""
    report = get_object_or_404(ServiceReportForm, id=report_id)
    
    # Build formatted names and address
    engineer_name = f"{report.engineer_first_name} {report.engineer_last_name}"
    customer_name = f"{report.customer_first_name} {report.customer_last_name}"
    full_address = f"{report.address_street}, {report.address_city}, {report.address_state}, {report.address_zip}"
    
    # Determine filled_by display name
    if report.user:
        filled_by = report.user.get_full_name() or report.user.username
    else:
        filled_by = engineer_name
    
    context = {
        'report': report,
        'engineer_name': engineer_name,
        'customer_name': customer_name,
        'full_address': full_address,
        'filled_by': filled_by,
    }
    
    return render(request, 'forms/service_report_pdf.html', context)


@login_required
def repairing_form_pdf(request, form_id):
    """Generate PDF view for a repairing form"""
    form = get_object_or_404(RepairingForm, id=form_id)
    filled_by = form.user.get_full_name() or form.user.username if form.user else "N/A"
    context = {'form': form, 'filled_by': filled_by}
    return render(request, 'forms/repairing_form_pdf.html', context)


@login_required
def inward_form_pdf(request, form_id):
    """Generate PDF view for an inward form"""
    form = get_object_or_404(InwardForm, id=form_id)
    filled_by = form.user.get_full_name() or form.user.username if form.user else "N/A"
    context = {'form': form, 'filled_by': filled_by}
    return render(request, 'forms/inward_form_pdf.html', context)


@login_required
def outward_form_pdf(request, form_id):
    """Generate PDF view for an outward form"""
    form = get_object_or_404(OutwardForm, id=form_id)
    filled_by = form.user.get_full_name() or form.user.username if form.user else "N/A"
    context = {'form': form, 'filled_by': filled_by}
    return render(request, 'forms/outward_form_pdf.html', context)


@login_required
def hierarchy_details(request):
    # Minimal hierarchy logic only
    from collections import defaultdict
    from .models import UserProfile
    profiles = UserProfile.objects.select_related('user', 'manager').all()
    manager_map = defaultdict(list)
    for profile in profiles:
        if profile.manager:
            manager_map[profile.manager.id].append(profile)
    top_level = [profile for profile in profiles if profile.manager is None]
    context = {
        'top_level': top_level,
        'manager_map': manager_map,
    }
    return render(request, 'Include/hierarchy.html', context)


# ─────────────────────────────
# FREIGHT CALCULATOR
# ─────────────────────────────

@login_required
def freight_calculator(request):
    results = None
    error = None
    missing_pincodes = []
    
    if request.method == 'POST':
        form = FreightCalculatorForm(request.POST)
        
        # DEBUG: Log all POST data received
        print("=" * 80)
        print("FREIGHT CALCULATOR - POST DATA RECEIVED:")
        print("=" * 80)
        for key, value in request.POST.items():
            if key != 'csrfmiddlewaretoken':
                print(f"  {key}: {value}")
        print("=" * 80)
        
        if form.is_valid():
            data = form.cleaned_data
            
            # DEBUG: Log cleaned data
            print("\nCLEANED FORM DATA:")
            print("-" * 80)
            for key, value in data.items():
                print(f"  {key}: {value}")
            print("-" * 80)
            
            # Build items list from form data (multiple items support)
            items = []
            item_index = 0
            while True:
                weight_key = f'weight_kg_{item_index}' if item_index > 0 else 'weight_kg'
                length_key = f'length_cm_{item_index}' if item_index > 0 else 'length_cm'
                breadth_key = f'breadth_cm_{item_index}' if item_index > 0 else 'breadth_cm'
                height_key = f'height_cm_{item_index}' if item_index > 0 else 'height_cm'
                
                # Check if this item exists in POST data
                if weight_key in request.POST or (item_index == 0 and 'weight_kg' in data):
                    weight = float(request.POST.get(weight_key, data.get('weight_kg', 0)))
                    length = float(request.POST.get(length_key, data.get('length_cm', 0)))
                    breadth = float(request.POST.get(breadth_key, data.get('breadth_cm', 0)))
                    height = float(request.POST.get(height_key, data.get('height_cm', 0)))
                    
                    print(f"\nITEM {item_index}: weight={weight}kg, L={length}cm, B={breadth}cm, H={height}cm")
                    
                    if weight > 0 and length > 0 and breadth > 0 and height > 0:
                        items.append(ShipmentItem(
                            weight_kg=weight,
                            length_cm=length,
                            breadth_cm=breadth,
                            height_cm=height,
                        ))
                    item_index += 1
                else:
                    break
            
            print(f"\nTOTAL ITEMS CREATED: {len(items)}")
            print("=" * 80)
            
            # If no items were added, create one from main form fields
            if not items:
                items = [ShipmentItem(
                    weight_kg=data['weight_kg'],
                    length_cm=data['length_cm'],
                    breadth_cm=data['breadth_cm'],
                    height_cm=data['height_cm'],
                )]
            
            q = QuoteInput(
                from_pincode=data['from_pincode'],
                to_pincode=data['to_pincode'],
                items=items,
                reverse_pickup=data.get('reverse_pickup') or False,
                insured_value=data.get('insured_value') or 0.0,
                days_in_transit_storage=data.get('days_in_transit_storage') or 0,
                gst_mode=data.get('gst_mode', '12pct'),
                bigship_service_type=data.get('bigship_service_type', 'LTL'),
            )
            try:
                # Check if pincodes exist in database
                import pathlib
                from forms.calculator.data_loader import load_pincode_master
                base_dir = str(pathlib.Path(__file__).resolve().parents[2])
                pins_db = load_pincode_master(base_dir)
                
                if data['from_pincode'] not in pins_db:
                    missing_pincodes.append(f"From Pincode: {data['from_pincode']}")
                if data['to_pincode'] not in pins_db:
                    missing_pincodes.append(f"To Pincode: {data['to_pincode']}")
                
                results = get_all_partner_quotes(q)
                # Normalize surcharge keys so template lookups never fail
                for r in results or []:
                    if hasattr(r, 'surcharges') and isinstance(r.surcharges, dict):
                        for k in ['waybill', 'docket', 'fuel_surcharge', 'oda', 'metro_charge', 'insurance', 'valuation', 'reverse_pickup', 'handling', 'demurrage']:
                            r.surcharges.setdefault(k, 0.0)
            except Exception as exc:
                error = f"Calculator error: {exc}"
                print(f"\n❌ CALCULATOR ERROR: {exc}")
                print("=" * 80)
        else:
            error = 'Please correct the errors below.'
            print("\n❌ FORM VALIDATION ERRORS:")
            print("-" * 80)
            for field, errors in form.errors.items():
                print(f"  {field}: {', '.join(errors)}")
            print("=" * 80)
    else:
        form = FreightCalculatorForm()

    return render(request, 'forms/freight_calculator.html', {
        'form': form,
        'results': results,
        'error': error,
        'missing_pincodes': missing_pincodes,
    })


@login_required
def pincode_lookup_api(request):
    pin = request.GET.get('pin') or request.GET.get('pincode')
    if not pin:
        return JsonResponse({'ok': False, 'error': 'pin required'}, status=400)
    try:
        import pathlib
        base_dir = str(pathlib.Path(__file__).resolve().parents[2])
        db = load_pincode_master(base_dir)
        rec = lookup_pincode(db, str(pin))
        if not rec:
            return JsonResponse({'ok': False, 'error': 'not found'}, status=404)
        data = {
            'pincode': rec.pincode,
            'city': rec.city,
            'state': rec.state,
            'metro': bool(rec.is_metro) if rec.is_metro is not None else None,
            'oda': bool(rec.is_oda) if rec.is_oda is not None else None,
            'safexpress_zone': rec.safexpress_zone,
            'bluedart_region': rec.bluedart_region,
            'global_cargo_region': rec.global_cargo_region,
        }
        return JsonResponse({'ok': True, 'data': data})
    except Exception as exc:
        return JsonResponse({'ok': False, 'error': str(exc)}, status=500)


@login_required
def awb_lookup_api(request):
    query = (request.GET.get('q') or '').strip()
    base_qs = LogisticBooking.objects.exclude(awb_number='')

    if not query:
        recent_awbs = base_qs.order_by('-created_at').values_list('awb_number', flat=True)
        suggestions = []
        seen = set()
        for awb in recent_awbs:
            if awb and awb not in seen:
                seen.add(awb)
                suggestions.append(awb)
            if len(suggestions) >= 8:
                break
        return JsonResponse({'ok': True, 'suggestions': suggestions, 'detail': None})

    matches = base_qs.filter(awb_number__icontains=query)
    raw_suggestions = matches.values_list('awb_number', flat=True)
    suggestions = []
    seen = set()
    for awb in raw_suggestions:
        if awb and awb not in seen:
            seen.add(awb)
            suggestions.append(awb)
        if len(suggestions) >= 8:
            break

    exact_matches = base_qs.filter(awb_number__iexact=query).order_by('-created_at', '-id')
    items = []
    for booking in exact_matches:
        items.append({
            'booking_id': booking.id,
            'awb_number': booking.awb_number,
            'invoice_number': booking.invoice_number,
            'customer_name': booking.customer_name,
            'delivery_name': booking.delivery_name,
            'delivery_contact': booking.delivery_contact,
            'pickup_pincode': booking.pickup_pincode,
            'pickup_state': booking.pickup_state,
            'pickup_district': booking.pickup_district,
            'pickup_city': booking.pickup_city,
            'pickup_address': booking.pickup_address,
            'delivery_pincode': booking.delivery_pincode,
            'delivery_state': booking.delivery_state,
            'delivery_district': booking.delivery_district,
            'delivery_city': booking.delivery_city,
            'delivery_address': booking.delivery_address,
            'courier_partner': booking.courier_partner,
            'object_type': booking.object_type,
            'object_capacity': booking.object_capacity,
            'object_variant': booking.object_variant,
            'object_serial_number': booking.object_serial_number,
            'object_quantity': booking.object_quantity,
        })

    detail = items[0] if items else None

    return JsonResponse({'ok': True, 'suggestions': suggestions, 'detail': detail, 'items': items})

# ─────────────────────────────
# LEAVE MANAGEMENT
# ─────────────────────────────

@login_required
@login_required
def leave_home(request):
    # Check if user is a manager (has team members) or superuser
    from django.contrib.auth.models import User
    is_manager = User.objects.filter(profile__manager=request.user).exists()
    is_superuser = request.user.is_superuser
    
    # Show team attendance button if user is manager or superuser
    show_team_attendance = is_manager or is_superuser
    
    # Get holiday data for display
    from forms.models import Holiday
    fixed_holidays = Holiday.get_fixed_holidays()
    floating_holidays = Holiday.get_floating_holidays()
    
    # Count floating holidays used by employee (in days, not dates)
    floating_used_this_year = LeaveRequest.objects.filter(
        user=request.user,
        status='approved',
        start_date__year=date.today().year
    )
    
    floating_count = 0.0
    for leave in floating_used_this_year:
        check_date = leave.start_date
        while check_date <= leave.end_date:
            if Holiday.is_floating_holiday(check_date):
                # Figure out if this is a half-day or full-day
                if check_date == leave.start_date and check_date == leave.end_date:
                    # Same day - use the breakdown
                    if leave.start_breakdown == 'half' or leave.end_breakdown == 'half':
                        floating_count += 0.5
                    else:
                        floating_count += 1.0
                elif check_date == leave.start_date:
                    # Start day
                    if leave.start_breakdown == 'half':
                        floating_count += 0.5
                    else:
                        floating_count += 1.0
                elif check_date == leave.end_date:
                    # End day
                    if leave.end_breakdown == 'half':
                        floating_count += 0.5
                    else:
                        floating_count += 1.0
                else:
                    # Middle days are always full days
                    floating_count += 1.0
            check_date += timedelta(days=1)
    
    return render(request, 'leave/leave_home.html', {
        'is_manager': show_team_attendance,
        'fixed_holidays': fixed_holidays,
        'floating_holidays': floating_holidays,
        'floating_used': floating_count,
        'floating_remaining': max(3 - floating_count, 0),
    })


@login_required
def apply_leave(request):
    def _count_non_holiday_days(start_date, end_date, start_breakdown, end_breakdown):
        """Count leave days excluding holidays"""
        from forms.models import Holiday
        
        days = 0.0
        current = start_date
        
        # Count all work days (excluding Sundays)
        while current <= end_date:
            if current.weekday() != 6:  # 6 = Sunday
                # Check if it's a holiday (fixed or floating)
                if not Holiday.is_holiday(current):
                    days += 1.0
            current += timedelta(days=1)
        
        # Apply half-day adjustments only for non-holiday days
        if start_date == end_date:
            # Same day case
            if start_date.weekday() != 6 and not Holiday.is_holiday(start_date):
                if start_breakdown == 'half' and end_breakdown == 'half':
                    return 0.5
                elif start_breakdown == 'half' or end_breakdown == 'half':
                    return 0.5
                else:
                    return 1.0
            return 0.0
        else:
            # Multiple days
            if start_date.weekday() != 6 and not Holiday.is_holiday(start_date) and start_breakdown == 'half':
                days -= 0.5
            if end_date.weekday() != 6 and not Holiday.is_holiday(end_date) and end_breakdown == 'half':
                days -= 0.5
        
        return max(days, 0.0)
    
    def _remaining_paid_leave(user, year):
        """Calculate remaining paid leave excluding holidays"""
        # Get all approved paid leaves for the user this year
        approved_leaves = LeaveRequest.objects.filter(
            user=user,
            leave_type='leave',
            start_date__year=year,
            status='approved'
        )
        
        # Count only non-holiday days
        used = 0.0
        for leave in approved_leaves:
            non_holiday_days = _count_non_holiday_days(
                leave.start_date,
                leave.end_date,
                leave.start_breakdown,
                leave.end_breakdown
            )
            used += non_holiday_days
        
        return max(22.0 - used, 0.0)  # 22 total paid days per year

    if request.method == 'POST':
        leave_type = request.POST.get('leave_type')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        start_breakdown = request.POST.get('start_breakdown', 'full')
        end_breakdown = request.POST.get('end_breakdown', 'full')
        reason = request.POST.get('reason', '')

        # Coerce dates early for validation and balance checks
        try:
            sd = date.fromisoformat(start_date)
            ed = date.fromisoformat(end_date)
        except Exception:
            return render(request, 'leave/apply_leave.html', {
                'error': 'Invalid dates provided.',
                'success': False,
            })

        lr = LeaveRequest(
            user=request.user,
            leave_type=leave_type,
            start_date=sd,
            end_date=ed,
            start_breakdown=start_breakdown,
            end_breakdown=end_breakdown,
            reason=reason,
            status='pending'
        )

        requested_days = lr.compute_total_days()

        if requested_days <= 0:
            return render(request, 'leave/apply_leave.html', {
                'error': 'Requested leave days calculate to zero. Please adjust dates (Sundays are excluded).',
                'success': False,
            })

        # ===== HOLIDAY VALIDATION =====
        from forms.models import Holiday
        
        # Check for fixed holidays in the date range
        fixed_holidays_in_range = []
        current_date = sd
        while current_date <= ed:
            if Holiday.is_fixed_holiday(current_date):
                holiday = Holiday.objects.get(date=current_date, is_floating=False)
                fixed_holidays_in_range.append(holiday)
            current_date += timedelta(days=1)
        
        # Cannot request leave on fixed holidays
        if fixed_holidays_in_range:
            holiday_names = ', '.join([h.name for h in fixed_holidays_in_range])
            return render(request, 'leave/apply_leave.html', {
                'error': f'Cannot request leave on fixed holidays: {holiday_names}. These are compulsory holidays.',
                'success': False,
            })
        
        # Check floating holiday usage (if applicable)
        floating_holidays_used = LeaveRequest.objects.filter(
            user=request.user,
            status='approved',
            leave_type=leave_type,
            start_date__year=sd.year
        ).exclude(id=lr.id)  # Exclude current request if updating
        
        # Count actual days of floating holidays in approved leaves
        floating_count = 0.0
        for leave in floating_holidays_used:
            check_date = leave.start_date
            while check_date <= leave.end_date:
                if Holiday.is_floating_holiday(check_date):
                    # Calculate day fraction for this floating holiday
                    if check_date == leave.start_date and check_date == leave.end_date:
                        # Same day
                        if leave.start_breakdown == 'half' or leave.end_breakdown == 'half':
                            floating_count += 0.5
                        else:
                            floating_count += 1.0
                    elif check_date == leave.start_date:
                        # Start day
                        if leave.start_breakdown == 'half':
                            floating_count += 0.5
                        else:
                            floating_count += 1.0
                    elif check_date == leave.end_date:
                        # End day
                        if leave.end_breakdown == 'half':
                            floating_count += 0.5
                        else:
                            floating_count += 1.0
                    else:
                        # Middle days are always full days
                        floating_count += 1.0
                check_date += timedelta(days=1)
        
        # Count floating holiday days in current request
        floating_in_request = 0.0
        check_date = sd
        while check_date <= ed:
            if Holiday.is_floating_holiday(check_date):
                # Calculate day fraction for this floating holiday
                if check_date == sd and check_date == ed:
                    # Same day
                    if start_breakdown == 'half' or end_breakdown == 'half':
                        floating_in_request += 0.5
                    else:
                        floating_in_request += 1.0
                elif check_date == sd:
                    # Start day
                    if start_breakdown == 'half':
                        floating_in_request += 0.5
                    else:
                        floating_in_request += 1.0
                elif check_date == ed:
                    # End day
                    if end_breakdown == 'half':
                        floating_in_request += 0.5
                    else:
                        floating_in_request += 1.0
                else:
                    # Middle days are always full days
                    floating_in_request += 1.0
            check_date += timedelta(days=1)
        
        # Warn if already used 3 floating holidays and trying to add more
        if floating_count >= 3 and floating_in_request > 0:
            return render(request, 'leave/apply_leave.html', {
                'error': f'You have already used your 3 floating holidays for {sd.year}. You cannot request more.',
                'success': False,
            })
        
        if floating_count + floating_in_request > 3:
            return render(request, 'leave/apply_leave.html', {
                'error': f'This request would exceed your 3 floating holiday limit. Already used: {floating_count:.1f} days, Requested: {floating_in_request:.1f} days.',
                'success': False,
            })
        # ===== END HOLIDAY VALIDATION =====

        remaining = _remaining_paid_leave(request.user, sd.year)
        
        # For balance check, only count non-holiday days
        requested_non_holiday_days = _count_non_holiday_days(sd, ed, start_breakdown, end_breakdown)
        
        if leave_type == 'leave' and requested_non_holiday_days > remaining:
            return render(request, 'leave/apply_leave.html', {
                'error': f'Insufficient balance. Remaining paid leave: {remaining:.1f} days. Requested (excluding holidays): {requested_non_holiday_days:.1f} days.',
                'success': False,
            })

        # Persist after validation
        lr.save()

        # Send email notification with approve/reject links
        approve_url = request.build_absolute_uri(
            reverse('forms:approve_leave_email', kwargs={'leave_id': lr.id})
        )
        reject_url = request.build_absolute_uri(
            reverse('forms:reject_leave_email', kwargs={'leave_id': lr.id})
        )
        
        email_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; background-color: #f4f4f4; padding: 20px; }}
                .container {{ background: white; padding: 30px; border-radius: 8px; max-width: 600px; margin: 0 auto; }}
                h2 {{ color: #2c3e50; margin-bottom: 20px; }}
                .detail-row {{ padding: 10px 0; border-bottom: 1px solid #eee; }}
                .label {{ font-weight: bold; color: #555; }}
                .value {{ color: #333; margin-left: 10px; }}
                .actions {{ margin-top: 30px; text-align: center; }}
                .btn {{ display: inline-block; padding: 12px 30px; margin: 10px; text-decoration: none; 
                       border-radius: 5px; font-weight: bold; color: white; }}
                .btn-approve {{ background: #27ae60; }}
                .btn-approve:hover {{ background: #229954; }}
                .btn-reject {{ background: #e74c3c; }}
                .btn-reject:hover {{ background: #c0392b; }}
                .footer {{ margin-top: 20px; padding-top: 20px; border-top: 2px solid #eee; 
                          text-align: center; color: #777; font-size: 12px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <h2>🔔 New Leave Request</h2>
                
                <div class="detail-row">
                    <span class="label">Employee:</span>
                    <span class="value">{request.user.get_full_name() or request.user.username}</span>
                </div>
                
                <div class="detail-row">
                    <span class="label">Request Type:</span>
                    <span class="value">{lr.get_leave_type_display()}</span>
                </div>
                
                <div class="detail-row">
                    <span class="label">Start Date:</span>
                    <span class="value">{lr.start_date.strftime('%d %b %Y')} ({lr.get_start_breakdown_display()})</span>
                </div>
                
                <div class="detail-row">
                    <span class="label">End Date:</span>
                    <span class="value">{lr.end_date.strftime('%d %b %Y')} ({lr.get_end_breakdown_display()})</span>
                </div>
                
                <div class="detail-row">
                    <span class="label">Total Days:</span>
                    <span class="value">{lr.total_days} day(s)</span>
                </div>
                
                <div class="detail-row">
                    <span class="label">Reason:</span>
                    <span class="value">{lr.reason or 'No reason provided'}</span>
                </div>
                
                <div class="actions">
                    <a href="{approve_url}" class="btn btn-approve">✓ APPROVE</a>
                    <a href="{reject_url}" class="btn btn-reject">✗ REJECT</a>
                </div>
                
                <div class="footer">
                    This is an automated notification from Deye Leave Management System.<br>
                    Click the buttons above to approve or reject this request.
                </div>
            </div>
        </body>
        </html>
        """
        
        email_text = f"""
New Leave Request

Employee: {request.user.get_full_name() or request.user.username}
Type: {lr.get_leave_type_display()}
Start Date: {lr.start_date.strftime('%d %b %Y')} ({lr.get_start_breakdown_display()})
End Date: {lr.end_date.strftime('%d %b %Y')} ({lr.get_end_breakdown_display()})
Total Days: {lr.total_days}
Reason: {lr.reason or 'No reason provided'}

Approve: {approve_url}
Reject: {reject_url}
        """
        
        # Send email to HR and CC manager via SendGrid Web API (non-blocking)
        try:
            from forms.emails import send_sendgrid_email
            import threading
            
            def _send_leave_email():
                # Get manager's email if exists
                cc_list = []
                try:
                    if hasattr(request.user, 'profile') and request.user.profile.manager:
                        manager_email = request.user.profile.manager.email
                        if manager_email:
                            cc_list.append(manager_email)
                            print(f"[EMAIL] CC to manager: {manager_email}")
                except Exception as e:
                    print(f"[EMAIL] ⚠️ Could not get manager email: {e}")
                
                send_sendgrid_email(
                    ['hr@deyeindia.com'],
                    f'New {lr.get_leave_type_display()} Request - {request.user.get_full_name() or request.user.username}',
                    email_html,
                    email_text,
                    cc_emails=cc_list if cc_list else None
                )
                print(f"[EMAIL] ✅ Leave request email sent for {lr.user.username}")
            
            email_thread = threading.Thread(target=_send_leave_email, daemon=False)
            email_thread.start()
            print(f"[EMAIL] 🧵 Leave request email thread started")
        except Exception as e:
            # Log error but don't crash - leave request is already saved
            print(f"[EMAIL] ⚠️ Email send failed (non-critical): {str(e)}")

        return redirect(f"{request.path}?success=1")
    else:
        success = request.GET.get('success') == '1'
        # Show remaining balance for current year
        today = date.today()
        remaining = _remaining_paid_leave(request.user, today.year)
        return render(request, 'leave/apply_leave.html', {'success': success, 'remaining': remaining})


@login_required
def leave_status(request):
    leaves = LeaveRequest.objects.filter(user=request.user).order_by('-id')
    return render(request, 'leave/leave_status.html', {'leaves': leaves})


@login_required
def leave_history(request):
    def _count_non_holiday_days(start_date, end_date, start_breakdown, end_breakdown):
        """Count leave days excluding holidays"""
        from forms.models import Holiday
        
        days = 0.0
        current = start_date
        
        # Count all work days (excluding Sundays)
        while current <= end_date:
            if current.weekday() != 6:  # 6 = Sunday
                # Check if it's a holiday (fixed or floating)
                if not Holiday.is_holiday(current):
                    days += 1.0
            current += timedelta(days=1)
        
        # Apply half-day adjustments only for non-holiday days
        if start_date == end_date:
            # Same day case
            if start_date.weekday() != 6 and not Holiday.is_holiday(start_date):
                if start_breakdown == 'half' and end_breakdown == 'half':
                    return 0.5
                elif start_breakdown == 'half' or end_breakdown == 'half':
                    return 0.5
                else:
                    return 1.0
            return 0.0
        else:
            # Multiple days
            if start_date.weekday() != 6 and not Holiday.is_holiday(start_date) and start_breakdown == 'half':
                days -= 0.5
            if end_date.weekday() != 6 and not Holiday.is_holiday(end_date) and end_breakdown == 'half':
                days -= 0.5
        
        return max(days, 0.0)
    
    def _remaining_paid_leave(user, year):
        """Calculate remaining paid leave excluding holidays"""
        # Get all approved paid leaves for the user this year
        approved_leaves = LeaveRequest.objects.filter(
            user=user,
            leave_type='leave',
            start_date__year=year,
            status='approved'
        )
        
        # Count only non-holiday days
        used = 0.0
        for leave in approved_leaves:
            non_holiday_days = _count_non_holiday_days(
                leave.start_date,
                leave.end_date,
                leave.start_breakdown,
                leave.end_breakdown
            )
            used += non_holiday_days
        
        return max(22.0 - used, 0.0)  # 22 total paid days per year

    leaves = LeaveRequest.objects.filter(
        user=request.user
    ).exclude(status='pending').order_by('-id')

    today = date.today()
    remaining = _remaining_paid_leave(request.user, today.year)

    return render(request, 'leave/leave_history.html', {
        'leaves': leaves,
        'remaining_leave': remaining,
        'current_year': today.year,
    })


@login_required
def team_attendance(request):
    """
    Team attendance dashboard for managers and superusers.
    Superusers can see all employees, managers see only their team.
    """
    from django.contrib.auth.models import User
    from .models import LeaveRequest
    
    # Check if user is superuser - they can see all employees
    is_superuser = request.user.is_superuser
    
    if is_superuser:
        # Superusers see all employees
        team_members = User.objects.filter(is_active=True).select_related('profile')
        viewer_role = 'superuser'
    else:
        # Get team members where current user is the manager
        team_members = User.objects.filter(profile__manager=request.user).select_related('profile')
        viewer_role = 'manager'
    
    # If user is not a manager and not superuser (no team members), show message
    if not team_members.exists():
        return render(request, 'leave/team_attendance.html', {
            'is_manager': False,
            'is_superuser': False,
            'team_status': []
        })
    
    # Get today's date
    today = date.today()
    
    # Get all approved leaves that overlap with today
    leaves_today = LeaveRequest.objects.filter(
        user__in=team_members,
        status='approved',
        start_date__lte=today,
        end_date__gte=today
    ).select_related('user')
    
    # Create a dictionary of user_id -> leave for quick lookup
    leave_map = {leave.user.id: leave for leave in leaves_today}
    
    # Build status for each team member
    team_status = []
    for member in team_members:
        leave_today = leave_map.get(member.id)
        
        # Determine status
        if leave_today:
            if leave_today.leave_type == 'wfh':
                status = 'wfh'
                status_display = 'Work From Home'
                status_class = 'info'
            else:
                status = 'on_leave'
                status_display = 'On Leave'
                status_class = 'warning'
        else:
            status = 'working'
            status_display = 'Working'
            status_class = 'success'
        
        team_status.append({
            'user': member,
            'status': status,
            'status_display': status_display,
            'status_class': status_class,
            'leave': leave_today
        })
    
    # Calculate stats
    total_team = len(team_status)
    working_count = sum(1 for t in team_status if t['status'] == 'working')
    on_leave_count = sum(1 for t in team_status if t['status'] == 'on_leave')
    wfh_count = sum(1 for t in team_status if t['status'] == 'wfh')
    
    return render(request, 'leave/team_attendance.html', {
        'is_manager': True,
        'is_superuser': is_superuser,
        'viewer_role': viewer_role,
        'team_status': team_status,
        'total_team': total_team,
        'working_count': working_count,
        'on_leave_count': on_leave_count,
        'wfh_count': wfh_count,
        'today': today
    })


# ─────────────────────────────
# ADMIN LEAVE MANAGEMENT
# ─────────────────────────────

def _is_admin(user):
    return user.is_authenticated and user.is_staff


@login_required
@user_passes_test(_is_admin)
def leave_admin(request):
    from django.contrib.auth.models import User
    
    leaves = LeaveRequest.objects.select_related('user', 'status_changed_by').order_by('-applied_at')
    status_filter = request.GET.get('status')
    type_filter = request.GET.get('type')

    if status_filter:
        leaves = leaves.filter(status=status_filter)
    if type_filter:
        leaves = leaves.filter(leave_type=type_filter)

    # Add team attendance data for today
    today = date.today()
    all_employees = User.objects.filter(is_active=True).select_related('profile')
    
    # Get all approved leaves that overlap with today
    leaves_today = LeaveRequest.objects.filter(
        user__in=all_employees,
        status='approved',
        start_date__lte=today,
        end_date__gte=today
    ).select_related('user')
    
    # Create a dictionary of user_id -> leave for quick lookup
    leave_map = {leave.user.id: leave for leave in leaves_today}
    
    # Build status for each employee
    team_status = []
    for employee in all_employees:
        leave_today = leave_map.get(employee.id)
        
        # Determine status
        if leave_today:
            if leave_today.leave_type == 'wfh':
                status = 'wfh'
                status_display = 'Work From Home'
                status_class = 'info'
            else:
                status = 'on_leave'
                status_display = 'On Leave'
                status_class = 'warning'
        else:
            status = 'working'
            status_display = 'Working'
            status_class = 'success'
        
        team_status.append({
            'user': employee,
            'status': status,
            'status_display': status_display,
            'status_class': status_class,
            'leave': leave_today
        })
    
    # Calculate stats
    total_employees = len(team_status)
    working_count = sum(1 for t in team_status if t['status'] == 'working')
    on_leave_count = sum(1 for t in team_status if t['status'] == 'on_leave')
    wfh_count = sum(1 for t in team_status if t['status'] == 'wfh')

    return render(request, 'leave/admin_leave_list.html', {
        'leaves': leaves,
        'status_filter': status_filter or '',
        'type_filter': type_filter or '',
        'team_status': team_status,
        'total_employees': total_employees,
        'working_count': working_count,
        'on_leave_count': on_leave_count,
        'wfh_count': wfh_count,
        'today': today,
    })


@login_required
@user_passes_test(_is_admin)
def update_leave_status(request, leave_id):
    if request.method != 'POST':
        return HttpResponseForbidden('Invalid method')

    leave = get_object_or_404(LeaveRequest, pk=leave_id)
    new_status = request.POST.get('status')
    if new_status not in dict(LeaveRequest.LEAVE_STATUS_CHOICES):
        return HttpResponseForbidden('Invalid status')

    leave.status = new_status
    leave.status_changed_at = timezone.now()
    if request.user.is_authenticated:
        leave.status_changed_by = request.user
    leave.save()

    # Send status update email via SendGrid Web API
    if leave.user.email:
        try:
            from forms.emails import send_sendgrid_email
            import threading
            
            def _send_status_email():
                message = (
                    f"Hello {leave.user.get_username()},\n\n"
                    f"Your leave request from {leave.start_date} to {leave.end_date} is now {leave.status.title()}.\n"
                    f"Total days: {leave.total_days}."
                )
                send_sendgrid_email(
                    [leave.user.email],
                    'Leave status updated',
                    f"<p>{message.replace(chr(10), '<br>')}</p>",
                    message
                )
            
            email_thread = threading.Thread(target=_send_status_email, daemon=False)
            email_thread.start()
        except Exception as e:
            print(f"[EMAIL] Email error (non-critical): {e}")
    
    return redirect('forms:leave_admin')


@login_required
@user_passes_test(_is_admin)
def leave_export_csv(request):
    leaves = LeaveRequest.objects.select_related('user', 'status_changed_by').order_by('-applied_at')
    status_filter = request.GET.get('status')
    type_filter = request.GET.get('type')

    if status_filter:
        leaves = leaves.filter(status=status_filter)
    if type_filter:
        leaves = leaves.filter(leave_type=type_filter)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="leave_requests.csv"'
    writer = csv.writer(response)
    writer.writerow([
        'Employee', 'Type', 'Start', 'End', 'Start Portion', 'End Portion',
        'Total Days', 'Status', 'Applied At', 'Status Changed At', 'Changed By', 'Reason'
    ])
    for leave in leaves:
        writer.writerow([
            leave.user.get_username(),
            leave.get_leave_type_display(),
            leave.start_date,
            leave.end_date,
            leave.start_breakdown,
            leave.end_breakdown,
            leave.total_days,
            leave.status,
            leave.applied_at,
            leave.status_changed_at,
            leave.status_changed_by.get_username() if leave.status_changed_by else '',
            leave.reason,
        ])
    return response


@login_required
@user_passes_test(_is_admin)
def leave_admin_report(request):
    """
    Employee-wise leave statistics report.
    Shows per-employee: total requests, approved, rejected, taken, remaining.
    """
    today = date.today()
    current_year = today.year

    # Get all users with at least one leave request
    employees = User.objects.filter(
        leaverequest__isnull=False
    ).distinct().order_by('username')

    report_data = []
    for employee in employees:
        # Total requests
        total_requests = LeaveRequest.objects.filter(
            user=employee,
            leave_type='leave'
        ).count()

        # Total approved (paid leave only)
        total_approved = LeaveRequest.objects.filter(
            user=employee,
            leave_type='leave',
            status='approved'
        ).aggregate(total=Sum('total_days'))['total'] or 0.0

        # Total rejected
        total_rejected = LeaveRequest.objects.filter(
            user=employee,
            leave_type='leave',
            status='rejected'
        ).count()

        # Total taken (approved leaves this year only)
        total_taken = LeaveRequest.objects.filter(
            user=employee,
            leave_type='leave',
            status='approved',
            start_date__year=current_year,
        ).aggregate(total=Sum('total_days'))['total'] or 0.0

        # Remaining leaves (for current year)
        remaining = max(21.0 - total_taken, 0.0)

        report_data.append({
            'employee': employee,
            'username': employee.get_username(),
            'total_requests': total_requests,
            'total_approved': total_approved,
            'total_rejected': total_rejected,
            'total_taken': total_taken,
            'remaining': remaining,
        })

    return render(request, 'leave/admin_leave_report.html', {
        'report_data': report_data,
        'current_year': current_year,
    })


def approve_leave_email(request, leave_id):
    """
    Handle leave approval via email link.
    Allows both HR and the employee's manager to approve.
    """
    leave = get_object_or_404(LeaveRequest, id=leave_id)
    
    if leave.status != 'pending':
        return render(request, 'leave/email_action_result.html', {
            'success': False,
            'message': f'This leave request has already been {leave.status}.',
            'leave': leave
        })
    
    # Track who approved (if user is logged in)
    approver = None
    if request.user.is_authenticated:
        # Check if user is HR (admin/staff) or the employee's manager
        is_hr = request.user.is_staff
        is_manager = False
        try:
            if hasattr(leave.user, 'profile') and leave.user.profile.manager:
                is_manager = leave.user.profile.manager == request.user
        except:
            pass
        
        if is_hr or is_manager:
            approver = request.user
    
    leave.status = 'approved'
    leave.status_changed_at = timezone.now()
    leave.status_changed_by = approver  # Track who approved
    leave.save()
    
    # Send confirmation email to employee via SendGrid Web API
    try:
        employee_email = leave.user.email
        if employee_email:
            from forms.emails import send_sendgrid_email
            import threading
            
            def _send_approval_email():
                message = f"""
Dear {leave.user.get_full_name() or leave.user.username},

Your leave request has been APPROVED.

Details:
- Type: {leave.get_leave_type_display()}
- Start Date: {leave.start_date.strftime('%d %b %Y')} ({leave.get_start_breakdown_display()})
- End Date: {leave.end_date.strftime('%d %b %Y')} ({leave.get_end_breakdown_display()})
- Total Days: {leave.total_days}

Regards,
Deye India HR Team
                """
                send_sendgrid_email(
                    [employee_email],
                    f'Leave Request Approved - {leave.start_date.strftime("%d %b %Y")}',
                    f"<pre>{message}</pre>",
                    message
                )
            
            email_thread = threading.Thread(target=_send_approval_email, daemon=False)
            email_thread.start()
    except:
        pass
    
    return render(request, 'leave/email_action_result.html', {
        'success': True,
        'action': 'approved',
        'message': 'Leave request has been successfully approved!',
        'leave': leave
    })


def reject_leave_email(request, leave_id):
    """
    Handle leave rejection via email link.
    """
    leave = get_object_or_404(LeaveRequest, id=leave_id)
    
    if leave.status != 'pending':
        return render(request, 'leave/email_action_result.html', {
            'success': False,
            'message': f'This leave request has already been {leave.status}.',
            'leave': leave
        })
    
    leave.status = 'rejected'
    leave.status_changed_at = timezone.now()
    leave.status_changed_by = None  # Email action, no specific user
    leave.save()
    
    # Send confirmation email to employee via SendGrid Web API
    try:
        employee_email = leave.user.email
        if employee_email:
            from forms.emails import send_sendgrid_email
            import threading
            
            def _send_rejection_email():
                message = f"""
Dear {leave.user.get_full_name() or leave.user.username},

Your leave request has been REJECTED.

Details:
- Type: {leave.get_leave_type_display()}
- Start Date: {leave.start_date.strftime('%d %b %Y')} ({leave.get_start_breakdown_display()})
- End Date: {leave.end_date.strftime('%d %b %Y')} ({leave.get_end_breakdown_display()})
- Total Days: {leave.total_days}

Please contact HR for more information.

Regards,
Deye India HR Team
                """
                send_sendgrid_email(
                    [employee_email],
                    f'Leave Request Rejected - {leave.start_date.strftime("%d %b %Y")}',
                    f"<pre>{message}</pre>",
                    message
                )
            
            email_thread = threading.Thread(target=_send_rejection_email, daemon=False)
            email_thread.start()
    except:
        pass
        pass
    
    return render(request, 'leave/email_action_result.html', {
        'success': True,
        'action': 'rejected',
        'message': 'Leave request has been rejected.',
        'leave': leave
    })


# ─────────────────────────────
# USER REGISTRATION
# ─────────────────────────────

def register(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data['password'])
            user.save()
            
            # Create UserProfile with date of birth
            UserProfile.objects.create(
                user=user,
                date_of_birth=form.cleaned_data.get('date_of_birth'),
                phone=form.cleaned_data.get('phone', ''),
                department=form.cleaned_data.get('department', '')
            )
            
            # Auto-login after registration
            login(request, user)
            return redirect('forms:simple_home')
    else:
        form = UserRegistrationForm()
    
    return render(request, 'registration/register.html', {'form': form})


@login_required
def birthday_greeting(request, user_id):
    """Display birthday greeting with fireworks animation"""
    from django.contrib.auth.models import User
    user = get_object_or_404(User, id=user_id)
    return render(request, 'birthday/birthday_greeting.html', {'user': user})


# ─────────────────────────────
# EMPLOYEE DATA DASHBOARD
# ─────────────────────────────

@login_required
def employee_data_view(request):
    """Display all form responses submitted by the logged-in employee with filters"""
    from datetime import datetime, timedelta
    
    user = request.user
    user_name = user.get_full_name() or user.username
    user_email = user.email
    
    # Get tab-specific filter parameters
    repairing_date_from = request.GET.get('repairing_date_from', '')
    repairing_date_to = request.GET.get('repairing_date_to', '')
    repairing_search = request.GET.get('repairing_search', '')
    
    inward_date_from = request.GET.get('inward_date_from', '')
    inward_date_to = request.GET.get('inward_date_to', '')
    inward_search = request.GET.get('inward_search', '')
    
    outward_date_from = request.GET.get('outward_date_from', '')
    outward_date_to = request.GET.get('outward_date_to', '')
    outward_search = request.GET.get('outward_search', '')
    
    service_date_from = request.GET.get('service_date_from', '')
    service_date_to = request.GET.get('service_date_to', '')
    service_search = request.GET.get('service_search', '')
    
    stock_req_date_from = request.GET.get('stock_req_date_from', '')
    stock_req_date_to = request.GET.get('stock_req_date_to', '')
    stock_req_search = request.GET.get('stock_req_search', '')
    stock_req_status = request.GET.get('stock_req_status', '')
    
    stock_disp_date_from = request.GET.get('stock_disp_date_from', '')
    stock_disp_date_to = request.GET.get('stock_disp_date_to', '')
    stock_disp_search = request.GET.get('stock_disp_search', '')
    
    # Repairing Forms (show only forms filled by the logged-in user)
    repairing_forms = RepairingForm.objects.filter(user=user).exclude(id__isnull=True)
    if repairing_date_from:
        from_date = datetime.strptime(repairing_date_from, '%Y-%m-%d')
        repairing_forms = repairing_forms.filter(created_at__gte=from_date)
    if repairing_date_to:
        to_date = datetime.strptime(repairing_date_to, '%Y-%m-%d')
        end_date = to_date + timedelta(days=1)
        repairing_forms = repairing_forms.filter(created_at__lt=end_date)
    if repairing_search:
        repairing_forms = repairing_forms.filter(
            Q(case_number__icontains=repairing_search) |
            Q(customer_abbrev__icontains=repairing_search) |
            Q(inverter_id__icontains=repairing_search)
        )
    repairing_forms = repairing_forms.order_by('-created_at')
    
    # Inward Forms (show only forms filled by the logged-in user)
    inward_forms = InwardForm.objects.filter(user=user).exclude(id__isnull=True)
    if inward_date_from:
        from_date = datetime.strptime(inward_date_from, '%Y-%m-%d')
        inward_forms = inward_forms.filter(created_at__gte=from_date)
    if inward_date_to:
        to_date = datetime.strptime(inward_date_to, '%Y-%m-%d')
        end_date = to_date + timedelta(days=1)
        inward_forms = inward_forms.filter(created_at__lt=end_date)
    if inward_search:
        inward_forms = inward_forms.filter(
            Q(customer_name__icontains=inward_search) |
            Q(inverter_id__icontains=inward_search) |
            Q(awb_lr_number__icontains=inward_search)
        )
    inward_forms = inward_forms.order_by('-created_at')
    
    # Outward Forms (show only forms filled by the logged-in user)
    outward_forms = OutwardForm.objects.filter(user=user).exclude(id__isnull=True)
    if outward_date_from:
        from_date = datetime.strptime(outward_date_from, '%Y-%m-%d')
        outward_forms = outward_forms.filter(created_at__gte=from_date)
    if outward_date_to:
        to_date = datetime.strptime(outward_date_to, '%Y-%m-%d')
        end_date = to_date + timedelta(days=1)
        outward_forms = outward_forms.filter(created_at__lt=end_date)
    if outward_search:
        outward_forms = outward_forms.filter(
            Q(inverter_id_outward__icontains=outward_search) |
            Q(sent_to_company__icontains=outward_search) |
            Q(awb_number__icontains=outward_search)
        )
    outward_forms = outward_forms.order_by('-created_at')
    
    # Service Report Forms (show only forms filled by the logged-in user)
    service_forms = ServiceReportForm.objects.filter(
        Q(user=user) |
        Q(email=user_email) |
        Q(engineer_first_name__icontains=user.first_name, engineer_last_name__icontains=user.last_name)
    ).exclude(id__isnull=True)
    if service_date_from:
        from_date = datetime.strptime(service_date_from, '%Y-%m-%d')
        service_forms = service_forms.filter(created_at__gte=from_date)
    if service_date_to:
        to_date = datetime.strptime(service_date_to, '%Y-%m-%d')
        end_date = to_date + timedelta(days=1)
        service_forms = service_forms.filter(created_at__lt=end_date)
    if service_search:
        service_forms = service_forms.filter(
            Q(serial_number__icontains=service_search) |
            Q(customer_first_name__icontains=service_search) |
            Q(customer_last_name__icontains=service_search)
        )
    service_forms = service_forms.order_by('-created_at')
    
    # Stock Requisitions (show only forms filled by the logged-in user)
    stock_requisitions = StockRequisition.objects.filter(user=user)
    if stock_req_date_from:
        from_date = datetime.strptime(stock_req_date_from, '%Y-%m-%d')
        stock_requisitions = stock_requisitions.filter(created_at__gte=from_date)
    if stock_req_date_to:
        to_date = datetime.strptime(stock_req_date_to, '%Y-%m-%d')
        end_date = to_date + timedelta(days=1)
        stock_requisitions = stock_requisitions.filter(created_at__lt=end_date)
    if stock_req_search:
        stock_requisitions = stock_requisitions.filter(
            Q(serial_number__icontains=stock_req_search) |
            Q(component_type__icontains=stock_req_search) |
            Q(description__icontains=stock_req_search)
        )
    if stock_req_status:
        stock_requisitions = stock_requisitions.filter(status=stock_req_status)
    stock_requisitions = stock_requisitions.order_by('-created_at')
    
    # Dispatched Stock
    # Dispatched Stock (show only forms filled by the logged-in user)
    dispatched_stock = DispatchedStock.objects.filter(user=user)
    if stock_disp_date_from:
        from_date = datetime.strptime(stock_disp_date_from, '%Y-%m-%d')
        dispatched_stock = dispatched_stock.filter(created_at__gte=from_date)
    if stock_disp_date_to:
        to_date = datetime.strptime(stock_disp_date_to, '%Y-%m-%d')
        end_date = to_date + timedelta(days=1)
        dispatched_stock = dispatched_stock.filter(created_at__lt=end_date)
    if stock_disp_search:
        dispatched_stock = dispatched_stock.filter(
            Q(serial_number__icontains=stock_disp_search) |
            Q(component_type__icontains=stock_disp_search) |
            Q(tracking_number__icontains=stock_disp_search)
        )
    dispatched_stock = dispatched_stock.order_by('-created_at')
    
    # Calculate duplicate serial numbers PER FORM TYPE (not across all forms)
    from collections import Counter
    
    # Count duplicates separately for each form type (excluding stock forms)
    # Build counters by object type for each form
    repairing_inv = Counter(f.inverter_id for f in repairing_forms if f.repairing_object == 'Inverter' and f.inverter_id)
    repairing_bat = Counter(f.battery_id for f in repairing_forms if f.repairing_object == 'Battery' and f.battery_id)
    repairing_pcb = Counter(f.pcb_serial_number for f in repairing_forms if f.repairing_object == 'PCB' and f.pcb_serial_number)

    inward_inv = Counter(f.inverter_id for f in inward_forms if f.inward_object == 'Inverter' and f.inverter_id)
    inward_bat = Counter(f.battery_id for f in inward_forms if f.inward_object == 'Battery' and f.battery_id)
    inward_pcb = Counter(f.pcb_serial_number for f in inward_forms if f.inward_object == 'PCB' and f.pcb_serial_number)

    outward_inv = Counter(f.inverter_id_outward for f in outward_forms if f.outward_object == 'Inverter' and f.inverter_id_outward)
    outward_bat = Counter(f.battery_id for f in outward_forms if f.outward_object == 'Battery' and f.battery_id)
    outward_pcb = Counter(f.pcb_serial_number for f in outward_forms if f.outward_object == 'PCB' and f.pcb_serial_number)

    service_serials = Counter(service_forms.values_list('serial_number', flat=True))

    # Annotate each object with duplicate count (only within its own form type/object type)
    for form in repairing_forms:
        if form.repairing_object == 'Inverter':
            form.duplicate_count = repairing_inv.get(form.inverter_id, 0) if repairing_inv.get(form.inverter_id, 0) > 1 else 0
        elif form.repairing_object == 'Battery':
            form.duplicate_count = repairing_bat.get(form.battery_id, 0) if repairing_bat.get(form.battery_id, 0) > 1 else 0
        elif form.repairing_object == 'PCB':
            form.duplicate_count = repairing_pcb.get(form.pcb_serial_number, 0) if repairing_pcb.get(form.pcb_serial_number, 0) > 1 else 0
        else:
            form.duplicate_count = 0

    for form in inward_forms:
        if form.inward_object == 'Inverter':
            form.duplicate_count = inward_inv.get(form.inverter_id, 0) if inward_inv.get(form.inverter_id, 0) > 1 else 0
        elif form.inward_object == 'Battery':
            form.duplicate_count = inward_bat.get(form.battery_id, 0) if inward_bat.get(form.battery_id, 0) > 1 else 0
        elif form.inward_object == 'PCB':
            form.duplicate_count = inward_pcb.get(form.pcb_serial_number, 0) if inward_pcb.get(form.pcb_serial_number, 0) > 1 else 0
        else:
            form.duplicate_count = 0

    for form in outward_forms:
        if form.outward_object == 'Inverter':
            form.duplicate_count = outward_inv.get(form.inverter_id_outward, 0) if outward_inv.get(form.inverter_id_outward, 0) > 1 else 0
        elif form.outward_object == 'Battery':
            form.duplicate_count = outward_bat.get(form.battery_id, 0) if outward_bat.get(form.battery_id, 0) > 1 else 0
        elif form.outward_object == 'PCB':
            form.duplicate_count = outward_pcb.get(form.pcb_serial_number, 0) if outward_pcb.get(form.pcb_serial_number, 0) > 1 else 0
        else:
            form.duplicate_count = 0

    for form in service_forms:
        form.duplicate_count = service_serials.get(form.serial_number, 0) if service_serials.get(form.serial_number, 0) > 1 else 0
    
    context = {
        'user': user,
        'repairing_forms': repairing_forms,
        'inward_forms': inward_forms,
        'outward_forms': outward_forms,
        'service_forms': service_forms,
        'stock_requisitions': stock_requisitions,
        'dispatched_stock': dispatched_stock,
        # Repairing filters
        'repairing_date_from': repairing_date_from,
        'repairing_date_to': repairing_date_to,
        'repairing_search': repairing_search,
        # Inward filters
        'inward_date_from': inward_date_from,
        'inward_date_to': inward_date_to,
        'inward_search': inward_search,
        # Outward filters
        'outward_date_from': outward_date_from,
        'outward_date_to': outward_date_to,
        'outward_search': outward_search,
        # Service filters
        'service_date_from': service_date_from,
        'service_date_to': service_date_to,
        'service_search': service_search,
        # Stock requisition filters
        'stock_req_date_from': stock_req_date_from,
        'stock_req_date_to': stock_req_date_to,
        'stock_req_search': stock_req_search,
        'stock_req_status': stock_req_status,
        # Stock dispatched filters
        'stock_disp_date_from': stock_disp_date_from,
        'stock_disp_date_to': stock_disp_date_to,
        'stock_disp_search': stock_disp_search,
    }
    
    return render(request, 'forms/employee_data.html', context)


@login_required
def edit_employee_form(request, form_type, form_id):
    """Edit form submissions by employee"""
    user = request.user
    user_name = user.get_full_name() or user.username
    
    # Get the appropriate form based on type
    if form_type == 'repairing':
        form_obj = get_object_or_404(RepairingForm, id=form_id)
        form_class = RepairingFormForm
        redirect_url = 'forms:my_data'
    elif form_type == 'inward':
        form_obj = get_object_or_404(InwardForm, id=form_id)
        form_class = InwardFormForm
        redirect_url = 'forms:my_data'
    elif form_type == 'outward':
        form_obj = get_object_or_404(OutwardForm, id=form_id)
        form_class = OutwardFormForm
        redirect_url = 'forms:my_data'
    elif form_type == 'service':
        form_obj = get_object_or_404(ServiceReportForm, id=form_id)
        form_class = ServiceReportFormForm
        redirect_url = 'forms:my_data'
    else:
        return HttpResponseForbidden("Invalid form type")

    # Only allow editing if the logged-in user filled the form, or is a superuser
    if not (form_obj.user == user or user.is_superuser):
        return HttpResponseForbidden("You don't have permission to edit this form")
    
    if request.method == 'POST':
        form = form_class(request.POST, instance=form_obj)
        if form.is_valid():
            form.save()
            return redirect(redirect_url)
    else:
        form = form_class(instance=form_obj)
    
    return render(request, 'forms/edit_employee_form.html', {
        'form': form,
        'form_type': form_type,
        'form_obj': form_obj,
    })


@login_required
def export_employee_data(request, form_type, format):
    """Export employee data to CSV or XLSX for a specific form type"""
    import io
    from datetime import datetime, timedelta
    
    user = request.user
    user_name = user.get_full_name() or user.username
    user_email = user.email
    
    # Get forms based on type with tab-specific filters (user's own entries only, but fields/headers match Team Data export)
    def get_user_display(obj):
        if hasattr(obj, 'user') and obj.user:
            return obj.user.get_full_name() or obj.user.username
        return '-'

    if form_type == 'repairing':
        date_from = request.GET.get('repairing_date_from', '')
        date_to = request.GET.get('repairing_date_to', '')
        search_query = request.GET.get('repairing_search', '')
        queryset = RepairingForm.objects.filter(
            Q(repaired_by__icontains=user_name) | Q(tested_by__icontains=user_name)
        )
        if date_from:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            queryset = queryset.filter(created_at__gte=from_date)
        if date_to:
            to_date = datetime.strptime(date_to, '%Y-%m-%d')
            end_date = to_date + timedelta(days=1)
            queryset = queryset.filter(created_at__lt=end_date)
        if search_query:
            queryset = queryset.filter(
                Q(case_number__icontains=search_query) |
                Q(customer_abbrev__icontains=search_query) |
                Q(inverter_id__icontains=search_query)
            )
        queryset = queryset.order_by('-created_at')
        filename_prefix = "repairing_forms"
        headers = [
            'Date', 'Case Number', 'Customer', 'Object', 'Inverter ID', 'PCB Serial Number', 'PCB Specification', 'PCB Rating',
            'Inverter Spec', 'Inverter Rating', 'Battery', 'Fault Location', 'Repair Content', 'Repaired By', 'Tested By', 'Filled By'
        ]
    elif form_type == 'inward':
        date_from = request.GET.get('inward_date_from', '')
        date_to = request.GET.get('inward_date_to', '')
        search_query = request.GET.get('inward_search', '')
        queryset = InwardForm.objects.filter(
            Q(email=user_email) | Q(received_by__icontains=user_name)
        )
        if date_from:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            queryset = queryset.filter(created_at__gte=from_date)
        if date_to:
            to_date = datetime.strptime(date_to, '%Y-%m-%d')
            end_date = to_date + timedelta(days=1)
            queryset = queryset.filter(created_at__lt=end_date)
        if search_query:
            queryset = queryset.filter(
                Q(customer_name__icontains=search_query) |
                Q(inverter_id__icontains=search_query) |
                Q(awb_lr_number__icontains=search_query)
            )
        queryset = queryset.order_by('-created_at')
        filename_prefix = "inward_forms"
        headers = [
            'Date', 'Inward Object', 'Customer Abbreviation', 'Customer Name', 'Inverter ID', 'Battery ID', 'PCB Serial Number',
            'Inverter Specification', 'Inverter Ratings', 'Battery Model', 'No of MPPT', 'Current/MPPT', 'PCB Quantity',
            'Received From Location', 'Received From District', 'Received From State', 'Pincode', 'Received By', 'Reason',
            'Transportation Mode', 'AWB Number', 'Filled By'
        ]
    elif form_type == 'outward':
        date_from = request.GET.get('outward_date_from', '')
        date_to = request.GET.get('outward_date_to', '')
        search_query = request.GET.get('outward_search', '')
        queryset = OutwardForm.objects.filter(
            Q(sent_by__icontains=user_name) | Q(approved_by__icontains=user_name)
        )
        if date_from:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            queryset = queryset.filter(created_at__gte=from_date)
        if date_to:
            to_date = datetime.strptime(date_to, '%Y-%m-%d')
            end_date = to_date + timedelta(days=1)
            queryset = queryset.filter(created_at__lt=end_date)
        if search_query:
            queryset = queryset.filter(
                Q(inverter_id_outward__icontains=search_query) |
                Q(sent_to_company__icontains=search_query) |
                Q(awb_number__icontains=search_query)
            )
        queryset = queryset.order_by('-created_at')
        filename_prefix = "outward_forms"
        headers = [
            'Date', 'Outward Object', 'Inverter ID (Outward)', 'Inverter Spec', 'Inverter Rating', 'Battery', 'Battery ID',
            'Sent To Company', 'Sent To Address', 'Sent To District', 'Sent To State', 'Pincode', 'Sent By', 'Approved By',
            'Control Card Changed', 'New Serial Number', 'Inverter ID (Inward)', 'Inverter Replaced', 'Delivered Through',
            'AWB Number', 'Remarks', 'Filled By'
        ]
    elif form_type == 'service':
        date_from = request.GET.get('service_date_from', '')
        date_to = request.GET.get('service_date_to', '')
        search_query = request.GET.get('service_search', '')
        queryset = ServiceReportForm.objects.filter(
            Q(email=user_email) | Q(engineer_first_name__icontains=user.first_name, engineer_last_name__icontains=user.last_name)
        )
        if date_from:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            queryset = queryset.filter(created_at__gte=from_date)
        if date_to:
            to_date = datetime.strptime(date_to, '%Y-%m-%d')
            end_date = to_date + timedelta(days=1)
            queryset = queryset.filter(created_at__lt=end_date)
        if search_query:
            queryset = queryset.filter(
                Q(serial_number__icontains=search_query) |
                Q(customer_first_name__icontains=search_query) |
                Q(customer_last_name__icontains=search_query)
            )
        queryset = queryset.order_by('-created_at')
        filename_prefix = "service_reports"
        headers = [
            'Engineer', 'Customer', 'Product Type', 'Serial Number', 'Service Date', 'Address', 'Battery Type', 'Battery Make',
            'Battery Voltage', 'PV Capacity (kW)', 'AC Cable Size', 'Physical Observation', 'Actual Work Done', 'Cause of Failure',
            'Conclusion', 'Customer Ratings', 'Suggestions', 'Created', 'Filled By'
        ]
    
    # Export based on format
    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        for obj in queryset:
            if form_type == 'repairing':
                writer.writerow([
                    obj.created_at.strftime('%d %b %Y'), obj.case_number, obj.customer_abbrev, obj.repairing_object, obj.inverter_id,
                    obj.pcb_serial_number, getattr(obj, 'pcb_specification', ''), getattr(obj, 'pcb_rating', ''), obj.inverter_spec, obj.inverter_rating, obj.battery,
                    obj.fault_location, obj.repair_content, obj.repaired_by, obj.tested_by, get_user_display(obj)
                ])
            elif form_type == 'inward':
                writer.writerow([
                    obj.created_at.strftime('%d %b %Y'), obj.inward_object, obj.customer_abbrev, obj.customer_name, obj.inverter_id, obj.battery_id,
                    obj.pcb_serial_number, obj.inverter_specs, obj.inverter_ratings, obj.battery, obj.no_of_mppt, obj.current_mppt, obj.pcb_quantity,
                    obj.received_from_location, obj.received_from_district, obj.received_from_state, obj.pincode, obj.received_by, obj.reason,
                    obj.transportation_mode, obj.awb_lr_number, get_user_display(obj)
                ])
            elif form_type == 'outward':
                writer.writerow([
                    obj.created_at.strftime('%d %b %Y'), obj.outward_object, obj.inverter_id_outward, obj.inverter_spec, obj.inverter_rating, obj.battery,
                    obj.battery_id, obj.sent_to_company, obj.sent_to_address, obj.sent_to_district, obj.sent_to_state, obj.pincode, obj.sent_by,
                    obj.approved_by, obj.control_card_changed, obj.new_serial_number, obj.inverter_id_inward, obj.inverter_replaced,
                    obj.delivered_through, obj.awb_number, obj.remarks, get_user_display(obj)
                ])
            elif form_type == 'service':
                writer.writerow([
                    f"{obj.engineer_first_name} {obj.engineer_last_name}", f"{obj.customer_first_name} {obj.customer_last_name}", obj.product_type,
                    obj.serial_number, obj.date_of_service.strftime('%d %b %Y') if obj.date_of_service else 'N/A',
                    f"{obj.address_street}, {obj.address_city}, {obj.address_state}, {obj.address_zip}", obj.battery_type, obj.battery_make,
                    obj.battery_voltage, obj.pv_capacity_kw, obj.ac_cable_size, obj.physical_observation, obj.actual_work_done,
                    obj.cause_of_failure, obj.conclusion, obj.customer_ratings, obj.suggestions, obj.created_at.strftime('%d %b %Y'), get_user_display(obj)
                ])
        return response
    elif format == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = filename_prefix.replace('_', ' ').title()
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for obj in queryset:
            if form_type == 'repairing':
                ws.append([
                    obj.created_at.strftime('%d %b %Y'), obj.case_number, obj.customer_abbrev, obj.repairing_object, obj.inverter_id,
                    obj.pcb_serial_number, getattr(obj, 'pcb_specification', ''), getattr(obj, 'pcb_rating', ''), obj.inverter_spec, obj.inverter_rating, obj.battery,
                    obj.fault_location, obj.repair_content, obj.repaired_by, obj.tested_by, get_user_display(obj)
                ])
            elif form_type == 'inward':
                ws.append([
                    obj.created_at.strftime('%d %b %Y'), obj.inward_object, obj.customer_abbrev, obj.customer_name, obj.inverter_id, obj.battery_id,
                    obj.pcb_serial_number, obj.inverter_specs, obj.inverter_ratings, obj.battery, obj.no_of_mppt, obj.current_mppt, obj.pcb_quantity,
                    obj.received_from_location, obj.received_from_district, obj.received_from_state, obj.pincode, obj.received_by, obj.reason,
                    obj.transportation_mode, obj.awb_lr_number, get_user_display(obj)
                ])
            elif form_type == 'outward':
                ws.append([
                    obj.created_at.strftime('%d %b %Y'), obj.outward_object, obj.inverter_id_outward, obj.inverter_spec, obj.inverter_rating, obj.battery,
                    obj.battery_id, obj.sent_to_company, obj.sent_to_address, obj.sent_to_district, obj.sent_to_state, obj.pincode, obj.sent_by,
                    obj.approved_by, obj.control_card_changed, obj.new_serial_number, obj.inverter_id_inward, obj.inverter_replaced,
                    obj.delivered_through, obj.awb_number, obj.remarks, get_user_display(obj)
                ])
            elif form_type == 'service':
                ws.append([
                    f"{obj.engineer_first_name} {obj.engineer_last_name}", f"{obj.customer_first_name} {obj.customer_last_name}", obj.product_type,
                    obj.serial_number, obj.date_of_service.strftime('%d %b %Y') if obj.date_of_service else 'N/A',
                    f"{obj.address_street}, {obj.address_city}, {obj.address_state}, {obj.address_zip}", obj.battery_type, obj.battery_make,
                    obj.battery_voltage, obj.pv_capacity_kw, obj.ac_cable_size, obj.physical_observation, obj.actual_work_done,
                    obj.cause_of_failure, obj.conclusion, obj.customer_ratings, obj.suggestions, obj.created_at.strftime('%d %b %Y'), get_user_display(obj)
                ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response
    else:
        return HttpResponse("Invalid format. Use 'csv' or 'xlsx'.", status=400)
    """Export employee data to CSV or XLSX"""
    import io
    from datetime import datetime
    
    user = request.user
    user_name = user.get_full_name() or user.username
    user_email = user.email
    
    # Get filter parameters
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    
    # Get all forms (same filtering logic as main view)
    repairing_forms = RepairingForm.objects.filter(
        Q(repaired_by__icontains=user_name) | 
        Q(tested_by__icontains=user_name)
    )
    inward_forms = InwardForm.objects.filter(
        Q(email=user_email) | 
        Q(received_by__icontains=user_name)
    )
    outward_forms = OutwardForm.objects.filter(
        Q(sent_by__icontains=user_name) | 
        Q(approved_by__icontains=user_name)
    )
    service_forms = ServiceReportForm.objects.filter(
        Q(email=user_email) |
        Q(engineer_first_name__icontains=user.first_name, engineer_last_name__icontains=user.last_name)
    )
    stock_requisitions = StockRequisition.objects.filter(
        manager_name__icontains=user_name
    )
    dispatched_stock = DispatchedStock.objects.filter(
        Q(dispatched_by__icontains=user_name) |
        Q(engineer_name__icontains=user_name)
    )
    
    # Apply filters
    if date_from:
        repairing_forms = repairing_forms.filter(created_at__gte=date_from)
        inward_forms = inward_forms.filter(created_at__gte=date_from)
        outward_forms = outward_forms.filter(created_at__gte=date_from)
        service_forms = service_forms.filter(created_at__gte=date_from)
        stock_requisitions = stock_requisitions.filter(created_at__gte=date_from)
        dispatched_stock = dispatched_stock.filter(created_at__gte=date_from)
    
    if date_to:
        repairing_forms = repairing_forms.filter(created_at__lte=date_to + ' 23:59:59')
        inward_forms = inward_forms.filter(created_at__lte=date_to + ' 23:59:59')
        outward_forms = outward_forms.filter(created_at__lte=date_to + ' 23:59:59')
        service_forms = service_forms.filter(created_at__lte=date_to + ' 23:59:59')
        stock_requisitions = stock_requisitions.filter(created_at__lte=date_to + ' 23:59:59')
        dispatched_stock = dispatched_stock.filter(created_at__lte=date_to + ' 23:59:59')
    
    if search_query:
        repairing_forms = repairing_forms.filter(
            Q(case_number__icontains=search_query) |
            Q(customer_abbrev__icontains=search_query) |
            Q(inverter_id__icontains=search_query)
        )
        inward_forms = inward_forms.filter(
            Q(customer_name__icontains=search_query) |
            Q(inverter_id__icontains=search_query) |
            Q(awb_lr_number__icontains=search_query)
        )
        outward_forms = outward_forms.filter(
            Q(inverter_id_outward__icontains=search_query) |
            Q(sent_to_company__icontains=search_query) |
            Q(awb_number__icontains=search_query)
        )
        service_forms = service_forms.filter(
            Q(serial_number__icontains=search_query) |
            Q(customer_first_name__icontains=search_query) |
            Q(customer_last_name__icontains=search_query)
        )
        stock_requisitions = stock_requisitions.filter(
            Q(serial_number__icontains=search_query) |
            Q(component_type__icontains=search_query) |
            Q(description__icontains=search_query)
        )
        dispatched_stock = dispatched_stock.filter(
            Q(serial_number__icontains=search_query) |
            Q(component_type__icontains=search_query) |
            Q(tracking_number__icontains=search_query)
        )
    
    if status_filter:
        stock_requisitions = stock_requisitions.filter(status=status_filter)
    
    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="my_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        
        # Repairing Forms
        writer.writerow(['=== REPAIRING FORMS ==='])
        writer.writerow(['Case Number', 'Customer', 'Object', 'Inverter ID', 'Repaired By', 'Tested By', 'Date', 'Status'])
        for form in repairing_forms:
            writer.writerow([
                form.case_number,
                form.customer_abbrev,
                form.repairing_object,
                form.inverter_id,
                form.repaired_by,
                form.tested_by,
                form.repaired_on_date,
                form.remark or 'Completed'
            ])
        writer.writerow([])
        
        # Inward Forms
        writer.writerow(['=== INWARD FORMS ==='])
        writer.writerow(['Customer', 'Inverter ID', 'From Location', 'Received By', 'Reason', 'AWB/LR', 'Date'])
        for form in inward_forms:
            writer.writerow([
                form.customer_name,
                form.inverter_id,
                f"{form.received_from_location}, {form.received_from_state}",
                form.received_by,
                form.reason,
                form.awb_lr_number,
                               form.created_at.strftime('%Y-%m-%d')
            ])
        writer.writerow([])
        
        # Outward Forms
        writer.writerow(['=== OUTWARD FORMS ==='])
        writer.writerow(['Inverter ID', 'Sent To', 'Location', 'Sent By', 'Approved By', 'AWB Number', 'Date'])
        for form in outward_forms:
            writer.writerow([
                form.inverter_id_outward,
                form.sent_to_company,
                f"{form.sent_to_district}, {form.sent_to_state}",
                form.sent_by,
                form.approved_by,
                form.awb_number,
                form.created_at.strftime('%Y-%m-%d')
            ])
        writer.writerow([])
        
        # Service Reports
        writer.writerow(['=== SERVICE REPORTS ==='])
        writer.writerow(['Engineer', 'Customer', 'Product Type', 'Serial Number', 'Service Date', 'Location', 'Created'])
        for form in service_forms:
            writer.writerow([
                f"{form.engineer_first_name} {form.engineer_last_name}",
                f"{form.customer_first_name} {form.customer_last_name}",
                form.product_type,
                form.serial_number,
                form.date_of_service.strftime('%Y-%m-%d') if form.date_of_service else 'N/A',
                f"{form.address_city}, {form.address_state}",
                form.created_at.strftime('%Y-%m-%d')
            ])
        writer.writerow([])
        
        # Stock Requisitions
        writer.writerow(['=== STOCK REQUISITIONS ==='])
        writer.writerow(['Serial Number', 'Component Type', 'Description', 'Qty Required', 'Qty Approved', 'Required To', 'Status', 'Created'])
        for req in stock_requisitions:
            writer.writerow([
                req.serial_number,
                req.component_type,
                req.description,
                req.quantity_required,
                req.approved_quantity or '—',
                req.required_to,
                req.get_status_display(),
                req.created_at.strftime('%Y-%m-%d')
            ])
        writer.writerow([])
        
        # Dispatched Stock
        writer.writerow(['=== STOCK DISPATCHED ==='])
        writer.writerow(['Serial Number', 'Component', 'Quantity', 'Engineer', 'Location', 'Courier', 'Tracking', 'Dispatch Date'])
        for stock in dispatched_stock:
            writer.writerow([
                stock.serial_number,
                stock.component_type,
                stock.quantity_dispatched,
                stock.engineer_name,
                stock.dispatch_location,
                stock.courier_name,
                stock.tracking_number,
                stock.dispatch_date.strftime('%Y-%m-%d')
            ])
        
        return response
    
    elif format == 'xlsx':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return HttpResponse("openpyxl library not installed. Please install it using: pip install openpyxl", status=500)
        
        wb = Workbook()
        ws = wb.active
        ws.title = filename_prefix.replace('_', ' ').title()
        
        # Set headers based on form type
        if form_type == 'repairing':
            headers = ['Case Number', 'Customer', 'Object', 'Inverter ID', 'Repaired By', 'Tested By', 'Date', 'Status']
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
            for form in queryset:
                ws.append([
                    form.case_number,
                    form.customer_abbrev,
                    form.repairing_object,
                    form.inverter_id,
                    form.repaired_by,
                    form.tested_by,
                    form.repaired_on_date,
                    form.remark or 'Completed'
                ])
        
        elif form_type == 'inward':
            headers = ['Customer', 'Inverter ID', 'From Location', 'Received By', 'Reason', 'AWB/LR', 'Date']
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
            for form in queryset:
                ws.append([
                    form.customer_name,
                    form.inverter_id,
                    f"{form.received_from_location}, {form.received_from_state}",
                    form.received_by,
                    form.reason,
                    form.awb_lr_number,
                    form.created_at.strftime('%Y-%m-%d')
                ])
        
        elif form_type == 'outward':
            headers = ['Inverter ID', 'Sent To', 'Location', 'Sent By', 'Approved By', 'AWB Number', 'Date']
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
            for form in queryset:
                ws.append([
                    form.inverter_id_outward,
                    form.sent_to_company,
                    f"{form.sent_to_district}, {form.sent_to_state}",
                    form.sent_by,
                    form.approved_by,
                    form.awb_number,
                    form.created_at.strftime('%Y-%m-%d')
                ])
        
        elif form_type == 'service':
            headers = ['Engineer', 'Customer', 'Product Type', 'Serial Number', 'Service Date', 'Location', 'Created']
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
            for form in queryset:
                ws.append([
                    f"{form.engineer_first_name} {form.engineer_last_name}",
                    f"{form.customer_first_name} {form.customer_last_name}",
                    form.product_type,
                    form.serial_number,
                    form.date_of_service.strftime('%Y-%m-%d') if form.date_of_service else 'N/A',
                    f"{form.address_city}, {form.address_state}",
                    form.created_at.strftime('%Y-%m-%d')
                ])
        
        elif form_type == 'stock-req':
            headers = ['Serial Number', 'Component Type', 'Description', 'Qty Required', 'Qty Approved', 'Required To', 'Status', 'Created']
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
            for req in queryset:
                ws.append([
                    req.serial_number,
                    req.component_type,
                    req.description,
                    req.quantity_required,
                    req.approved_quantity or '—',
                    req.required_to,
                    req.get_status_display(),
                    req.created_at.strftime('%Y-%m-%d')
                ])
        
        elif form_type == 'stock-disp':
            headers = ['Serial Number', 'Component', 'Quantity', 'Engineer', 'Location', 'Courier', 'Tracking', 'Dispatch Date']
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
            for stock in queryset:
                ws.append([
                    stock.serial_number,
                    stock.component_type,
                    stock.quantity_dispatched,
                    stock.engineer_name,
                    stock.dispatch_location,
                    stock.courier_name,
                    stock.tracking_number,
                    stock.dispatch_date.strftime('%Y-%m-%d')
                ])
        
        # Save to response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        return response
    
    else:
        return HttpResponse("Invalid format. Use 'csv' or 'xlsx'.", status=400)

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseForbidden
from django.db.models import Q
from django.contrib.auth.models import User
from .models import (RepairingForm, InwardForm, OutwardForm, 
                     ServiceReportForm, StockRequisition, DispatchedStock)
import csv


@login_required
def team_data_view(request):
    """Display all form responses from team members (for managers) or all data (for superusers)"""
    from datetime import datetime, timedelta
    from collections import Counter
    
    user = request.user
    
    # Check if user is superuser or manager
    if user.is_superuser:
        # Superuser sees all data
        team_users = User.objects.all()
        view_mode = 'all'
    elif hasattr(user, 'profile') and user.profile.get_team_members().exists():
        # Manager sees their team's data and their own data
        team_users = user.profile.get_team_members()
        if user not in team_users:
            team_users = team_users | User.objects.filter(pk=user.pk)
        view_mode = 'team'
    else:
        # Regular user - redirect to personal data view
        from django.contrib import messages
        messages.warning(request, "You don't have permission to view team data. Showing your personal data instead.")
        return redirect('forms:my_data')
    
    # Get team members names and emails
    team_names = [u.get_full_name() or u.username for u in team_users]
    team_emails = list(team_users.values_list('email', flat=True))
    
    # Get tab-specific filter parameters
    repairing_date_from = request.GET.get('repairing_date_from', '')
    repairing_date_to = request.GET.get('repairing_date_to', '')
    repairing_search = request.GET.get('repairing_search', '')
    
    inward_date_from = request.GET.get('inward_date_from', '')
    inward_date_to = request.GET.get('inward_date_to', '')
    inward_search = request.GET.get('inward_search', '')
    
    outward_date_from = request.GET.get('outward_date_from', '')
    outward_date_to = request.GET.get('outward_date_to', '')
    outward_search = request.GET.get('outward_search', '')
    
    service_date_from = request.GET.get('service_date_from', '')
    service_date_to = request.GET.get('service_date_to', '')
    service_search = request.GET.get('service_search', '')
    
    stock_req_date_from = request.GET.get('stock_req_date_from', '')
    stock_req_date_to = request.GET.get('stock_req_date_to', '')
    stock_req_search = request.GET.get('stock_req_search', '')
    stock_req_status = request.GET.get('stock_req_status', '')
    
    stock_disp_date_from = request.GET.get('stock_disp_date_from', '')
    stock_disp_date_to = request.GET.get('stock_disp_date_to', '')
    stock_disp_search = request.GET.get('stock_disp_search', '')
    
    # Build queries for team data
    # For superusers, get all records. For managers, filter by team members
    if view_mode == 'all':
        # Superuser - get all records
        repairing_forms = RepairingForm.objects.all()
    else:
        # Manager - filter by team members (user who filled the form or name match)
        repairing_q = Q(user__in=team_users)
        for name in team_names:
            repairing_q |= Q(repaired_by__icontains=name) | Q(tested_by__icontains=name)
        repairing_forms = RepairingForm.objects.filter(repairing_q)
    if repairing_date_from:
        from_date = datetime.strptime(repairing_date_from, '%Y-%m-%d')
        repairing_forms = repairing_forms.filter(created_at__gte=from_date)
    if repairing_date_to:
        to_date = datetime.strptime(repairing_date_to, '%Y-%m-%d')
        end_date = to_date + timedelta(days=1)
        repairing_forms = repairing_forms.filter(created_at__lt=end_date)
    if repairing_search:
        repairing_forms = repairing_forms.filter(
            Q(case_number__icontains=repairing_search) |
            Q(customer_abbrev__icontains=repairing_search) |
            Q(inverter_id__icontains=repairing_search)
        )
    repairing_forms = repairing_forms.order_by('-created_at')
    
    # Inward Forms
    if user.is_superuser:
        inward_forms = InwardForm.objects.all()
    elif view_mode == 'all':
        inward_forms = InwardForm.objects.all()
    else:
        inward_q = Q(user__in=team_users)
        for email in team_emails:
            inward_q |= Q(email=email)
        for name in team_names:
            inward_q |= Q(received_by__icontains=name)
        inward_forms = InwardForm.objects.filter(inward_q)
    if inward_date_from:
        from_date = datetime.strptime(inward_date_from, '%Y-%m-%d')
        inward_forms = inward_forms.filter(created_at__gte=from_date)
    if inward_date_to:
        to_date = datetime.strptime(inward_date_to, '%Y-%m-%d')
        end_date = to_date + timedelta(days=1)
        inward_forms = inward_forms.filter(created_at__lt=end_date)
    if inward_search:
        inward_forms = inward_forms.filter(
            Q(customer_name__icontains=inward_search) |
            Q(inverter_id__icontains=inward_search) |
            Q(awb_lr_number__icontains=inward_search)
        )
    inward_forms = inward_forms.order_by('-created_at')
    
    # Outward Forms
    if view_mode == 'all':
        outward_forms = OutwardForm.objects.all()
    else:
        outward_q = Q(user__in=team_users)
        for name in team_names:
            outward_q |= Q(sent_by__icontains=name) | Q(approved_by__icontains=name)
        outward_forms = OutwardForm.objects.filter(outward_q)
    if outward_date_from:
        from_date = datetime.strptime(outward_date_from, '%Y-%m-%d')
        outward_forms = outward_forms.filter(created_at__gte=from_date)
    if outward_date_to:
        to_date = datetime.strptime(outward_date_to, '%Y-%m-%d')
        end_date = to_date + timedelta(days=1)
        outward_forms = outward_forms.filter(created_at__lt=end_date)
    if outward_search:
        outward_forms = outward_forms.filter(
            Q(inverter_id_outward__icontains=outward_search) |
            Q(sent_to_company__icontains=outward_search) |
            Q(awb_number__icontains=outward_search)
        )
    outward_forms = outward_forms.order_by('-created_at')
    
    # Service Report Forms
    if view_mode == 'all':
        service_forms = ServiceReportForm.objects.all()
    else:
        service_q = Q(user__in=team_users)
        for email in team_emails:
            service_q |= Q(email=email)
        for user_obj in team_users:
            service_q |= Q(engineer_first_name__icontains=user_obj.first_name, engineer_last_name__icontains=user_obj.last_name)
        service_forms = ServiceReportForm.objects.filter(service_q)
    if service_date_from:
        from_date = datetime.strptime(service_date_from, '%Y-%m-%d')
        service_forms = service_forms.filter(created_at__gte=from_date)
    if service_date_to:
        to_date = datetime.strptime(service_date_to, '%Y-%m-%d')
        end_date = to_date + timedelta(days=1)
        service_forms = service_forms.filter(created_at__lt=end_date)
    if service_search:
        service_forms = service_forms.filter(
            Q(serial_number__icontains=service_search) |
            Q(customer_first_name__icontains=service_search) |
            Q(customer_last_name__icontains=service_search)
        )
    service_forms = service_forms.order_by('-created_at')
    
    # Stock Requisitions
    if view_mode == 'all':
        stock_requisitions = StockRequisition.objects.all()
    else:
        stock_req_q = Q(user__in=team_users)
        for name in team_names:
            stock_req_q |= Q(manager_name__icontains=name)
        stock_requisitions = StockRequisition.objects.filter(stock_req_q)
    if stock_req_date_from:
        from_date = datetime.strptime(stock_req_date_from, '%Y-%m-%d')
        stock_requisitions = stock_requisitions.filter(created_at__gte=from_date)
    if stock_req_date_to:
        to_date = datetime.strptime(stock_req_date_to, '%Y-%m-%d')
        end_date = to_date + timedelta(days=1)
        stock_requisitions = stock_requisitions.filter(created_at__lt=end_date)
    if stock_req_search:
        stock_requisitions = stock_requisitions.filter(
            Q(serial_number__icontains=stock_req_search) |
            Q(component_type__icontains=stock_req_search) |
            Q(description__icontains=stock_req_search)
        )
    if stock_req_status:
        stock_requisitions = stock_requisitions.filter(status=stock_req_status)
    stock_requisitions = stock_requisitions.order_by('-created_at')
    
    # Dispatched Stock
    if view_mode == 'all':
        dispatched_stock = DispatchedStock.objects.all()
    else:
        dispatched_q = Q(user__in=team_users)
        for name in team_names:
            dispatched_q |= Q(dispatched_by__icontains=name) | Q(engineer_name__icontains=name)
        dispatched_stock = DispatchedStock.objects.filter(dispatched_q)
    if stock_disp_date_from:
        from_date = datetime.strptime(stock_disp_date_from, '%Y-%m-%d')
        dispatched_stock = dispatched_stock.filter(created_at__gte=from_date)
    if stock_disp_date_to:
        to_date = datetime.strptime(stock_disp_date_to, '%Y-%m-%d')
        end_date = to_date + timedelta(days=1)
        dispatched_stock = dispatched_stock.filter(created_at__lt=end_date)
    if stock_disp_search:
        dispatched_stock = dispatched_stock.filter(
            Q(serial_number__icontains=stock_disp_search) |
            Q(component_type__icontains=stock_disp_search) |
            Q(tracking_number__icontains=stock_disp_search)
        )
    dispatched_stock = dispatched_stock.order_by('-created_at')
    
    # Calculate duplicate serial numbers PER FORM TYPE (not across all forms)
    from collections import Counter
    
    # Count duplicates separately for each form type (excluding stock forms)
    repairing_serials = Counter(repairing_forms.values_list('inverter_id', flat=True))
    inward_serials = Counter(inward_forms.values_list('inverter_id', flat=True))
    outward_serials = Counter(outward_forms.values_list('inverter_id_outward', flat=True))
    service_serials = Counter(service_forms.values_list('serial_number', flat=True))
    
    # Annotate each object with duplicate count (only within its own form type)
    for form in repairing_forms:
        form.duplicate_count = repairing_serials.get(form.inverter_id, 0) if repairing_serials.get(form.inverter_id, 0) > 1 else 0
    
    for form in inward_forms:
        form.duplicate_count = inward_serials.get(form.inverter_id, 0) if inward_serials.get(form.inverter_id, 0) > 1 else 0
    
    for form in outward_forms:
        form.duplicate_count = outward_serials.get(form.inverter_id_outward, 0) if outward_serials.get(form.inverter_id_outward, 0) > 1 else 0
    
    for form in service_forms:
        form.duplicate_count = service_serials.get(form.serial_number, 0) if service_serials.get(form.serial_number, 0) > 1 else 0
    
    # Stock Inverters: InwardForm entries whose inverter_id is NOT in OutwardForm, filtered by team
    outward_ids = set(OutwardForm.objects.values_list('inverter_id_outward', flat=True))
    if view_mode == 'all':
        stock_inverters_qs = InwardForm.objects.exclude(inverter_id__in=outward_ids)
    else:
        stock_inverters_qs = InwardForm.objects.exclude(inverter_id__in=outward_ids).filter(user__in=team_users)
    # Apply Stock Inverters tab filters (date/search)
    stock_inv_date_from = request.GET.get('stock_inv_date_from', '')
    stock_inv_date_to = request.GET.get('stock_inv_date_to', '')
    stock_inv_search = request.GET.get('stock_inv_search', '')
    if stock_inv_date_from:
        from_date = datetime.strptime(stock_inv_date_from, '%Y-%m-%d')
        stock_inverters_qs = stock_inverters_qs.filter(created_at__gte=from_date)
    if stock_inv_date_to:
        to_date = datetime.strptime(stock_inv_date_to, '%Y-%m-%d')
        end_date = to_date + timedelta(days=1)
        stock_inverters_qs = stock_inverters_qs.filter(created_at__lt=end_date)
    if stock_inv_search:
        stock_inverters_qs = stock_inverters_qs.filter(
            Q(inverter_id__icontains=stock_inv_search) |
            Q(customer_name__icontains=stock_inv_search) |
            Q(received_from_location__icontains=stock_inv_search) |
            Q(received_from_state__icontains=stock_inv_search) |
            Q(received_by__icontains=stock_inv_search) |
            Q(reason__icontains=stock_inv_search) |
            Q(awb_lr_number__icontains=stock_inv_search)
        )
    stock_inverters = stock_inverters_qs.order_by('-created_at')

    # Add 'Filled By' and 'Date Filled' columns for stock inverters
    stock_inverters_display = []
    for form in stock_inverters:
        stock_inverters_display.append({
            'form': form,
            'filled_by': form.user.get_full_name() if form.user else '',
            'date_filled': form.created_at.strftime('%Y-%m-%d'),
        })

    # Build unified all_entries list for 'All Entries' tab
    all_entries = []
    for form in repairing_forms:
        all_entries.append({
            'form_type': 'Repairing',
            'serial': getattr(form, 'inverter_id', ''),
            'customer': getattr(form, 'customer_abbrev', ''),
            'date': form.created_at.strftime('%Y-%m-%d'),
            'remark': getattr(form, 'remark', getattr(form, 'status', '')),
        })
    for form in inward_forms:
        all_entries.append({
            'form_type': 'Inward',
            'serial': getattr(form, 'inverter_id', ''),
            'customer': getattr(form, 'customer_name', ''),
            'date': form.created_at.strftime('%Y-%m-%d'),
            'remark': getattr(form, 'remark', '')
        })
    for form in outward_forms:
        all_entries.append({
            'form_type': 'Outward',
            'serial': getattr(form, 'inverter_id_outward', ''),
            'customer': getattr(form, 'sent_to_company', ''),
            'date': form.created_at.strftime('%Y-%m-%d'),
            'remark': getattr(form, 'remark', '')
        })
    for form in service_forms:
        all_entries.append({
            'form_type': 'Service',
            'serial': getattr(form, 'serial_number', ''),
            'customer': f"{getattr(form, 'customer_first_name', '')} {getattr(form, 'customer_last_name', '')}",
            'date': form.created_at.strftime('%Y-%m-%d'),
            'remark': getattr(form, 'actual_work_done', getattr(form, 'suggestions', ''))
        })
    for form in stock_requisitions:
        all_entries.append({
            'form_type': 'Stock Requisition',
            'serial': getattr(form, 'serial_number', ''),
            'customer': getattr(form, 'component_type', ''),
            'date': form.created_at.strftime('%Y-%m-%d'),
            'remark': getattr(form, 'description', '')
        })
    for form in dispatched_stock:
        all_entries.append({
            'form_type': 'Dispatched Stock',
            'serial': getattr(form, 'serial_number', ''),
            'customer': getattr(form, 'engineer_name', ''),
            'date': form.created_at.strftime('%Y-%m-%d'),
            'remark': getattr(form, 'dispatch_remarks', '')
        })
    context = {
        'user': user,
        'view_mode': view_mode,
        'team_users': team_users,
        'repairing_forms': repairing_forms,
        'inward_forms': inward_forms,
        'outward_forms': outward_forms,
        'service_forms': service_forms,
        'stock_requisitions': stock_requisitions,
        'dispatched_stock': dispatched_stock,
        'stock_inverters': stock_inverters,
        'stock_inverters_display': stock_inverters_display,
        'stock_inv_date_from': stock_inv_date_from,
        'stock_inv_date_to': stock_inv_date_to,
        'stock_inv_search': stock_inv_search,
        # Repairing filters
        'repairing_date_from': repairing_date_from,
        'repairing_date_to': repairing_date_to,
        'repairing_search': repairing_search,
        # Inward filters
        'inward_date_from': inward_date_from,
        'inward_date_to': inward_date_to,
        'inward_search': inward_search,
        # Outward filters
        'outward_date_from': outward_date_from,
        'outward_date_to': outward_date_to,
        'outward_search': outward_search,
        # Service filters
        'service_date_from': service_date_from,
        'service_date_to': service_date_to,
        'service_search': service_search,
        # Stock requisition filters
        'stock_req_date_from': stock_req_date_from,
        'stock_req_date_to': stock_req_date_to,
        'stock_req_search': stock_req_search,
        'stock_req_status': stock_req_status,
        # Stock dispatched filters
        'stock_disp_date_from': stock_disp_date_from,
        'stock_disp_date_to': stock_disp_date_to,
        'stock_disp_search': stock_disp_search,
        'all_entries': all_entries,
    }
    
    return render(request, 'forms/team_data.html', context)


@login_required
def export_team_data(request, form_type, format):
    """Export team data to CSV or XLSX for managers/superusers"""
    import io
    from datetime import datetime, timedelta
    
    user = request.user
    
    # Check permissions
    if user.is_superuser:
        team_users = User.objects.all()
        view_mode = 'all'
    elif hasattr(user, 'profile') and user.profile.team_members.exists():
        team_users = user.profile.get_team_members()
        view_mode = 'team'
    else:
        return HttpResponseForbidden("You don't have permission to export team data")
    
    team_names = [u.get_full_name() or u.username for u in team_users]
    team_emails = list(team_users.values_list('email', flat=True))
    
    # Get forms based on type with tab-specific filters
    if form_type == 'repairing':
        date_from = request.GET.get('repairing_date_from', '')
        date_to = request.GET.get('repairing_date_to', '')
        search_query = request.GET.get('repairing_search', '')
        
        if view_mode == 'all':
            queryset = RepairingForm.objects.all()
        else:
            repairing_q = Q()
            for name in team_names:
                repairing_q |= Q(repaired_by__icontains=name) | Q(tested_by__icontains=name)
            queryset = RepairingForm.objects.filter(repairing_q)
        if date_from:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            queryset = queryset.filter(created_at__gte=from_date)
        if date_to:
            to_date = datetime.strptime(date_to, '%Y-%m-%d')
            end_date = to_date + timedelta(days=1)
            queryset = queryset.filter(created_at__lt=end_date)
        if search_query:
            queryset = queryset.filter(
                Q(case_number__icontains=search_query) |
                Q(customer_abbrev__icontains=search_query) |
                Q(inverter_id__icontains=search_query)
            )
        queryset = queryset.order_by('-created_at')
        filename_prefix = "team_repairing_forms"
        headers = [
            'Date', 'Case Number', 'Customer', 'Object', 'Inverter ID', 'PCB Serial Number', 'PCB Specification', 'PCB Rating',
            'Inverter Spec', 'Inverter Rating', 'Battery', 'Fault Location', 'Repair Content', 'Repaired By', 'Tested By', 'Filled By'
        ]
        
    elif form_type == 'inward':
        date_from = request.GET.get('inward_date_from', '')
        date_to = request.GET.get('inward_date_to', '')
        search_query = request.GET.get('inward_search', '')
        
        if view_mode == 'all':
            queryset = InwardForm.objects.all()
        else:
            inward_q = Q()
            for email in team_emails:
                inward_q |= Q(email=email)
            for name in team_names:
                inward_q |= Q(received_by__icontains=name)
            queryset = InwardForm.objects.filter(inward_q)
        if date_from:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            queryset = queryset.filter(created_at__gte=from_date)
        if date_to:
            to_date = datetime.strptime(date_to, '%Y-%m-%d')
            end_date = to_date + timedelta(days=1)
            queryset = queryset.filter(created_at__lt=end_date)
        if search_query:
            queryset = queryset.filter(
                Q(customer_name__icontains=search_query) |
                Q(inverter_id__icontains=search_query) |
                Q(awb_lr_number__icontains=search_query)
            )
        queryset = queryset.order_by('-created_at')
        filename_prefix = "team_inward_forms"
        headers = [
            'Date', 'Inward Object', 'Customer Abbreviation', 'Customer Name', 'Inverter ID', 'Battery ID', 'PCB Serial Number',
            'Inverter Specification', 'Inverter Ratings', 'Battery Model', 'No of MPPT', 'Current/MPPT', 'PCB Quantity',
            'Received From Location', 'Received From District', 'Received From State', 'Pincode', 'Received By', 'Reason',
            'Transportation Mode', 'AWB Number', 'Filled By'
        ]
        
    elif form_type == 'outward':
        date_from = request.GET.get('outward_date_from', '')
        date_to = request.GET.get('outward_date_to', '')
        search_query = request.GET.get('outward_search', '')
        
        if view_mode == 'all':
            queryset = OutwardForm.objects.all()
        else:
            outward_q = Q()
            for name in team_names:
                outward_q |= Q(sent_by__icontains=name) | Q(approved_by__icontains=name)
            queryset = OutwardForm.objects.filter(outward_q)
        if date_from:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            queryset = queryset.filter(created_at__gte=from_date)
        if date_to:
            to_date = datetime.strptime(date_to, '%Y-%m-%d')
            end_date = to_date + timedelta(days=1)
            queryset = queryset.filter(created_at__lt=end_date)
        if search_query:
            queryset = queryset.filter(
                Q(inverter_id_outward__icontains=search_query) |
                Q(sent_to_company__icontains=search_query) |
                Q(awb_number__icontains=search_query)
            )
        queryset = queryset.order_by('-created_at')
        filename_prefix = "team_outward_forms"
        headers = [
            'Date', 'Outward Object', 'Inverter ID (Outward)', 'Inverter Spec', 'Inverter Rating', 'Battery', 'Battery ID',
            'Sent To Company', 'Sent To Address', 'Sent To District', 'Sent To State', 'Pincode', 'Sent By', 'Approved By',
            'Control Card Changed', 'New Serial Number', 'Inverter ID (Inward)', 'Inverter Replaced', 'Delivered Through',
            'AWB Number', 'Remarks', 'Filled By'
        ]
        
    elif form_type == 'service':
        date_from = request.GET.get('service_date_from', '')
        date_to = request.GET.get('service_date_to', '')
        search_query = request.GET.get('service_search', '')
        
        if view_mode == 'all':
            queryset = ServiceReportForm.objects.all()
        else:
            service_q = Q()
            for email in team_emails:
                service_q |= Q(email=email)
            for user_obj in team_users:
                service_q |= Q(engineer_first_name__icontains=user_obj.first_name, engineer_last_name__icontains=user_obj.last_name)
            queryset = ServiceReportForm.objects.filter(service_q)
        if date_from:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            queryset = queryset.filter(created_at__gte=from_date)
        if date_to:
            to_date = datetime.strptime(date_to, '%Y-%m-%d')
            end_date = to_date + timedelta(days=1)
            queryset = queryset.filter(created_at__lt=end_date)
        if search_query:
            queryset = queryset.filter(
                Q(serial_number__icontains=search_query) |
                Q(customer_first_name__icontains=search_query) |
                Q(customer_last_name__icontains=search_query)
            )
        queryset = queryset.order_by('-created_at')
        filename_prefix = "team_service_reports"
        headers = [
            'Engineer', 'Customer', 'Product Type', 'Serial Number', 'Service Date', 'Address', 'Battery Type', 'Battery Make',
            'Battery Voltage', 'PV Capacity (kW)', 'AC Cable Size', 'Physical Observation', 'Actual Work Done', 'Cause of Failure',
            'Conclusion', 'Customer Ratings', 'Suggestions', 'Created', 'Filled By'
        ]
        
    elif form_type == 'stock-req':
        date_from = request.GET.get('stock_req_date_from', '')
        date_to = request.GET.get('stock_req_date_to', '')
        search_query = request.GET.get('stock_req_search', '')
        status_filter = request.GET.get('stock_req_status', '')
        
        if view_mode == 'all':
            queryset = StockRequisition.objects.all()
        else:
            stock_req_q = Q()
            for name in team_names:
                stock_req_q |= Q(manager_name__icontains=name)
            queryset = StockRequisition.objects.filter(stock_req_q)
        if date_from:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            queryset = queryset.filter(created_at__gte=from_date)
        if date_to:
            to_date = datetime.strptime(date_to, '%Y-%m-%d')
            end_date = to_date + timedelta(days=1)
            queryset = queryset.filter(created_at__lt=end_date)
        if search_query:
            queryset = queryset.filter(
                Q(serial_number__icontains=search_query) |
                Q(component_type__icontains=search_query) |
                Q(description__icontains=search_query)
            )
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        queryset = queryset.order_by('-created_at')
        filename_prefix = "team_stock_requisitions"
        headers = ['Serial Number', 'Component', 'Quantity', 'Manager', 'Status', 'Date']
        
    elif form_type == 'stock-disp':
        date_from = request.GET.get('stock_disp_date_from', '')
        date_to = request.GET.get('stock_disp_date_to', '')
        search_query = request.GET.get('stock_disp_search', '')
        
        if view_mode == 'all':
            queryset = DispatchedStock.objects.all()
        else:
            dispatched_q = Q()
            for name in team_names:
                dispatched_q |= Q(dispatched_by__icontains=name) | Q(engineer_name__icontains=name)
            queryset = DispatchedStock.objects.filter(dispatched_q)
        if date_from:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            queryset = queryset.filter(created_at__gte=from_date)
        if date_to:
            to_date = datetime.strptime(date_to, '%Y-%m-%d')
            end_date = to_date + timedelta(days=1)
            queryset = queryset.filter(created_at__lt=end_date)
        if search_query:
            queryset = queryset.filter(
                Q(serial_number__icontains=search_query) |
                Q(component_type__icontains=search_query) |
                Q(tracking_number__icontains=search_query)
            )
        queryset = queryset.order_by('-created_at')
        filename_prefix = "team_dispatched_stock"
        headers = ['Serial Number', 'Component', 'Quantity', 'Engineer', 'Courier', 'Tracking', 'Date']
    
    else:
        return HttpResponse("Invalid form type", status=400)
    
    # Export based on format
    def get_user_display(obj):
        if hasattr(obj, 'user') and obj.user:
            return obj.user.get_full_name() or obj.user.username
        return '-'

    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(headers)
        
        for obj in queryset:
            if form_type == 'repairing':
                writer.writerow([
                    obj.created_at.strftime('%d %b %Y'), obj.case_number, obj.customer_abbrev, obj.repairing_object, obj.inverter_id,
                    obj.pcb_serial_number, obj.pcb_specification, obj.pcb_rating, obj.inverter_spec, obj.inverter_rating, obj.battery,
                    obj.fault_location, obj.repair_content, obj.repaired_by, obj.tested_by, get_user_display(obj)
                ])
            elif form_type == 'inward':
                writer.writerow([
                    obj.created_at.strftime('%d %b %Y'), obj.inward_object, obj.customer_abbrev, obj.customer_name, obj.inverter_id, obj.battery_id,
                    obj.pcb_serial_number, obj.inverter_specs, obj.inverter_ratings, obj.battery, obj.no_of_mppt, obj.current_mppt, obj.pcb_quantity,
                    obj.received_from_location, obj.received_from_district, obj.received_from_state, obj.pincode, obj.received_by, obj.reason,
                    obj.transportation_mode, obj.awb_lr_number, get_user_display(obj)
                ])
            elif form_type == 'outward':
                writer.writerow([
                    obj.created_at.strftime('%d %b %Y'), obj.outward_object, obj.inverter_id_outward, obj.inverter_spec, obj.inverter_rating, obj.battery,
                    obj.battery_id, obj.sent_to_company, obj.sent_to_address, obj.sent_to_district, obj.sent_to_state, obj.pincode, obj.sent_by,
                    obj.approved_by, obj.control_card_changed, obj.new_serial_number, obj.inverter_id_inward, obj.inverter_replaced,
                    obj.delivered_through, obj.awb_number, obj.remarks, get_user_display(obj)
                ])
            elif form_type == 'service':
                writer.writerow([
                    f"{obj.engineer_first_name} {obj.engineer_last_name}", f"{obj.customer_first_name} {obj.customer_last_name}", obj.product_type,
                    obj.serial_number, obj.date_of_service.strftime('%d %b %Y') if obj.date_of_service else 'N/A',
                    f"{obj.address_street}, {obj.address_city}, {obj.address_state}, {obj.address_zip}", obj.battery_type, obj.battery_make,
                    obj.battery_voltage, obj.pv_capacity_kw, obj.ac_cable_size, obj.physical_observation, obj.actual_work_done,
                    obj.cause_of_failure, obj.conclusion, obj.customer_ratings, obj.suggestions, obj.created_at.strftime('%d %b %Y'), get_user_display(obj)
                ])
            elif form_type == 'stock-req':
                writer.writerow([obj.serial_number, obj.component_type, obj.quantity_required, obj.manager_name, obj.status, obj.created_at.strftime('%Y-%m-%d')])
            elif form_type == 'stock-disp':
                writer.writerow([obj.serial_number, obj.component_type, obj.quantity_dispatched, obj.engineer_name, obj.courier_name, obj.tracking_number, obj.dispatch_date.strftime('%Y-%m-%d')])
        
        return response

    elif format == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = form_type.replace('-', ' ').title()
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        for obj in queryset:
            if form_type == 'repairing':
                ws.append([
                    obj.created_at.strftime('%d %b %Y'), obj.case_number, obj.customer_abbrev, obj.repairing_object, obj.inverter_id,
                    obj.pcb_serial_number, obj.pcb_specification, obj.pcb_rating, obj.inverter_spec, obj.inverter_rating, obj.battery,
                    obj.fault_location, obj.repair_content, obj.repaired_by, obj.tested_by, get_user_display(obj)
                ])
            elif form_type == 'inward':
                ws.append([
                    obj.created_at.strftime('%d %b %Y'), obj.inward_object, obj.customer_abbrev, obj.customer_name, obj.inverter_id, obj.battery_id,
                    obj.pcb_serial_number, obj.inverter_specs, obj.inverter_ratings, obj.battery, obj.no_of_mppt, obj.current_mppt, obj.pcb_quantity,
                    obj.received_from_location, obj.received_from_district, obj.received_from_state, obj.pincode, obj.received_by, obj.reason,
                    obj.transportation_mode, obj.awb_lr_number, get_user_display(obj)
                ])
            elif form_type == 'outward':
                ws.append([
                    obj.created_at.strftime('%d %b %Y'), obj.outward_object, obj.inverter_id_outward, obj.inverter_spec, obj.inverter_rating, obj.battery,
                    obj.battery_id, obj.sent_to_company, obj.sent_to_address, obj.sent_to_district, obj.sent_to_state, obj.pincode, obj.sent_by,
                    obj.approved_by, obj.control_card_changed, obj.new_serial_number, obj.inverter_id_inward, obj.inverter_replaced,
                    obj.delivered_through, obj.awb_number, obj.remarks, get_user_display(obj)
                ])
            elif form_type == 'service':
                ws.append([
                    f"{obj.engineer_first_name} {obj.engineer_last_name}", f"{obj.customer_first_name} {obj.customer_last_name}", obj.product_type,
                    obj.serial_number, obj.date_of_service.strftime('%d %b %Y') if obj.date_of_service else 'N/A',
                    f"{obj.address_street}, {obj.address_city}, {obj.address_state}, {obj.address_zip}", obj.battery_type, obj.battery_make,
                    obj.battery_voltage, obj.pv_capacity_kw, obj.ac_cable_size, obj.physical_observation, obj.actual_work_done,
                    obj.cause_of_failure, obj.conclusion, obj.customer_ratings, obj.suggestions, obj.created_at.strftime('%d %b %Y'), get_user_display(obj)
                ])
            elif form_type == 'stock-req':
                ws.append([obj.serial_number, obj.component_type, obj.quantity_required, obj.manager_name, obj.status, obj.created_at.strftime('%Y-%m-%d')])
            elif form_type == 'stock-disp':
                ws.append([obj.serial_number, obj.component_type, obj.quantity_dispatched, obj.engineer_name, obj.courier_name, obj.tracking_number, obj.dispatch_date.strftime('%Y-%m-%d')])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        return response

    else:
        return HttpResponse("Invalid format. Use 'csv' or 'xlsx'.", status=400)


# Diagnostic endpoint for Bigship calculator debugging
def bigship_diagnostic(request):
    """Diagnostic endpoint to test Bigship Excel file loading and calculator"""
    from django.http import JsonResponse
    import pathlib
    from forms.calculator.data_loader import load_pincode_master
    from forms.calculator.bigship_calculator import Bigship, BigshipPincodeDB
    from forms.calculator.config import Settings
    
    diagnostics = {
        "status": "ok",
        "errors": [],
        "warnings": [],
        "info": {}
    }
    
    try:
        # Get base directory
        base_dir = str(pathlib.Path(__file__).resolve().parents[2])
        diagnostics["info"]["base_dir"] = base_dir
        
        # Check if Excel file exists
        excel_file = pathlib.Path(base_dir) / "Bigship Serviceable Pincode.xlsx"
        diagnostics["info"]["excel_file_path"] = str(excel_file)
        diagnostics["info"]["excel_file_exists"] = excel_file.exists()
        
        if not excel_file.exists():
            diagnostics["errors"].append(f"[CRITICAL] Excel file not found at {excel_file}")
            diagnostics["status"] = "error"
        else:
            # Try to load the pincode database
            try:
                bigship_pins = BigshipPincodeDB(str(excel_file))
                diagnostics["info"]["serviceable_pincodes_count"] = len(bigship_pins._serviceable)
                diagnostics["info"]["oda_pincodes_count"] = len(bigship_pins._oda_pincodes)
                
                if len(bigship_pins._serviceable) == 0:
                    diagnostics["errors"].append("[CRITICAL] No serviceable pincodes loaded from Excel file")
                    diagnostics["status"] = "error"
                else:
                    diagnostics["info"]["sample_pincodes"] = list(bigship_pins._serviceable)[:5]
            except Exception as e:
                diagnostics["errors"].append(f"[CRITICAL] Failed to load Bigship pincode database: {str(e)}")
                diagnostics["status"] = "error"
                import traceback
                diagnostics["errors"].append(traceback.format_exc())
        
        # Try to initialize Bigship calculator
        try:
            settings = Settings()
            bigship = Bigship(settings, base_dir=base_dir)
            diagnostics["info"]["bigship_initializer"] = "success"
        except Exception as e:
            diagnostics["errors"].append(f"[CRITICAL] Failed to initialize Bigship calculator: {str(e)}")
            diagnostics["status"] = "error"
            import traceback
            diagnostics["errors"].append(traceback.format_exc())
        
        # List all files in project root
        project_root = pathlib.Path(base_dir)
        xlsx_files = sorted([f.name for f in project_root.glob("*.xlsx")])
        diagnostics["info"]["xlsx_files_in_project_root"] = xlsx_files
        
    except Exception as e:
        diagnostics["status"] = "error"
        diagnostics["errors"].append(f"[CRITICAL] Unexpected error: {str(e)}")
        import traceback
        diagnostics["errors"].append(traceback.format_exc())
    
    return JsonResponse(diagnostics, indent=2)


# Import login_required decorator
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import never_cache


# ===== CHECK-IN / CHECK-OUT SYSTEM =====

@login_required
@never_cache
def checkin_page(request):
    """Display check-in/check-out page with geolocation"""
    from django.shortcuts import render
    from .models import CheckInOut
    
    # Get today's check-in record if exists
    today_checkin = CheckInOut.get_today_checkin(request.user)
    
    context = {
        'today_checkin': today_checkin,
        'is_checked_in': today_checkin and today_checkin.is_checked_in(),
    }
    
    return render(request, 'forms/checkin.html', context)


@login_required
def checkin_submit(request):
    """Handle check-in submission with geolocation data"""
    from .models import CheckInOut
    from django.utils import timezone
    import json
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            latitude = data.get('latitude')
            longitude = data.get('longitude')
            location = data.get('location', '')
            
            if not latitude or not longitude:
                return JsonResponse({
                    'success': False,
                    'error': 'Location data not available. Please enable location services.'
                }, status=400)
            
            # Check if already checked in today
            today_checkin = CheckInOut.get_today_checkin(request.user)
            
            if today_checkin and today_checkin.is_checked_in():
                return JsonResponse({
                    'success': False,
                    'error': 'Already checked in today. Please check out first.'
                }, status=400)
            
            # Create new check-in record
            checkin = CheckInOut.objects.create(
                user=request.user,
                check_in_time=timezone.now(),
                check_in_latitude=Decimal(str(latitude)),
                check_in_longitude=Decimal(str(longitude)),
                check_in_location=location[:255] if location else 'Location Not Captured'
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Checked in successfully!',
                'check_in_time': checkin.check_in_time.strftime('%I:%M %p'),
                'location': checkin.check_in_location
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)


@login_required
def checkout_submit(request):
    """Handle check-out submission with geolocation data"""
    from .models import CheckInOut
    from django.utils import timezone
    import json
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            latitude = data.get('latitude')
            longitude = data.get('longitude')
            location = data.get('location', '')
            
            if not latitude or not longitude:
                return JsonResponse({
                    'success': False,
                    'error': 'Location data not available. Please enable location services.'
                }, status=400)
            
            # Get active check-in
            checkin = CheckInOut.get_active_checkin(request.user)
            
            if not checkin:
                return JsonResponse({
                    'success': False,
                    'error': 'No active check-in found. Please check in first.'
                }, status=400)
            
            # Update check-out time and location
            checkin.check_out_time = timezone.now()
            checkin.check_out_latitude = Decimal(str(latitude))
            checkin.check_out_longitude = Decimal(str(longitude))
            checkin.check_out_location = location[:255] if location else 'Location Not Captured'
            
            # Calculate duration
            checkin.calculate_duration()
            checkin.save()
            
            # Generate daily travel summary for this day
            from .models import DailyTravelSummary
            DailyTravelSummary.generate_summary(request.user, checkin.date)
            
            return JsonResponse({
                'success': True,
                'message': 'Checked out successfully!',
                'check_out_time': checkin.check_out_time.strftime('%I:%M %p'),
                'duration': f"{int(checkin.duration_hours)}h {int((checkin.duration_hours - int(checkin.duration_hours)) * 60)}m",
                'location': checkin.check_out_location
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)


@login_required
def get_checkin_status(request):
    """Get current check-in status and location"""
    from .models import CheckInOut
    import json
    
    try:
        today_checkin = CheckInOut.get_today_checkin(request.user)
        
        if today_checkin:
            is_checked_in = today_checkin.is_checked_in()
            
            response = {
                'success': True,
                'is_checked_in': is_checked_in,
                'check_in_time': today_checkin.check_in_time.strftime('%H:%M:%S'),
                'check_in_location': today_checkin.check_in_location,
                'check_in_coords': {
                    'lat': float(today_checkin.check_in_latitude),
                    'lng': float(today_checkin.check_in_longitude)
                }
            }
            
            if today_checkin.check_out_time:
                response.update({
                    'check_out_time': today_checkin.check_out_time.strftime('%H:%M:%S'),
                    'check_out_location': today_checkin.check_out_location,
                    'check_out_coords': {
                        'lat': float(today_checkin.check_out_latitude),
                        'lng': float(today_checkin.check_out_longitude)
                    },
                    'duration_hours': float(today_checkin.duration_hours)
                })
            
            return JsonResponse(response)
        
        else:
            return JsonResponse({
                'success': True,
                'is_checked_in': False,
                'message': 'No check-in record for today'
            })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def attendance_history(request):
    """View employee's check-in/check-out history"""
    from django.shortcuts import render
    from django.core.paginator import Paginator
    from .models import CheckInOut
    
    # Get all check-ins for the logged-in user, ordered by date descending
    checkins = CheckInOut.objects.filter(user=request.user).order_by('-date', '-check_in_time')
    
    # Pagination
    paginator = Paginator(checkins, 50)  # 50 records per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'checkins': page_obj,
        'total_records': paginator.count,
    }
    
    return render(request, 'forms/attendance_history.html', context)


@login_required
def manager_attendance_view(request):
    """Manager view of team's attendance (for managers)"""
    from django.shortcuts import render
    from django.contrib.auth.models import User
    from .models import CheckInOut, UserProfile
    from django.utils import timezone
    
    # Check if user is a manager or superuser
    has_team_members = User.objects.filter(profile__manager=request.user).exists()
    
    if not (request.user.is_superuser or has_team_members):
        return HttpResponseForbidden("Access Denied: Manager access required.")
    
    # Get team members under this manager
    team_members = User.objects.filter(
        profile__manager=request.user
    ).select_related('profile').order_by('first_name', 'last_name')
    
    # Get today's attendance for all team members
    today_checkins = CheckInOut.objects.filter(
        user__in=team_members,
        date=timezone.now().date()
    ).select_related('user').order_by('-check_in_time')
    
    context = {
        'team_members': team_members,
        'today_checkins': today_checkins,
        'attended': today_checkins.count(),
    }
    
    return render(request, 'forms/manager_attendance.html', context)


# ═══════════════════════════════════════════════════════════════════
# LOCATION TRACKING VIEWS - Hourly Pings & Travel History
# ═══════════════════════════════════════════════════════════════════

@login_required
def save_location_ping(request):
    """
    Save periodic location ping during check-in session.
    Tracks location every hour and handles coverage gaps.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST method required'})
    
    from .models import LocationTracking, CheckInOut
    from django.utils import timezone
    import json
    
    try:
        data = json.loads(request.body)
        
        # Get current active check-in session
        active_checkin = CheckInOut.get_active_checkin(request.user)
        
        # Extract location data
        ping_type = data.get('ping_type', 'hourly')
        latitude = data.get('latitude')
        longitude = data.get('longitude')
        location_address = data.get('location', '')
        accuracy = data.get('accuracy')
        device_info = data.get('device_info', '')
        coverage_gap_seconds = data.get('coverage_gap_seconds', 0)
        
        # Create location ping record
        location_ping = LocationTracking.record_ping(
            user=request.user,
            checkin=active_checkin,
            ping_type=ping_type,
            latitude=latitude,
            longitude=longitude,
            location_address=location_address,
            accuracy=accuracy,
            device_info=device_info,
            coverage_gap_seconds=coverage_gap_seconds
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Location ping saved successfully',
            'ping_id': location_ping.id,
            'ping_time': location_ping.ping_time.strftime('%I:%M %p'),
            'is_available': location_ping.is_location_available
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
def travel_history_view(request, username=None):
    """
    View employee's detailed travel history with all location pings.
    Shows TODAY'S data by default.
    Access: Users see own data, Managers see team data, Superusers see all.
    """
    from django.shortcuts import render, get_object_or_404
    from django.contrib.auth.models import User
    from .models import LocationTracking, CheckInOut, UserProfile
    from django.core.paginator import Paginator
    from django.utils import timezone
    from datetime import timedelta, datetime, time
    from django.db import models
    
    # Determine accessible employees based on role
    if request.user.is_superuser:
        # Superuser: Can see all employees
        accessible_users = User.objects.filter(is_active=True).exclude(username='admin')
        can_switch_user = True
    else:
        # Check if user is a manager
        try:
            user_profile = request.user.profile
            team_members = user_profile.get_team_members()
            if team_members.exists():
                # Manager: Can see self and team members
                accessible_users = User.objects.filter(
                    models.Q(id=request.user.id) | models.Q(id__in=team_members.values_list('id', flat=True))
                ).filter(is_active=True)
                can_switch_user = True
            else:
                # Regular user: Can only see own data
                accessible_users = User.objects.filter(id=request.user.id)
                can_switch_user = False
        except UserProfile.DoesNotExist:
            # No profile: Can only see own data
            accessible_users = User.objects.filter(id=request.user.id)
            can_switch_user = False
    
    # Get target user (with access control)
    if username and can_switch_user:
        target_user = get_object_or_404(User, username=username)
        # Verify access
        if target_user not in accessible_users:
            return HttpResponseForbidden("Access Denied: You don't have permission to view this user's data.")
    else:
        target_user = request.user
    
    # Get today's date range by default
    today = timezone.now().date()
    start_datetime = timezone.make_aware(datetime.combine(today, time.min))
    end_datetime = timezone.make_aware(datetime.combine(today, time.max))
    
    # Get location tracking records for TODAY only
    location_pings = LocationTracking.objects.filter(
        user=target_user,
        ping_time__gte=start_datetime,
        ping_time__lte=end_datetime
    ).select_related('checkin', 'user').order_by('-ping_time')
    
    # Get today's check-in session
    today_checkin = CheckInOut.objects.filter(
        user=target_user,
        date=today
    ).first()
    
    # Statistics
    total_pings = location_pings.count()
    available_pings = location_pings.filter(is_location_available=True).count()
    unavailable_pings = total_pings - available_pings
    total_coverage_gap = sum(ping.coverage_gap_seconds for ping in location_pings)
    
    # Paginate location pings (50 per page)
    paginator = Paginator(location_pings, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get list of accessible employees (for dropdown)
    all_employees = accessible_users.order_by('first_name', 'last_name')
    
    context = {
        'target_user': target_user,
        'location_pings': page_obj,
        'today_checkin': today_checkin,
        'today_date': today,
        'total_pings': total_pings,
        'available_pings': available_pings,
        'unavailable_pings': unavailable_pings,
        'total_coverage_gap': total_coverage_gap,
        'coverage_gap_hours': total_coverage_gap // 3600,
        'coverage_gap_minutes': (total_coverage_gap % 3600) // 60,
        'all_employees': all_employees,
        'can_switch_user': can_switch_user,
        'is_authorized': True,
    }
    
    return render(request, 'forms/travel_history.html', context)


@login_required
def daily_travel_summary_view(request, username=None):
    """
    View employee's day-wise travel summary with total distance.
    Shows 6 locations max per day (check-in + 4 hourly + check-out).
    Access: Users see own data, Managers see team data, Superusers see all.
    """
    from django.shortcuts import render, get_object_or_404
    from django.contrib.auth.models import User
    from .models import DailyTravelSummary, LocationTracking, CheckInOut, UserProfile
    from django.core.paginator import Paginator
    from django.utils import timezone
    from datetime import timedelta, datetime
    from django.db import models
    
    # Determine accessible employees based on role
    if request.user.is_superuser:
        # Superuser: Can see all employees
        accessible_users = User.objects.filter(is_active=True).exclude(username='admin')
        can_switch_user = True
    else:
        # Check if user is a manager
        try:
            user_profile = request.user.profile
            team_members = user_profile.get_team_members()
            if team_members.exists():
                # Manager: Can see self and team members
                accessible_users = User.objects.filter(
                    models.Q(id=request.user.id) | models.Q(id__in=team_members.values_list('id', flat=True))
                ).filter(is_active=True)
                can_switch_user = True
            else:
                # Regular user: Can only see own data
                accessible_users = User.objects.filter(id=request.user.id)
                can_switch_user = False
        except UserProfile.DoesNotExist:
            # No profile: Can only see own data
            accessible_users = User.objects.filter(id=request.user.id)
            can_switch_user = False
    
    # Get target user (with access control)
    if username and can_switch_user:
        target_user = get_object_or_404(User, username=username)
        # Verify access
        if target_user not in accessible_users:
            return HttpResponseForbidden("Access Denied: You don't have permission to view this user's data.")
    else:
        target_user = request.user
    
    # Custom date filtering
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    
    # Default date range: last 30 days
    today = timezone.now().date()
    if to_date_str:
        try:
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
        except ValueError:
            to_date = today
    else:
        to_date = today
    
    if from_date_str:
        try:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
        except ValueError:
            from_date = today - timedelta(days=30)
    else:
        from_date = today - timedelta(days=30)
    
    # Ensure from_date is not after to_date
    if from_date > to_date:
        from_date, to_date = to_date, from_date
    
    start_date = from_date
    end_date = to_date
    
    # Get or generate daily summaries
    # First ensure all summaries exist for days with check-ins
    checkins = CheckInOut.objects.filter(
        user=target_user,
        date__gte=start_date,
        date__lte=end_date
    )
    
    for checkin in checkins:
        DailyTravelSummary.generate_summary(target_user, checkin.date)
    
    # Get daily summaries
    daily_summaries = DailyTravelSummary.objects.filter(
        user=target_user,
        date__gte=start_date,
        date__lte=end_date
    ).order_by('-date')
    
    # Statistics
    total_days = daily_summaries.count()
    total_distance = sum(summary.total_distance_km for summary in daily_summaries)
    avg_distance = round(total_distance / total_days, 2) if total_days > 0 else 0
    total_locations = sum(summary.location_count for summary in daily_summaries)
    
    # Paginate daily summaries (20 per page)
    paginator = Paginator(daily_summaries, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get list of accessible employees (for dropdown)
    all_employees = accessible_users.order_by('first_name', 'last_name')
    
    context = {
        'target_user': target_user,
        'daily_summaries': page_obj,
        'total_days': total_days,
        'total_distance': total_distance,
        'avg_distance': avg_distance,
        'total_locations': total_locations,
        'from_date': from_date,
        'to_date': to_date,
        'all_employees': all_employees,
        'can_switch_user': can_switch_user,
        'is_authorized': True,
    }
    
    return render(request, 'forms/daily_travel_summary.html', context)


@login_required
def day_route_map_view(request, username, date_str):
    """
    Display all locations for a specific day on an interactive map.
    Shows the complete route with all 6 locations (check-in + 4 hourly + check-out).
    Access: Users see own data, Managers see team data, Superusers see all.
    """
    from django.shortcuts import render, get_object_or_404
    from django.contrib.auth.models import User
    from .models import LocationTracking, DailyTravelSummary, UserProfile
    from datetime import datetime
    import json
    
    # Get target user
    target_user = get_object_or_404(User, username=username)
    
    # Access control check
    if request.user.is_superuser:
        # Superuser: Can see all
        pass
    elif target_user == request.user:
        # User viewing own data
        pass
    else:
        # Check if user is manager of target user
        try:
            user_profile = request.user.profile
            team_members = user_profile.get_team_members()
            if target_user not in team_members:
                return HttpResponseForbidden("Access Denied: You don't have permission to view this user's data.")
        except UserProfile.DoesNotExist:
            return HttpResponseForbidden("Access Denied: You don't have permission to view this user's data.")
    
    # Parse date
    try:
        date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponseForbidden("Invalid date format")
    
    # Get all locations for this day
    locations = LocationTracking.get_daily_locations(target_user, date)
    
    # Get daily summary
    try:
        daily_summary = DailyTravelSummary.objects.get(user=target_user, date=date)
    except DailyTravelSummary.DoesNotExist:
        daily_summary = DailyTravelSummary.generate_summary(target_user, date)
    
    # Prepare location data for map
    location_data = []
    for i, loc in enumerate(locations, 1):
        location_data.append({
            'seq': i,
            'lat': float(loc.latitude),
            'lng': float(loc.longitude),
            'address': loc.location_address,
            'time': loc.ping_time.strftime('%I:%M %p'),
            'type': loc.ping_type,
        })
    
    context = {
        'target_user': target_user,
        'date': date,
        'daily_summary': daily_summary,
        'locations': json.dumps(location_data),  # Convert to JSON string
        'location_count': len(location_data),
    }
    
    return render(request, 'forms/day_route_map.html', context)







