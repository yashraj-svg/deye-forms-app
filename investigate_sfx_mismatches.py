#!/usr/bin/env python
"""Investigate Safexpress mismatches in detail."""
import openpyxl
from forms.calculator.data_loader import load_pincode_master
from forms.calculator.freight_calculator import Safexpress, QuoteInput, ShipmentItem
from forms.calculator.config import Settings

wb = openpyxl.load_workbook('SFX00017299 DEYE INVERTER TECHNOLOGY PRIVATE LIMITED MIS  (1).xlsx', data_only=True)
ws = wb.active

pins = load_pincode_master('.')
calc = Safexpress(Settings())

mismatch_rows = [9, 10, 14, 23, 27]

print("="*120)
print("INVESTIGATING SAFEXPRESS MISMATCHES")
print("="*120)

for row_num in mismatch_rows:
    from_pin = str(ws.cell(row_num, 4).value or '').strip()
    to_pin = str(ws.cell(row_num, 5).value or '').strip()
    weight = ws.cell(row_num, 9).value
    waybill = ws.cell(row_num, 3).value
    
    # Excel charges
    excel_basic = ws.cell(row_num, 10).value or 0
    excel_value = ws.cell(row_num, 13).value or 0
    excel_waybill = ws.cell(row_num, 14).value or 0
    excel_ucc = ws.cell(row_num, 16).value or 0
    excel_delivery_safe = ws.cell(row_num, 18).value or 0
    excel_fuel = ws.cell(row_num, 27).value or 0
    excel_subtotal = ws.cell(row_num, 30).value or 0
    excel_gst = ws.cell(row_num, 31).value or 0
    excel_total = ws.cell(row_num, 32).value or 0
    
    # Our calculation
    items = [ShipmentItem(weight_kg=float(weight), length_cm=0, breadth_cm=0, height_cm=0)]
    result = calc.calculate_quote(QuoteInput(from_pincode=from_pin, to_pincode=to_pin, items=items), pins)
    
    fp = pins.get(from_pin)
    tp = pins.get(to_pin)
    from_state = fp.state if fp else 'Unknown'
    to_state = tp.state if tp else 'Unknown'
    
    print(f"\n{'='*120}")
    print(f"Row {row_num}: {waybill}")
    print(f"Route: {from_pin} ({from_state}) → {to_pin} ({to_state})")
    print(f"Weight: {weight}kg | CW: {result.chargeable_weight_kg}kg")
    print(f"Zone: {result.from_zone} → {result.to_zone} | Rate: {result.rate_per_kg}/kg")
    print(f"{'='*120}")
    
    print(f"\n{'Charge':<30} {'Excel':<15} {'Calculator':<15} {'Diff':<10}")
    print(f"{'-'*70}")
    
    print(f"{'Basic Freight':<30} {excel_basic:<15.2f} {result.base_freight:<15.2f} {result.base_freight - excel_basic:+.2f}")
    print(f"{'Value Surcharge':<30} {excel_value:<15.2f} {result.surcharges.get('value_surcharge', 0):<15.2f}")
    print(f"{'Waybill Charges':<30} {excel_waybill:<15.2f} {result.surcharges.get('waybill', 0):<15.2f}")
    print(f"{'UCC Charges':<30} {excel_ucc:<15.2f} {result.surcharges.get('ucc', 0):<15.2f}")
    
    if excel_delivery_safe > 0:
        print(f"{'Delivery Safe Extension':<30} {excel_delivery_safe:<15.2f} {'NOT CALC':<15}")
    
    print(f"{'OSC (if applied)':<30} {0:<15.2f} {result.surcharges.get('osc', 0):<15.2f}")
    print(f"{'Fuel Surcharges':<30} {excel_fuel:<15.2f} {result.surcharges.get('fuel_surcharge', 0):<15.2f} {result.surcharges.get('fuel_surcharge', 0) - excel_fuel:+.2f}")
    
    print(f"{'-'*70}")
    print(f"{'Freight Subtotal':<30} {excel_subtotal:<15.2f} {result.total_before_gst:<15.2f} {result.total_before_gst - excel_subtotal:+.2f}")
    print(f"{'GST 18%':<30} {excel_gst:<15.2f} {result.gst_amount:<15.2f}")
    print(f"{'TOTAL FREIGHT':<30} {excel_total:<15.2f} {result.total_after_gst:<15.2f} {result.total_after_gst - excel_total:+.2f}")
    
    print(f"\nAll surcharges: {result.surcharges}")
    
    # Check if Delivery Safe is the issue
    if excel_delivery_safe > 0 and result.surcharges.get('oda') == 0:
        print(f"\n⚠️  NOTE: Excel shows Delivery Safe Extension ₹{excel_delivery_safe} but we show ODA ₹{result.surcharges.get('oda', 'N/A')}")
        print(f"    This might be why totals differ!")
