#!/usr/bin/env python
"""Validate all Safexpress rows against Excel."""
import openpyxl
from forms.calculator.data_loader import load_pincode_master
from forms.calculator.freight_calculator import Safexpress, QuoteInput, ShipmentItem
from forms.calculator.config import Settings

wb = openpyxl.load_workbook('SFX00017299 DEYE INVERTER TECHNOLOGY PRIVATE LIMITED MIS  (1).xlsx', data_only=True)
ws = wb.active

pins = load_pincode_master('.')
calc = Safexpress(Settings())

matches = 0
mismatches = 0
errors = 0
mismatch_details = []

print("="*100)
print("SAFEXPRESS - VALIDATING ALL ROWS")
print("="*100)

# Get total rows
total_rows = ws.max_row
print(f"\nChecking {total_rows - 1} shipments (rows 2 to {total_rows})...\n")

for row_num in range(2, total_rows + 1):
    from_pin = str(ws.cell(row_num, 4).value or '').strip()
    to_pin = str(ws.cell(row_num, 5).value or '').strip()
    weight = ws.cell(row_num, 9).value
    waybill = ws.cell(row_num, 3).value
    
    excel_total = ws.cell(row_num, 32).value
    
    if not from_pin or not to_pin or not weight or not excel_total:
        continue
    
    try:
        items = [ShipmentItem(weight_kg=float(weight), length_cm=0, breadth_cm=0, height_cm=0)]
        result = calc.calculate_quote(QuoteInput(from_pincode=from_pin, to_pincode=to_pin, items=items), pins)
        
        diff = result.total_after_gst - float(excel_total)
        
        if abs(diff) < 0.5:
            matches += 1
            print(f"Row {row_num:3d}: ✓ {waybill} | {from_pin}→{to_pin} ({weight}kg) | Calc: {result.total_after_gst:8.2f} | Excel: {excel_total:8.2f}")
        else:
            mismatches += 1
            mismatch_details.append({
                'row': row_num,
                'waybill': waybill,
                'from': from_pin,
                'to': to_pin,
                'weight': weight,
                'calc': result.total_after_gst,
                'excel': float(excel_total),
                'diff': diff
            })
            print(f"Row {row_num:3d}: ✗ {waybill} | {from_pin}→{to_pin} ({weight}kg) | Calc: {result.total_after_gst:8.2f} | Excel: {excel_total:8.2f} | Diff: {diff:+8.2f}")
    
    except Exception as e:
        errors += 1
        print(f"Row {row_num:3d}: ERROR - {waybill} | {from_pin}→{to_pin} - {str(e)[:50]}")

print("\n" + "="*100)
print("RESULTS:")
print("="*100)
print(f"✓ Exact matches:     {matches}")
print(f"✗ Mismatches:        {mismatches}")
print(f"⚠ Errors:            {errors}")
print(f"Total validated:     {matches + mismatches}")
print(f"Success rate:        {matches/(matches+mismatches)*100:.1f}%" if (matches+mismatches) > 0 else "N/A")

if mismatches > 0:
    print("\n" + "="*100)
    print("MISMATCH DETAILS:")
    print("="*100)
    for m in mismatch_details:
        print(f"\nRow {m['row']}: {m['waybill']} | {m['from']}→{m['to']} ({m['weight']}kg)")
        print(f"  Calculated: ₹{m['calc']:.2f}")
        print(f"  Excel:      ₹{m['excel']:.2f}")
        print(f"  Difference: {m['diff']:+.2f}")
