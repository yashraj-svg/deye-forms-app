"""
Check for items with blank serial numbers but valid component/specification
"""

import openpyxl

def check_blank_serials():
    file_path = 'Akshay India Shipping List 2020 - 2025.xlsx'
    print(f'\n📂 Checking Excel for blank serial numbers...\n')
    
    wb = openpyxl.load_workbook(file_path)
    
    blank_serial_items = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Extract year
        try:
            year = int(''.join(filter(str.isdigit, sheet_name)))
        except ValueError:
            year = 2025
        
        # Find header row
        header_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            if row and any('PCBA SN' in str(cell) for cell in row if cell):
                header_row = row_idx
                break
        
        if not header_row:
            continue
        
        # Get headers
        headers = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
        col_indices = {}
        for idx, header in enumerate(headers):
            if header:
                header_str = str(header).lower().strip()
                if 'pcba' in header_str and 'old' in header_str:
                    col_indices['pcba_sn_old'] = idx
                elif 'pcba' in header_str and 'new' in header_str:
                    col_indices['pcba_sn_new'] = idx
                elif 'component type' in header_str or 'component' in header_str:
                    col_indices['component_type'] = idx
                elif 'specification' in header_str:
                    col_indices['specification'] = idx
                elif 'quantity' in header_str or 'pcs' in header_str:
                    col_indices['quantity'] = idx
        
        # Check data rows
        for row_num, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
            pcba_sn_new = row[col_indices['pcba_sn_new']] if 'pcba_sn_new' in col_indices else None
            component_type = row[col_indices['component_type']] if 'component_type' in col_indices else None
            specification = row[col_indices['specification']] if 'specification' in col_indices else None
            quantity = row[col_indices['quantity']] if 'quantity' in col_indices else None
            
            # Check if serial is blank but has component/spec info
            if (not pcba_sn_new or str(pcba_sn_new).strip() == '' or str(pcba_sn_new).strip() == 'None'):
                if component_type or specification:
                    blank_serial_items.append({
                        'sheet': sheet_name,
                        'year': year,
                        'row': row_num,
                        'pcba_sn_new': pcba_sn_new,
                        'component_type': component_type,
                        'specification': specification,
                        'quantity': quantity
                    })
    
    print(f'Found {len(blank_serial_items)} items with BLANK serial numbers but valid component/spec info:\n')
    print('='*100)
    
    for idx, item in enumerate(blank_serial_items, 1):
        print(f'\n{idx}. Sheet: {item["sheet"]} (Year: {item["year"]}, Row: {item["row"]})')
        print(f'   Serial: {item["pcba_sn_new"]} (BLANK)')
        print(f'   Component: {item["component_type"]}')
        print(f'   Specification: {item["specification"]}')
        print(f'   Quantity: {item["quantity"]}')
    
    print(f'\n{"="*100}')
    print(f'✓ Total items with blank serial but valid info: {len(blank_serial_items)}')


if __name__ == '__main__':
    check_blank_serials()
