#!/usr/bin/env python
"""Check Safexpress charges with all columns."""
import openpyxl

# Load the file
wb = openpyxl.load_workbook('SFX00017299 DEYE INVERTER TECHNOLOGY PRIVATE LIMITED MIS  (1).xlsx', data_only=True)
ws = wb.active

# Column headers mapping
headers = {
    1: "SFX Code",
    2: "Bill No",
    3: "Waybill No",
    4: "From location",
    5: "To Location",
    6: "Booking date",
    7: "Destination",
    8: "Total Packages",
    9: "Charge Weight",
    10: "BASIC FREIGHT",
    11: "GREEN SURCHARGE BKG",
    12: "GREEN SURCHARGE DLY",
    13: "VALUE SURCHARGE",
    14: "WAYBILL CHARGES",
    15: "HANDLING CHARGES",
    16: "UCC CHARGES",
    17: "BOOKING SAFE EXTENSION",
    18: "DELIVERY SAFE EXTENSION",
    19: "DOD/DACC CHARGES",
    20: "HUB DELIVERY CHARGES",
    21: "POD CHARGES",
    22: "SCHEDULED CHARGES",
    23: "TO-PAY SERVICE CHARGES",
    24: "STATE SURCHARGES",
    25: "INCREMENTAL CHARGES",
    26: "OTHER SERVICE CHARGES",
    27: "FUEL SURCHARGES",
    28: "SDS CHARGES",
    29: "OLD FREIGHT CHARGES",
    30: "Freight",
    31: "GST Amount",
    32: "Total Freight"
}

print("="*100)
print("SAFEXPRESS PRICING BREAKDOWN - First 5 Rows")
print("="*100)

for row_num in range(2, 7):
    waybill = ws.cell(row_num, 3).value
    from_loc = ws.cell(row_num, 4).value
    to_loc = ws.cell(row_num, 5).value
    weight = ws.cell(row_num, 9).value
    
    print(f"\nRow {row_num}: {waybill} | {from_loc}→{to_loc} | Weight: {weight}kg")
    print("-" * 100)
    
    # Basic charges
    basic = ws.cell(row_num, 10).value
    green_bkg = ws.cell(row_num, 11).value
    green_dly = ws.cell(row_num, 12).value
    val_surcharge = ws.cell(row_num, 13).value
    waybill_chg = ws.cell(row_num, 14).value
    handling = ws.cell(row_num, 15).value
    ucc = ws.cell(row_num, 16).value
    booking_safe = ws.cell(row_num, 17).value
    delivery_safe = ws.cell(row_num, 18).value
    dod = ws.cell(row_num, 19).value
    hub_delivery = ws.cell(row_num, 20).value
    pod = ws.cell(row_num, 21).value
    scheduled = ws.cell(row_num, 22).value
    to_pay = ws.cell(row_num, 23).value
    state_surcharge = ws.cell(row_num, 24).value
    incremental = ws.cell(row_num, 25).value
    other = ws.cell(row_num, 26).value
    fuel = ws.cell(row_num, 27).value
    sds = ws.cell(row_num, 28).value
    old_freight = ws.cell(row_num, 29).value
    
    freight_subtotal = ws.cell(row_num, 30).value
    gst = ws.cell(row_num, 31).value
    total = ws.cell(row_num, 32).value
    
    print(f"  Basic Freight:        {basic}")
    print(f"  Green Surcharge BKG:  {green_bkg}")
    print(f"  Green Surcharge DLY:  {green_dly}")
    print(f"  Value Surcharge:      {val_surcharge}")
    print(f"  Waybill Charges:      {waybill_chg}")
    print(f"  Handling Charges:     {handling}")
    print(f"  UCC Charges:          {ucc}")
    print(f"  Booking Safe Ext:     {booking_safe}")
    print(f"  Delivery Safe Ext:    {delivery_safe}")
    print(f"  DOD/DACC Charges:     {dod}")
    print(f"  Hub Delivery Charges: {hub_delivery}")
    print(f"  POD Charges:          {pod}")
    print(f"  Scheduled Charges:    {scheduled}")
    print(f"  To-Pay Service:       {to_pay}")
    print(f"  State Surcharges:     {state_surcharge}")
    print(f"  Incremental Charges:  {incremental}")
    print(f"  Other Service:        {other}")
    print(f"  FUEL SURCHARGES:      {fuel}")
    print(f"  SDS Charges:          {sds}")
    print(f"  Old Freight Charges:  {old_freight}")
    print(f"  " + "-"*95)
    print(f"  Freight Subtotal:     {freight_subtotal}")
    print(f"  GST Amount:           {gst}")
    print(f"  TOTAL FREIGHT:        {total}")
