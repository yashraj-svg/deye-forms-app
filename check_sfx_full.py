#!/usr/bin/env python
"""Analyze Safexpress rates - check all columns and pricing."""
import openpyxl

wb = openpyxl.load_workbook('SFX00017299 DEYE INVERTER TECHNOLOGY PRIVATE LIMITED MIS  (1).xlsx', data_only=True)
ws = wb.active

print("="*100)
print("SAFEXPRESS INVOICE STRUCTURE")
print("="*100)

# Print all headers (row 1)
print("\nHeaders (Row 1):")
headers = {}
for col in range(1, 25):
    val = ws.cell(1, col).value
    if val:
        headers[col] = val
        print(f"  Col {chr(64+col)}: {val}")

print("\n" + "="*100)
print("SAMPLE SHIPMENTS - Full breakdown:")
print("="*100)

# Analyze first 5 shipments
for row_num in range(2, 7):
    data = {}
    for col, header in headers.items():
        data[header] = ws.cell(row_num, col).value
    
    print(f"\nRow {row_num}:")
    print(f"  Waybill: {data.get('Waybill No')}")
    print(f"  Route: {data.get('From location')}→{data.get('To Location')}")
    print(f"  Destination: {data.get('Destination')}")
    print(f"  Weight: {data.get('Charge Weight')}kg")
    print(f"  BASIC FREIGHT: ₹{data.get('BASIC FREIGHT')}")
    print(f"  GREEN BKG: ₹{data.get('GREEN SURCHARGE BKG')}")
    print(f"  GREEN DLY: ₹{data.get('GREEN SURCHARGE DLY')}")
    print(f"  VALUE SURCHARGE: ₹{data.get('VALUE SURCHARGE')}")
    print(f"  WAYBILL CHARGES: ₹{data.get('WAYBILL CHARGES')}")
    print(f"  HANDLING CHARGES: ₹{data.get('HANDLING CHARGES')}")
    
    # Look for more columns (total, tax, etc)
    for col in range(len(headers)+1, 30):
        val = ws.cell(row_num, col).value
        if val is not None:
            header = ws.cell(1, col).value or f"Col{chr(64+col)}"
            print(f"  {header}: {val}")
