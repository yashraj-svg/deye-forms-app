#!/usr/bin/env python
"""Compare Safexpress calculator with Excel values."""
import openpyxl
from forms.calculator.data_loader import load_pincode_master
from forms.calculator.freight_calculator import Safexpress, QuoteInput, ShipmentItem
from forms.calculator.config import Settings

wb = openpyxl.load_workbook('SFX00017299 DEYE INVERTER TECHNOLOGY PRIVATE LIMITED MIS  (1).xlsx', data_only=True)
ws = wb.active

pins = load_pincode_master('.')
calc = Safexpress(Settings())

print("="*120)
print("SAFEXPRESS CALCULATOR COMPARISON - Excel vs Our Calculator")
print("="*120)

for row_num in range(2, 7):
    from_pin = str(ws.cell(row_num, 4).value or '').strip()
    to_pin = str(ws.cell(row_num, 5).value or '').strip()
    weight = ws.cell(row_num, 9).value
    
    # Excel values
    excel_basic = ws.cell(row_num, 10).value or 0
    excel_value_sur = ws.cell(row_num, 13).value or 0
    excel_waybill = ws.cell(row_num, 14).value or 0
    excel_ucc = ws.cell(row_num, 16).value or 0
    excel_safe_ext = ws.cell(row_num, 18).value or 0  # Delivery Safe Extension
    excel_fuel = ws.cell(row_num, 27).value or 0
    excel_subtotal = ws.cell(row_num, 30).value or 0
    excel_gst = ws.cell(row_num, 31).value or 0
    excel_total = ws.cell(row_num, 32).value or 0
    
    # Our calculation
    try:
        items = [ShipmentItem(weight_kg=float(weight), length_cm=0, breadth_cm=0, height_cm=0)]
        result = calc.calculate_quote(QuoteInput(from_pincode=from_pin, to_pincode=to_pin, items=items), pins)
        
        print(f"\nRow {row_num}: {from_pin}→{to_pin}, Weight: {weight}kg")
        print("-" * 120)
        print(f"{'Charge':<30} {'Excel':<15} {'Calculator':<15} {'Match':<10}")
        print("-" * 120)
        
        # Basic freight
        match = "✓" if abs(result.base_freight - excel_basic) < 1 else "✗"
        print(f"{'Basic Freight':<30} {excel_basic:<15.2f} {result.base_freight:<15.2f} {match:<10}")
        
        # Surcharges
        waybill_calc = result.surcharges.get('waybill', 0)
        match = "✓" if abs(waybill_calc - excel_waybill) < 1 else "✗"
        print(f"{'Waybill Charges':<30} {excel_waybill:<15.2f} {waybill_calc:<15.2f} {match:<10}")
        
        value_calc = result.surcharges.get('value_surcharge', 0)
        match = "✓" if abs(value_calc - excel_value_sur) < 1 else "✗"
        print(f"{'Value Surcharge':<30} {excel_value_sur:<15.2f} {value_calc:<15.2f} {match:<10}")
        
        ucc_calc = result.surcharges.get('ucc', 0)
        match = "✓" if abs(ucc_calc - excel_ucc) < 1 else "✗"
        print(f"{'UCC Charges':<30} {excel_ucc:<15.2f} {ucc_calc:<15.2f} {match:<10}")
        
        # Check for Delivery Safe Extension
        if excel_safe_ext > 0:
            print(f"{'Delivery Safe Extension':<30} {excel_safe_ext:<15.2f} {'NOT CALC':<15} {'✗':<10}")
        
        fuel_calc = result.surcharges.get('fuel_surcharge', 0)
        match = "✓" if abs(fuel_calc - excel_fuel) < 5 else "✗"  # Allow 5 tolerance for fuel rounding
        print(f"{'Fuel Surcharges':<30} {excel_fuel:<15.2f} {fuel_calc:<15.2f} {match:<10}")
        
        # Totals
        match = "✓" if abs(result.total_before_gst - excel_subtotal) < 1 else "✗"
        print(f"{'Freight Subtotal':<30} {excel_subtotal:<15.2f} {result.total_before_gst:<15.2f} {match:<10}")
        
        match = "✓" if abs(result.gst_amount - excel_gst) < 1 else "✗"
        print(f"{'GST Amount':<30} {excel_gst:<15.2f} {result.gst_amount:<15.2f} {match:<10}")
        
        match = "✓" if abs(result.total_after_gst - excel_total) < 1 else "✗"
        print(f"{'TOTAL FREIGHT':<30} {excel_total:<15.2f} {result.total_after_gst:<15.2f} {match:<10}")
        
        # Show all surcharges calculated
        print(f"\nAll surcharges calculated: {result.surcharges}")
        
    except Exception as e:
        print(f"\nRow {row_num}: ERROR - {e}")

print("\n" + "="*120)
