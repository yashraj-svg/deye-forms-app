import openpyxl

# Load the Excel file
wb = openpyxl.load_workbook('Current stock updated.xlsx', data_only=True)

print("=" * 80)
print("EXCEL FILE ANALYSIS: Current stock updated.xlsx")
print("=" * 80)

print(f"\nSheet names: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*80}")
    print(f"Total rows: {ws.max_row}")
    print(f"Total columns: {ws.max_column}")
    print(f"\nFirst 10 rows:")
    print("-" * 80)
    
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        print(f"Row {i}: {row}")
    
    print("\n")

wb.close()
