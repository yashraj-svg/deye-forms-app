#!/usr/bin/env python
"""Analyze Safexpress rates from Excel invoice."""
import openpyxl

wb = openpyxl.load_workbook('SFX00017299 DEYE INVERTER TECHNOLOGY PRIVATE LIMITED MIS  (1).xlsx', data_only=True)

# List all sheets
print("Available sheets:")
for sheet in wb.sheetnames:
    print(f"  - {sheet}")

# Try to find data in first sheet
ws = wb.active
print(f"\nActive sheet: {ws.title}")
print("\n" + "="*80)
print("SHEET STRUCTURE:")
print("="*80)

# Print first 20 rows and 15 columns to understand structure
for row in range(1, 25):
    row_data = []
    for col in range(1, 16):
        val = ws.cell(row, col).value
        if val is not None:
            row_data.append(f"Col{chr(64+col)}={val}")
    if row_data:
        print(f"Row {row}: {', '.join(row_data[:8])}")
        if len(row_data) > 8:
            print(f"         {', '.join(row_data[8:])}")
