#!/usr/bin/env python3
import openpyxl

wb = openpyxl.load_workbook('SFX00017299 DEYE INVERTER TECHNOLOGY PRIVATE LIMITED MIS  (1).xlsx')
ws = wb.active

# Check header row to identify columns
print("HEADER ROW:")
for col in range(1, 33):
    val = ws.cell(1, col).value
    print(f"  Col {col}: {val}")

print("\n" + "="*100)
print("MISMATCHED ROWS DETAIL:")
rows_to_check = [9, 14, 23]
for row_num in rows_to_check:
    print(f"\nRow {row_num}: {ws.cell(row_num, 4).value} -> {ws.cell(row_num, 5).value}  Weight: {ws.cell(row_num, 9).value}kg")
    print("-" * 100)
    print("All charges:")
    for col in range(10, 32):
        val = ws.cell(row_num, col).value
        header = ws.cell(1, col).value
        if val:
            print(f"  Col {col:2d} ({header}): Rs.{val}")
    print(f"  Col 32 (TOTAL FREIGHT): Rs.{ws.cell(row_num, 32).value}")
