#!/usr/bin/env python
"""Find all rate mismatches between Excel and our matrix."""
import openpyxl
from forms.calculator.data_loader import load_pincode_master
from forms.calculator.freight_calculator import GlobalCourierCargo
from forms.calculator.config import Settings

wb = openpyxl.load_workbook('ML model for global courier.xlsx', data_only=True)
ws = wb.active

pins = load_pincode_master('.')
calc = GlobalCourierCargo(Settings())

rate_mismatches = {}

print("Checking all 74 shipments for rate discrepancies...\n")

for row_num in range(2, 76):  # Rows 2-75
    from_pin = str(ws.cell(row_num, 5).value or '').strip()
    to_pin = str(ws.cell(row_num, 6).value or '').strip()
    weight = ws.cell(row_num, 10).value
    excel_base = ws.cell(row_num, 12).value
    
    if not from_pin or not to_pin or not weight or not excel_base:
        continue
    
    fp = pins.get(from_pin)
    tp = pins.get(to_pin)
    
    if not fp or not tp:
        continue
    
    from_zone = fp.global_cargo_region or 'Unknown'
    to_zone = tp.global_cargo_region or 'Unknown'
    
    # Calculate rates (excluding ODA from base)
    excel_rate = excel_base / weight if weight > 0 else 0
    our_rate = calc.base_rate_per_kg(from_zone, to_zone, weight)
    
    # Check if there's a mismatch (allow 0.1 tolerance)
    if abs(excel_rate - our_rate) > 0.1:
        zone_pair = f"{from_zone}→{to_zone}"
        if zone_pair not in rate_mismatches:
            rate_mismatches[zone_pair] = {
                'excel_rate': excel_rate,
                'our_rate': our_rate,
                'examples': []
            }
        rate_mismatches[zone_pair]['examples'].append(f"{from_pin}→{to_pin} ({weight}kg)")

print("="*80)
print(f"Found {len(rate_mismatches)} zone pairs with rate discrepancies:")
print("="*80)

for zone_pair, data in sorted(rate_mismatches.items()):
    print(f"\n{zone_pair}:")
    print(f"  Excel rate: {data['excel_rate']:.1f}/kg")
    print(f"  Our rate:   {data['our_rate']:.1f}/kg")
    print(f"  Difference: {data['our_rate'] - data['excel_rate']:+.1f}/kg")
    print(f"  Examples: {', '.join(data['examples'][:3])}")
    if len(data['examples']) > 3:
        print(f"            ... and {len(data['examples'])-3} more")
