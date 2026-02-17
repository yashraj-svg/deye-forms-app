#!/usr/bin/env python
"""Extract Global Cargo rates from Excel file"""
import openpyxl
import json

excel_file = 'Global Cargo Rates  Reupdated (14-02-2026).xlsx'
wb = openpyxl.load_workbook(excel_file)

# Extract ZONE data
print("=" * 80)
print("ZONE DEFINITIONS")
print("=" * 80)
ws_zone = wb['ZONE']
zones = {}
for row_idx in range(2, ws_zone.max_row + 1):
    code = ws_zone.cell(row_idx, 2).value
    details = ws_zone.cell(row_idx, 3).value
    zone_name = ws_zone.cell(row_idx, 4).value
    if code:
        print(f"{code}: {details} -> {zone_name}")
        zones[code] = zone_name

# Extract RATE matrix
print("\n" + "=" * 80)
print("RATE MATRIX (Rs/kg)")
print("=" * 80)
ws_rate = wb['RATE']
rate_matrix = {}

# Get column headers (zones)
columns = []
for col_idx in range(2, ws_rate.max_column + 1):
    col_val = ws_rate.cell(1, col_idx).value
    if col_val:
        columns.append(col_val)
        print(f"Col {col_idx}: {col_val}")

print("\nRate Matrix:")
for row_idx in range(2, ws_rate.max_row + 1):
    from_zone = ws_rate.cell(row_idx, 1).value
    if from_zone:
        rate_row = {}
        for col_idx, to_zone in enumerate(columns, start=2):
            rate = ws_rate.cell(row_idx, col_idx).value
            if rate:
                rate_row[to_zone] = float(rate)
        if rate_row:
            rate_matrix[from_zone] = rate_row
            print(f"{from_zone}: {rate_row}")

# Extract VAS (Value Added Services)
print("\n" + "=" * 80)
print("VALUE ADDED SERVICES & CHARGES")
print("=" * 80)
ws_vas = wb['VAS']
for row_idx in range(2, ws_vas.max_row + 1):
    col1 = ws_vas.cell(row_idx, 1).value
    col3 = ws_vas.cell(row_idx, 3).value
    col4 = ws_vas.cell(row_idx, 4).value
    col5 = ws_vas.cell(row_idx, 5).value
    if col3 or col4:
        print(f"{col3} | {col4}: {col5}")

print("\n" + "=" * 80)
print("SUMMARY")
print("=" * 80)
print(f"Zones: {len(rate_matrix)} zones")
print(f"Rate Matrix Keys: {list(rate_matrix.keys())}")
