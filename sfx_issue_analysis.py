#!/usr/bin/env python
"""Detailed Safexpress rate analysis to identify the issue."""
import openpyxl
from forms.calculator.data_loader import load_pincode_master
from forms.calculator.freight_calculator import Safexpress, QuoteInput, ShipmentItem
from forms.calculator.config import Settings

wb = openpyxl.load_workbook('SFX00017299 DEYE INVERTER TECHNOLOGY PRIVATE LIMITED MIS  (1).xlsx', data_only=True)
ws = wb.active

pins = load_pincode_master('.')
sfx = Safexpress(Settings())

print("SAFEXPRESS ISSUE ANALYSIS")
print("="*90)

# Analyze a few specific cases
test_rows = [3, 4, 10, 14, 17]  # Rows with different issues

for row_num in test_rows:
    from_pin = str(ws.cell(row_num, 4).value or '').strip()
    to_pin = str(ws.cell(row_num, 5).value or '').strip()
    weight = ws.cell(row_num, 9).value
    excel_basic = ws.cell(row_num, 10).value
    
    fp = pins.get(from_pin)
    tp = pins.get(to_pin)
    
    print(f"\nRow {row_num}: {from_pin} to {to_pin} ({weight}kg)")
    print(f"  States: {fp.state if fp else 'N/A'} to {tp.state if tp else 'N/A'}")
    
    # Get region mapping
    from_region = sfx._region_for_state(fp.state if fp else "")
    to_region = sfx._region_for_state(tp.state if tp else "")
    band = sfx.ZONE_MATRIX.get(from_region, {}).get(to_region, "C")
    
    print(f"  Routing: {from_region} to {to_region}, Band: {band}")
    
    # Get rate
    rate = sfx.ZONE_RATE_A_TO_E.get(band, 10.0)
    print(f"  Rate per kg for band {band}: {rate}/kg")
    
    # Calculate chargeable weight with slab
    items = [ShipmentItem(weight_kg=float(weight), length_cm=0, breadth_cm=0, height_cm=0)]
    cw = sfx.chargeable_weight(items)
    print(f"  Weight slab applied: {weight}kg -> {cw}kg")
    
    # Calculate basic freight
    our_basic = rate * cw
    print(f"  Our calculation: {rate} * {cw} = {our_basic}")
    print(f"  Excel basic freight: {excel_basic}")
    
    # Check if Excel uses different rate
    if excel_basic and weight > 0:
        excel_rate = excel_basic / weight
        print(f"  Excel implied rate: {excel_basic} / {weight} = {excel_rate}/kg")
    
    print()

print("="*90)
print("POSSIBLE ISSUES:")
print("1. Band/Zone mapping might be different in Excel")
print("2. Weight slab logic might be different")
print("3. Base rate values might be different")
print("4. Fuel surcharge might not apply to basic rate in Excel (show 0)")
