#!/usr/bin/env python
"""Check weight slab logic impact."""
import openpyxl

wb = openpyxl.load_workbook('SFX00017299 DEYE INVERTER TECHNOLOGY PRIVATE LIMITED MIS  (1).xlsx', data_only=True)
ws = wb.active

print("Weight Analysis from Excel:")
print("="*80)
print(f"{'Row':<5} {'Weight':<10} {'Basic':<10} {'Rate':<10} {'Expected (rate*weight)':<20} {'Match'}")
print("="*80)

for row_num in range(2, 28):
    weight = ws.cell(row_num, 9).value
    basic = ws.cell(row_num, 10).value
    
    if not weight or not basic:
        continue
    
    # Check all possible rates
    calc_rate = basic / weight if weight > 0 else 0
    
    expected_notaion = f"{basic} = {weight}kg × {calc_rate:.0f}/kg"
    match = "✓"
    
    # If weight should round to next slab...
    WEIGHT_SLABS = [10, 20, 30, 40, 50, 60, 70, 80, 90, 100, 110, 120, 130, 140, 150, 200, 250, 300, 500, 1000]
    slab_weight = weight
    for s in WEIGHT_SLABS:
        if weight <= s:
            slab_weight = s
            break
    
    if slab_weight != weight:
        match = f"Slab: {slab_weight}kg"
    
    print(f"{row_num:<5} {weight:<10} {basic:<10.0f} {calc_rate:<10.0f} {expected_notaion:<20} {match}")
