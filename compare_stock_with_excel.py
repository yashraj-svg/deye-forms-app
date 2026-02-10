"""
Compare database stock items with Excel file to find mismatches
"""

import os
import django
import openpyxl

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'deye_config.settings')
django.setup()

from forms.models import StockItem
from collections import defaultdict

def load_excel_data(file_path='Akshay India Shipping List 2020 - 2025.xlsx'):
    """Load all data from Excel file"""
    print(f'\n📂 Loading Excel data from: {file_path}')
    
    wb = openpyxl.load_workbook(file_path)
    excel_data = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Extract year
        try:
            year = int(''.join(filter(str.isdigit, sheet_name)))
        except ValueError:
            year = 2025
        
        # Find header row
        header_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            if row and any('PCBA SN' in str(cell) for cell in row if cell):
                header_row = row_idx
                break
        
        if not header_row:
            continue
        
        # Get headers
        headers = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
        col_indices = {}
        for idx, header in enumerate(headers):
            if header:
                header_str = str(header).lower().strip()
                if 'pcba' in header_str and 'old' in header_str:
                    col_indices['pcba_sn_old'] = idx
                elif 'pcba' in header_str and 'new' in header_str:
                    col_indices['pcba_sn_new'] = idx
                elif 'component type' in header_str or 'component' in header_str:
                    col_indices['component_type'] = idx
                elif 'specification' in header_str:
                    col_indices['specification'] = idx
                elif 'quantity' in header_str or 'pcs' in header_str:
                    col_indices['quantity'] = idx
        
        # Read data rows
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            pcba_sn_new = row[col_indices['pcba_sn_new']] if 'pcba_sn_new' in col_indices else None
            
            # Convert to string and clean
            if pcba_sn_new:
                pcba_sn_new = str(pcba_sn_new).strip()
                if '.' in pcba_sn_new and 'e' not in pcba_sn_new.lower():
                    pcba_sn_new = pcba_sn_new.split('.')[0]
                if pcba_sn_new == '' or pcba_sn_new.lower() == 'none':
                    pcba_sn_new = None
            else:
                pcba_sn_new = None
            
            component_type = row[col_indices['component_type']] if 'component_type' in col_indices else None
            if component_type:
                component_type = str(component_type).strip()
                if component_type == '' or component_type == 'None':
                    component_type = None
            
            specification = row[col_indices['specification']] if 'specification' in col_indices else None
            if specification:
                specification = str(specification).strip()
                if specification == '' or specification == 'None':
                    specification = None
            
            quantity_raw = row[col_indices['quantity']] if 'quantity' in col_indices else None
            try:
                quantity = float(quantity_raw) if quantity_raw else 0
            except (ValueError, TypeError):
                quantity = 0
            
            # Skip only if both serial AND (component_type or specification) are missing
            # Allow items with blank serial if they have component/spec info
            if not pcba_sn_new and not component_type and not specification:
                continue
            
            excel_data.append({
                'pcba_sn_new': pcba_sn_new,
                'component_type': component_type,
                'specification': specification,
                'quantity': quantity,
                'year': year
            })
    
    print(f'✓ Loaded {len(excel_data)} items from Excel')
    return excel_data


def compare_data():
    """Compare Excel data with database"""
    print('\n' + '='*80)
    print('STOCK COMPARISON REPORT: Excel vs Database')
    print('='*80)
    
    # Load Excel data
    excel_data = load_excel_data()
    
    # Load database data
    db_items = StockItem.objects.all()
    print(f'📊 Database has {db_items.count()} items')
    
    # Index by serial number
    excel_by_sn = defaultdict(list)
    for item in excel_data:
        excel_by_sn[item['pcba_sn_new']].append(item)
    
    db_by_sn = defaultdict(list)
    for item in db_items:
        db_by_sn[item.pcba_sn_new].append(item)
    
    # Find differences
    missing_in_db = []
    missing_in_excel = []
    quantity_mismatches = []
    description_mismatches = []
    
    # Check Excel items against DB
    for sn, excel_items in excel_by_sn.items():
        if sn not in db_by_sn:
            missing_in_db.extend(excel_items)
        else:
            # Compare quantities and descriptions
            excel_total_qty = sum(item['quantity'] for item in excel_items)
            db_total_qty = sum(float(item.quantity) for item in db_by_sn[sn])
            
            if abs(excel_total_qty - db_total_qty) > 0.01:
                quantity_mismatches.append({
                    'serial': sn,
                    'excel_qty': excel_total_qty,
                    'db_qty': db_total_qty,
                    'difference': excel_total_qty - db_total_qty,
                    'excel_items': excel_items,
                    'db_items': db_by_sn[sn]
                })
            
            # Check descriptions
            for excel_item in excel_items:
                matching_db = False
                for db_item in db_by_sn[sn]:
                    if (excel_item['component_type'] == db_item.component_type and 
                        excel_item['specification'] == db_item.specification):
                        matching_db = True
                        break
                
                if not matching_db:
                    description_mismatches.append({
                        'serial': sn,
                        'excel_type': excel_item['component_type'],
                        'excel_spec': excel_item['specification'],
                        'db_items': [(i.component_type, i.specification) for i in db_by_sn[sn]]
                    })
    
    # Check DB items against Excel
    for sn, db_items_list in db_by_sn.items():
        if sn not in excel_by_sn:
            missing_in_excel.extend(db_items_list)
    
    # Print report
    print('\n' + '-'*80)
    print(f'📋 SUMMARY')
    print('-'*80)
    print(f'Total Excel items: {len(excel_data)}')
    print(f'Total DB items: {db_items.count()}')
    print(f'Unique serial numbers in Excel: {len(excel_by_sn)}')
    print(f'Unique serial numbers in DB: {len(db_by_sn)}')
    
    if missing_in_db:
        print(f'\n⚠️  MISSING IN DATABASE: {len(missing_in_db)} items')
        print('-'*80)
        for item in missing_in_db[:10]:  # Show first 10
            print(f'  Serial: {item["pcba_sn_new"]}')
            print(f'  Component: {item["component_type"]}')
            print(f'  Specification: {item["specification"]}')
            print(f'  Quantity: {item["quantity"]}')
            print(f'  Year: {item["year"]}')
            print()
        if len(missing_in_db) > 10:
            print(f'  ... and {len(missing_in_db) - 10} more')
    
    if missing_in_excel:
        print(f'\n⚠️  EXTRA IN DATABASE (not in Excel): {len(missing_in_excel)} items')
        print('-'*80)
        for item in missing_in_excel[:10]:
            print(f'  Serial: {item.pcba_sn_new}')
            print(f'  Component: {item.component_type}')
            print(f'  Specification: {item.specification}')
            print(f'  Quantity: {item.quantity}')
            print(f'  Year: {item.year}')
            print()
        if len(missing_in_excel) > 10:
            print(f'  ... and {len(missing_in_excel) - 10} more')
    
    if quantity_mismatches:
        print(f'\n⚠️  QUANTITY MISMATCHES: {len(quantity_mismatches)} items')
        print('-'*80)
        for item in quantity_mismatches[:10]:
            print(f'  Serial: {item["serial"]}')
            print(f'  Excel Quantity: {item["excel_qty"]}')
            print(f'  Database Quantity: {item["db_qty"]}')
            print(f'  Difference: {item["difference"]:+.2f}')
            print()
        if len(quantity_mismatches) > 10:
            print(f'  ... and {len(quantity_mismatches) - 10} more')
    
    if description_mismatches:
        print(f'\n⚠️  DESCRIPTION MISMATCHES: {len(description_mismatches)} items')
        print('-'*80)
        for item in description_mismatches[:10]:
            print(f'  Serial: {item["serial"]}')
            print(f'  Excel Type: {item["excel_type"]}')
            print(f'  Excel Spec: {item["excel_spec"]}')
            print(f'  DB Types/Specs: {item["db_items"]}')
            print()
        if len(description_mismatches) > 10:
            print(f'  ... and {len(description_mismatches) - 10} more')
    
    if not missing_in_db and not missing_in_excel and not quantity_mismatches and not description_mismatches:
        print('\n✅ PERFECT MATCH! All data matches between Excel and Database!')
    
    print('\n' + '='*80)
    print('END OF REPORT')
    print('='*80)


if __name__ == '__main__':
    compare_data()
