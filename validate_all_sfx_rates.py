#!/usr/bin/env python3
"""
Comprehensive Safexpress Rate Validation Script
Validates all rows from Excel against our calculator to ensure MOU compliance
"""

import openpyxl
import sys
from pathlib import Path
from dataclasses import dataclass

# Add forms to path
sys.path.insert(0, str(Path(__file__).parent))

from forms.calculator.data_loader import PincodeDB, load_pincode_master
from forms.calculator.freight_calculator import (
    Safexpress, QuoteInput, ShipmentItem, _round_money, Settings
)

# Load pincode database properly
db = load_pincode_master('.')

# Excel file path
EXCEL_FILE = Path(__file__).parent / "SFX00017299 DEYE INVERTER TECHNOLOGY PRIVATE LIMITED MIS  (1).xlsx"

@dataclass
class RowValidation:
    row_num: int
    from_pin: str
    to_pin: str
    weight: float
    excel_total: float
    calc_total: float
    match: bool
    difference: float
    error: str = ""

def validate_all_rows():
    """Validate all rows from Excel file"""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        results = []
        print("\n" + "="*120)
        print("SAFEXPRESS RATE VALIDATION - ALL ROWS")
        print("="*120 + "\n")
        
        # Find header row
        header_row = 1
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=False), 1):
            first_cell = row[0].value
            if first_cell and "SFX" in str(first_cell):
                header_row = row_idx
                break
        
        print(f"Header row detected at row {header_row}\n")
        
        # Column indices (0-based, but we use 1-based when reading from Excel)
        # A=1: SFX Code, D=4: From location, E=5: To Location, I=9: Charge Weight, AF=32: Total Freight
        # Using openpyxl, we read with cell(row, col) where col is 1-based
        
        total_rows = 0
        matched_rows = 0
        failed_rows = 0
        
        # Iterate through all data rows
        for row_idx in range(header_row + 1, ws.max_row + 1):
            from_pin_cell = ws.cell(row_idx, 4).value  # Column D
            to_pin_cell = ws.cell(row_idx, 5).value    # Column E
            weight_cell = ws.cell(row_idx, 9).value    # Column I
            total_cell = ws.cell(row_idx, 32).value    # Column AF
            
            if from_pin_cell is None:
                break
            
            
            total_rows += 1
            from_pin_str = str(from_pin_cell or "").strip()
            to_pin_str = str(to_pin_cell or "").strip()
            weight = float(weight_cell or 0)
            excel_total = float(total_cell or 0)
            
            if not from_pin_str or not to_pin_str:
                continue
            
            # Create quote input
            try:
                inp = QuoteInput(
                    from_pincode=from_pin_str,
                    to_pincode=to_pin_str,
                    items=[ShipmentItem(weight_kg=weight, length_cm=0, breadth_cm=0, height_cm=0)],
                    insured_value=None
                )
                
                # Calculate
                calc = Safexpress(Settings())
                result = calc.calculate_quote(inp, db)
                calc_total = result.total_after_gst
                
                # Check match (within Rs2 tolerance for rounding)
                tolerance = 2.0
                match = abs(excel_total - calc_total) <= tolerance
                difference = excel_total - calc_total
                
                if match:
                    matched_rows += 1
                    status = "OK"
                else:
                    failed_rows += 1
                    status = "NO"
                
                validation = RowValidation(
                    row_num=row_idx,
                    from_pin=from_pin_str,
                    to_pin=to_pin_str,
                    weight=weight,
                    excel_total=excel_total,
                    calc_total=calc_total,
                    match=match,
                    difference=difference
                )
                results.append(validation)
                
                # Print row result
                print(f"Row {row_idx:3d} {status}  {from_pin_str}->{to_pin_str}  Weight: {weight:6.1f}kg  "
                      f"Excel: Rs.{excel_total:8.2f}  Calc: Rs.{calc_total:8.2f}  Diff: Rs.{difference:7.2f}")
                
            except Exception as e:
                failed_rows += 1
                print(f"Row {row_idx:3d} X  {from_pin_str}->{to_pin_str}  ERROR: {str(e)[:60]}")
                validation = RowValidation(
                    row_num=row_idx,
                    from_pin=from_pin_str,
                    to_pin=to_pin_str,
                    weight=weight,
                    excel_total=excel_total,
                    calc_total=0,
                    match=False,
                    difference=0,
                    error=str(e)
                )
                results.append(validation)
        
        # Summary
        print("\n" + "="*120)
        print("VALIDATION SUMMARY")
        print("="*120)
        print(f"Total Rows Processed: {total_rows}")
        print(f"Matched Rows (within ₹2): {matched_rows} ({100*matched_rows/max(total_rows, 1):.1f}%)")
        print(f"Failed/Mismatched Rows: {failed_rows} ({100*failed_rows/max(total_rows, 1):.1f}%)")
        print("="*120 + "\n")
        
        # Show failed rows details
        if failed_rows > 0:
            print("MISMATCHED ROWS DETAILS:")
            print("-"*120)
            for v in results:
                if not v.match:
                    print(f"Row {v.row_num}: {v.from_pin}->{v.to_pin} Weight: {v.weight}kg")
                    print(f"  Excel: Rs.{v.excel_total:.2f} | Calculator: Rs.{v.calc_total:.2f} | Diff: Rs.{v.difference:.2f}")
                    if v.error:
                        print(f"  Error: {v.error}")
            print("-"*120 + "\n")
        
        return matched_rows, failed_rows, total_rows
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()
        return 0, 0, 0

if __name__ == "__main__":
    matched, failed, total = validate_all_rows()
    
    if failed == 0 and total > 0:
        print("[SUCCESS] ALL RATES VALIDATED SUCCESSFULLY!")
        sys.exit(0)
    else:
        print(f"[WARNING] {failed} rows need attention")
        sys.exit(1)
