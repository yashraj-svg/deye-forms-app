#!/usr/bin/env python
"""Compare Safexpress calculator against Excel invoice prices."""
import openpyxl
from forms.calculator.data_loader import load_pincode_master
from forms.calculator.freight_calculator import Safexpress, QuoteInput, ShipmentItem
from forms.calculator.config import Settings

wb = openpyxl.load_workbook('SFX00017299 DEYE INVERTER TECHNOLOGY PRIVATE LIMITED MIS  (1).xlsx', data_only=True)
ws = wb.active

pins = load_pincode_master('.')
sfx = Safexpress(Settings())

print("="*100)
print("SAFEXPRESS RATE VALIDATION")
print("="*100)

# Columns from Excel
HEADERS = {
    'Waybill': 'C',
    'From': 'D',
    'To': 'E',
    'Weight': 'I',
    'Basic': 'J',
    'Value': 'M',
    'Waybill': 'N',
    'Handling': 'O',
    'Fuel': 'V'  # Fuel surcharges column
}

matches = 0
mismatches = 0

for row_num in range(2, 50):  # Check first 48 shipments
    from_pin = str(ws.cell(row_num, 4).value or '').strip()  # Col D
    to_pin = str(ws.cell(row_num, 5).value or '').strip()    # Col E
    weight = ws.cell(row_num, 9).value                       # Col I
    
    excel_basic = ws.cell(row_num, 10).value                 # Col J
    excel_value = ws.cell(row_num, 13).value                 # Col M
    excel_waybill = ws.cell(row_num, 14).value               # Col N
    excel_handling = ws.cell(row_num, 15).value              # Col O
    excel_fuel = ws.cell(row_num, 22).value                  # Col V (FUEL SURCHARGES)
    
    # Skip if missing data
    if not from_pin or not to_pin or not weight or not excel_basic:
        continue
    
    # Calculate using our logic
    items = [ShipmentItem(weight_kg=float(weight), length_cm=0, breadth_cm=0, height_cm=0)]
    result = sfx.calculate_quote(QuoteInput(from_pincode=from_pin, to_pincode=to_pin, items=items), pins)
    
    # Extract our calculated values
    our_basic = result.base_freight
    our_waybill = result.surcharges.get('waybill', 0)
    our_value = result.surcharges.get('value_surcharge', 0)
    our_fuel = result.surcharges.get('fuel_surcharge', 0)
    our_total_before_tax = result.total_before_gst
    
    # Compare
    basic_match = abs(our_basic - (excel_basic or 0)) < 1
    waybill_match = abs(our_waybill - (excel_waybill or 0)) < 1
    value_match = abs(our_value - (excel_value or 0)) < 1
    fuel_match = abs(our_fuel - (excel_fuel or 0)) < 1
    
    if basic_match and waybill_match and value_match and fuel_match:
        matches += 1
    else:
        mismatches += 1
        print(f"\nRow {row_num}: {from_pin} to {to_pin} ({weight}kg)")
        print(f"  From: {from_pin}, To: {to_pin}")
        fp = pins.get(from_pin)
        tp = pins.get(to_pin)
        if fp and tp:
            print(f"  States: {fp.state} to {tp.state}")
        print(f"  {'Charge':<15} {'Excel':<12} {'Our Calc':<12} {'Match':<7}")
        print(f"  {'-'*45}")
        print(f"  {'Basic Freight':<15} {excel_basic:<11} {our_basic:<11.2f} {'OK' if basic_match else 'X':<7}")
        print(f"  {'Value Surcharge':<15} {excel_value:<11} {our_value:<11.2f} {'OK' if value_match else 'X':<7}")
        print(f"  {'Waybill Charge':<15} {excel_waybill:<11} {our_waybill:<11.2f} {'OK' if waybill_match else 'X':<7}")
        print(f"  {'Fuel Surcharge':<15} {excel_fuel or 0:<11} {our_fuel:<11.2f} {'OK' if fuel_match else 'X':<7}")

print("\n" + "="*100)
print(f"RESULTS: {matches} matches, {mismatches} mismatches (from {matches+mismatches} shipments)")
print("="*100)
