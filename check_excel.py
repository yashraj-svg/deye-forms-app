import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('Akshay India Shipping List 2020 - 2025.xlsx')
print('Sheet names:', wb.sheetnames)
ws = wb.active
print(f'Total rows: {ws.max_row}')
print(f'Total columns: {ws.max_column}')

print('\nFirst 10 rows:')
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    print(f'Row {i}: {row}')
    if i >= 10:
        break

print('\nColumn headers (Row 1):')
header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
for idx, header in enumerate(header_row, 1):
    print(f'Column {idx}: {header}')
