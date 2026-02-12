import openpyxl
import os

# Find the Excel file
excel_files = [f for f in os.listdir('.') if f.endswith('.xlsx')]
print(f"Excel files found: {excel_files}")

if excel_files:
    wb = openpyxl.load_workbook(excel_files[0])
    print(f"\nSheet names: {wb.sheetnames}")
    
    # Check if 'Current Stock' sheet exists
    if 'Current Stock' in wb.sheetnames:
        ws = wb['Current Stock']
        print(f"\n✓ Current Stock sheet found!")
        print(f"Dimensions: {ws.dimensions}")
        print(f"\nFirst 5 rows:")
        for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True), 1):
            print(f"Row {i}: {row}")
        
        # Count total rows with data
        max_row = ws.max_row
        print(f"\nTotal rows: {max_row}")
        
        # Get total quantity
        total_qty = 0
        for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=10, values_only=True):
            print(f"Row sample: {row}")
            break
    else:
        print("\n✗ 'Current Stock' sheet not found!")
else:
    print("No Excel files found!")
