#!/usr/bin/env python
"""Find rate mismatches excluding minimum charge and ODA cases."""
import openpyxl
from forms.calculator.data_loader import load_pincode_master
from forms.calculator.freight_calculator import GlobalCourierCargo
from forms.calculator.config import Settings

wb = openpyxl.load_workbook('ML model for global courier.xlsx', data_only=True)
ws = wb.active

pins = load_pincode_master('.')
calc = GlobalCourierCargo(Settings())

rate_mismatches = {}

print("Analyzing 74 shipments (excluding minimum charge and ODA cases)...\n")

for row_num in range(2, 76):
    from_pin = str(ws.cell(row_num, 5).value or '').strip()
    to_pin = str(ws.cell(row_num, 6).value or '').strip()
    weight = ws.cell(row_num, 10).value
    excel_base = ws.cell(row_num, 12).value
    
    if not from_pin or not to_pin or not weight or not excel_base:
        continue
    
    fp = pins.get(from_pin)
    tp = pins.get(to_pin)
    
    if not fp or not tp:
        continue
    
    from_zone = fp.global_cargo_region or 'Unknown'
    to_zone = tp.global_cargo_region or 'Unknown'
    
    # Skip if Excel base is exactly minimum (450) or minimum+ODA (1050)
    # These don't reveal the actual rate
    if excel_base in [450, 1050]:
        continue
    
    # Calculate rates
    excel_rate = excel_base / weight if weight > 0 else 0
    our_rate = calc.base_rate_per_kg(from_zone, to_zone, weight)
    
    # Check if there's a mismatch (allow 0.5 tolerance for rounding)
    if abs(excel_rate - our_rate) > 0.5:
        zone_pair = f"{from_zone}→{to_zone}"
        if zone_pair not in rate_mismatches:
            rate_mismatches[zone_pair] = {
                'excel_rate': excel_rate,
                'our_rate': our_rate,
                'examples': []
            }
        rate_mismatches[zone_pair]['examples'].append(
            f"Row {row_num}: {from_pin}→{to_pin} ({weight}kg, base={excel_base})"
        )

print("="*80)
print(f"Found {len(rate_mismatches)} zone pairs with genuine rate discrepancies:")
print("="*80)

for zone_pair, data in sorted(rate_mismatches.items()):
    print(f"\n{zone_pair}:")
    print(f"  Excel rate: {data['excel_rate']:.2f}/kg")
    print(f"  Our rate:   {data['our_rate']:.2f}/kg")
    print(f"  Difference: {data['our_rate'] - data['excel_rate']:+.2f}/kg")
    for ex in data['examples'][:5]:
        print(f"  - {ex}")
    if len(data['examples']) > 5:
        print(f"  ... and {len(data['examples'])-5} more")

if not rate_mismatches:
    print("\n✓ No genuine rate discrepancies found!")
    print("All mismatches are due to minimum charge (₹450) or ODA reclassification.")
